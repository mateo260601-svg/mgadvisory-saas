from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

from app.config import OUTPUTS_DIR


def main() -> None:
    workbooks = sorted(Path(OUTPUTS_DIR).glob("*_BP_Model.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not workbooks:
        raise SystemExit("No BP workbook found in outputs.")

    path = workbooks[0]
    wb = load_workbook(path, data_only=False)
    required = [
        "Dashboard",
        "Config",
        "Assumptions",
        "Historicals",
        "Revenue",
        "Costs Payroll",
        "Debt",
        "3FS",
        "Checks",
    ]
    missing = [name for name in required if name not in wb.sheetnames]
    if missing:
        raise SystemExit(f"Missing sheets: {missing}")
    if len(wb.sheetnames) > 10:
        raise SystemExit(f"Too many sheets for configurable BP export: {len(wb.sheetnames)}")

    formulas = 0
    inputs = 0
    validations = 0
    protected = []
    for ws in wb.worksheets:
        if ws.protection.sheet:
            protected.append(ws.title)
        validations += len(ws.data_validations.dataValidation)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == "00D9EAF7":
                    inputs += 1

    if protected:
        raise SystemExit(f"Workbook should be editable by default, but protected sheets were found: {protected}")
    if formulas < 1000:
        raise SystemExit(f"Formula count too low: {formulas}")
    if validations < 6:
        raise SystemExit(f"Validation count too low: {validations}")
    if wb["Config"]["B16"].value != 250000:
        raise SystemExit("Saved BP Builder opening cash was not written to Config.")
    if wb["Assumptions"]["A5"].value != "Smoke revenue stream":
        raise SystemExit("Saved BP Builder revenue stream was not written to Assumptions.")
    if wb["Debt"]["E4"].value != 72:
        raise SystemExit("Saved BP Builder debt term was not written to Debt.")
    print(f"OK {path.name} sheets={len(wb.sheetnames)} formulas={formulas} input_cells={inputs} validations={validations}")


if __name__ == "__main__":
    main()
