from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUTS = ROOT / "outputs"


def find_bolt_template() -> Path | None:
    candidates = [
        Path(r"C:\Users\MateoGirard\JRC CORPORATE CONSULTING\JRC CORP CONSUL - JRC CLIENTS\JRC\Template & Example\Project Bolt - Financial Model - vF.xlsx"),
        Path(r"C:\Users\MateoGirard\OneDrive - JRC CORPORATE CONSULTING\JRC CORP CONSUL - JRC CLIENTS\JRC\Template & Example\Project Bolt - Financial Model - vF.xlsx"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    for base in [Path(r"C:\Users\MateoGirard\JRC CORPORATE CONSULTING"), Path(r"C:\Users\MateoGirard\OneDrive - JRC CORPORATE CONSULTING")]:
        if not base.exists():
            continue
        matches = list(base.rglob("*Project Bolt*Financial Model*vF.xlsx"))
        if matches:
            return matches[0]
    return None


def latest_generated_model() -> Path:
    models = sorted(OUTPUTS.glob("*BP_Model.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not models:
        raise FileNotFoundError("No generated BP model found in outputs/. Run scripts/smoke_test.py first.")
    return models[0]


def audit_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=False, read_only=False)
    sheet_metrics = []
    total_formulas = 0
    total_inputs = 0
    fills = Counter()
    validations = 0
    merged_ranges = 0
    for ws in wb.worksheets:
        formulas = 0
        inputs = 0
        non_empty = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                non_empty += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
                else:
                    inputs += 1
                fill = str(getattr(cell.fill.fgColor, "rgb", "") or "")
                if fill:
                    fills[fill] += 1
        validations += len(ws.data_validations.dataValidation)
        merged_ranges += len(ws.merged_cells.ranges)
        total_formulas += formulas
        total_inputs += inputs
        sheet_metrics.append(
            {
                "sheet": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "non_empty": non_empty,
                "formulas": formulas,
                "formula_density": round(formulas / non_empty, 3) if non_empty else 0,
                "validations": len(ws.data_validations.dataValidation),
                "merged": len(ws.merged_cells.ranges),
            }
        )
    return {
        "file": str(path),
        "sheets": len(wb.worksheets),
        "sheet_names": wb.sheetnames,
        "formulas": total_formulas,
        "inputs_or_labels": total_inputs,
        "formula_density": round(total_formulas / (total_formulas + total_inputs), 3) if total_formulas + total_inputs else 0,
        "validations": validations,
        "merged_ranges": merged_ranges,
        "top_fills": fills.most_common(10),
        "largest_sheets": sorted(sheet_metrics, key=lambda x: x["non_empty"], reverse=True)[:12],
        "most_formula_dense": sorted(sheet_metrics, key=lambda x: x["formulas"], reverse=True)[:12],
    }


def main() -> None:
    bolt = find_bolt_template()
    generated = latest_generated_model()
    report = {"generated": audit_workbook(generated)}
    if bolt:
        report["bolt"] = audit_workbook(bolt)
        report["parity"] = {
            "sheet_count_ratio": round(report["generated"]["sheets"] / max(1, report["bolt"]["sheets"]), 2),
            "formula_count_ratio": round(report["generated"]["formulas"] / max(1, report["bolt"]["formulas"]), 2),
            "validation_count_ratio": round(report["generated"]["validations"] / max(1, report["bolt"]["validations"]), 2),
            "generated_extra_sheets": [s for s in report["generated"]["sheet_names"] if s not in report["bolt"]["sheet_names"]],
            "bolt_sheet_count": report["bolt"]["sheets"],
            "generated_sheet_count": report["generated"]["sheets"],
        }
    else:
        report["bolt_error"] = "Project BOLT template not found from known local paths."
    print(json.dumps(report, indent=2))


if __name__ == "__main__":
    main()
