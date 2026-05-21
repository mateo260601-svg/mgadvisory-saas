from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook

from app.services.financial_mapping_service import normalize_project_financials
from app.services.project_service import create_project, project_dir
from app.schemas.project_schema import ProjectCreate


def main() -> None:
    project = create_project(ProjectCreate(company_name="EXTRACT TEST", project_type="QoE", currency="EUR"))
    documents = project_dir(project["id"]) / "documents"
    documents.mkdir(parents=True, exist_ok=True)
    workbook_path = documents / "management_accounts.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"
    ws["A4"] = "Line Item"
    ws["B4"] = "FY2023"
    ws["C4"] = "FY2024"
    ws["D4"] = "FY2025"
    ws.append(["Revenue", 1111, 2222, 3333])
    ws.append(["COGS", -400, -800, -1200])
    ws.append(["EBITDA", 300, 600, 900])
    wb.save(workbook_path)

    normalized = normalize_project_financials(project["id"], use_ai=False)
    revenue = next(line for line in normalized["income_statement"] if line["name"] == "Revenue")
    if revenue["values"].get("FY2023") != 1111:
        raise SystemExit(f"Revenue extraction failed: {revenue}")
    if revenue["values"].get("FY2025") != 3333:
        raise SystemExit(f"Revenue extraction failed: {revenue}")
    print(f"OK extracted periods={normalized['periods']} revenue={revenue['values']}")


if __name__ == "__main__":
    main()
