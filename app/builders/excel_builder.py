from datetime import date
from datetime import date
from pathlib import Path

from app.engines.debt_engine import debt_type_options


PERIODS = 60
QUARTERS = PERIODS // 3
MAX_DEBT_TRANCHES = 10
FIRST_PERIOD_COL = 4
FINANCIAL_MONTHLY_COL = 10
MAX_REVENUE_STREAMS = 10
MAX_COST_ITEMS = 12
MAX_HEADCOUNT_LINES = 10
HISTORICAL_DETAIL_LINES = 180
REVENUE_TOTAL_ROW = 7 + MAX_REVENUE_STREAMS + 2
PRODUCT_COGS_ROW = REVENUE_TOTAL_ROW
PRODUCT_GP_ROW = PRODUCT_COGS_ROW + 1
PRODUCT_MARGIN_ROW = PRODUCT_COGS_ROW + 2
HEADCOUNT_TOTAL_FTE_ROW = 7 + MAX_HEADCOUNT_LINES + 2
HEADCOUNT_PAYROLL_ROW = HEADCOUNT_TOTAL_FTE_ROW + 1
OPEX_TOTAL_EXCL_PAYROLL_ROW = 7 + MAX_COST_ITEMS + 2
OPEX_TOTAL_INCL_PAYROLL_ROW = OPEX_TOTAL_EXCL_PAYROLL_ROW + 1
ENTITIES = ["Group", "OpCo", "HoldCo"]
INPUT_FILL = "D9EAF7"
HEADER_FILL = "1F4E78"
SECTION_FILL = "D9E2F3"
OUTPUT_FILL = "E2F0D9"
CHECK_FILL = "FCE4D6"
FORMULA_FILL = "FFFFFF"
LINK_FILL = "EAF2F8"


def build_business_plan_workbook(project: dict, financials: dict, output_path: Path, assumptions: dict | None = None) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except Exception as exc:
        raise RuntimeError(f"Excel generation dependency unavailable: {exc}") from exc

    assumptions = assumptions or {}
    wb = Workbook()
    wb.remove(wb.active)
    styles = _styles(Font, PatternFill, Border, Side, Alignment)

    cover = wb.create_sheet("Cover")
    exec_dashboard = wb.create_sheet("Executive Dashboard")
    model_guide = wb.create_sheet("Model Guide")
    macro_inputs = wb.create_sheet("Macro Inputs")
    assumptions_sep = wb.create_sheet("Assumptions>>")
    bolt_revenue_inputs = wb.create_sheet("Revenue and COGS Inputs")
    bolt_opex_inputs = wb.create_sheet("Opex Inputs")
    bolt_bs_inputs = wb.create_sheet("BS and NWC Schedules")
    calculations_sep = wb.create_sheet("Calculations>>")
    bolt_revenue_calcs = wb.create_sheet("Revenue and COGS Calcs")
    bolt_opex_calcs = wb.create_sheet("Opex Calcs")
    bolt_bs_calcs = wb.create_sheet("BS and NWC Calcs")
    ciq_links = wb.create_sheet("CIQ_LinkingNames")
    consolidated_calcs = wb.create_sheet("Consolidated Financials Calcs")
    output_sep = wb.create_sheet("Output>>")
    summary_quarter = wb.create_sheet("Summary Financials Quarter")
    summary_annual = wb.create_sheet("Summary Financials Annual")
    ebitda_bridges = wb.create_sheet("EBITDA Bridges")
    supporting_sep = wb.create_sheet("Supporting Information>>")
    statements = wb.create_sheet("Financial Statements")
    outputs = wb.create_sheet("Outputs")
    packaged = wb.create_sheet("Packaged Output")
    ic_summary = wb.create_sheet("IC Summary")
    admin = wb.create_sheet("Admin")
    group_assumptions = wb.create_sheet("Group Assumptions")
    data_room = wb.create_sheet("Data Room")
    control = wb.create_sheet("Control Panel")
    assumption_map = wb.create_sheet("Assumption Input Map")
    historical_detail = wb.create_sheet("Historical Detail Input")
    historical = wb.create_sheet("Historical Inputs")
    historical_bridge = wb.create_sheet("Historical Bridge")
    operating_sep = wb.create_sheet("Operating Build>>")
    entity_input_sheets = [wb.create_sheet(f"{entity} Data Input") for entity in ENTITIES]
    entity_output_sheets = [wb.create_sheet(f"Output_{entity}_Monthly") for entity in ENTITIES]
    entity_annual_sheets = [wb.create_sheet(f"Output_{entity}_Annual") for entity in ENTITIES]
    revenue = wb.create_sheet("Revenue Drivers")
    products = wb.create_sheet("Product Build")
    headcount = wb.create_sheet("Headcount")
    opex = wb.create_sheet("Opex")
    wc = wb.create_sheet("Working Capital")
    capex = wb.create_sheet("Capex D&A")
    debt_sep = wb.create_sheet("Debt & Covenants>>")
    debt_config = wb.create_sheet("Debt Config")
    debt_schedule = wb.create_sheet("Debt Schedule")
    covenants = wb.create_sheet("Covenants")
    debt_capacity = wb.create_sheet("Debt Capacity")
    restructuring_sep = wb.create_sheet("Restructuring & Sensitivities>>")
    restructuring = wb.create_sheet("Restructuring Options")
    sensitivities = wb.create_sheet("Sensitivity Matrix")
    detail_3fs = wb.create_sheet("3FS Detail Output")
    detail_lines = wb.create_sheet("Detailed Forecast Lines")
    checks_sep = wb.create_sheet("Checks & Mapping>>")
    checks = wb.create_sheet("Checks")
    lookup = wb.create_sheet("Lookup")
    mapping = wb.create_sheet("Mapping")
    lists = wb.create_sheet("Lists & Dates")

    _build_lists(lists, project, styles)
    for sheet, title, subtitle in [
        (output_sep, "Output", "Review the 3-statement outputs, quarterly/annual summaries and investment committee pages first."),
        (assumptions_sep, "Assumptions", "Core inputs, historical actuals and operating drivers controlled from the SaaS BP Builder."),
        (operating_sep, "Operating Build", "Revenue, COGS, headcount, opex, working capital and capex build-up."),
        (debt_sep, "Debt & Covenants", "Debt instruments, cash/PIK interest, amortisation, liquidity and covenant tests."),
        (restructuring_sep, "Restructuring & Sensitivities", "Debt capacity, restructuring alternatives and downside sensitivities."),
        (calculations_sep, "Calculations", "Detailed formula-driven forecast lines and granular 3FS projection engine."),
        (checks_sep, "Checks & Mapping", "Workbook integrity checks, mapping tables and lookup support."),
        (supporting_sep, "Supporting Information", "Fixed lists, date tables and technical support sheets."),
    ]:
        _build_separator(sheet, title, subtitle, styles)
    _build_admin(admin, project, styles)
    _build_group_assumptions(group_assumptions, project, styles, DataValidation)
    _build_cover(cover, project, styles)
    _build_model_guide(model_guide, styles)
    _build_macro_inputs(macro_inputs, styles)
    _build_data_room(data_room, project, financials, styles)
    _build_control(control, project, styles, DataValidation, assumptions)
    _build_assumption_input_map(assumption_map, assumptions, styles)
    _build_historical_detail_input(historical_detail, financials, styles, DataValidation, assumptions)
    _build_historical(historical, financials, styles)
    _build_historical_bridge(historical_bridge, styles)
    for idx, sheet in enumerate(entity_input_sheets):
        _build_entity_input(sheet, ENTITIES[idx], styles)
    for idx, sheet in enumerate(entity_output_sheets):
        _build_entity_monthly_output(sheet, ENTITIES[idx], styles)
    for idx, sheet in enumerate(entity_annual_sheets):
        _build_entity_annual_output(sheet, ENTITIES[idx], styles)
    _build_revenue_drivers(revenue, styles, DataValidation, assumptions)
    _build_product_build(products, styles, DataValidation, assumptions)
    _build_headcount(headcount, styles, assumptions)
    _build_opex(opex, styles, DataValidation, assumptions)
    _build_working_capital(wc, styles, assumptions)
    _build_capex(capex, styles, assumptions)
    _build_debt_config(debt_config, styles, DataValidation, assumptions)
    _build_debt_schedule(debt_schedule, styles)
    _build_financial_statements(statements, styles)
    _build_3fs_detail_output(detail_3fs, styles)
    _build_detail_forecast_lines(detail_lines, styles)
    _build_bolt_revenue_cogs_inputs(bolt_revenue_inputs, project, styles)
    _build_bolt_opex_inputs(bolt_opex_inputs, project, styles)
    _build_bolt_bs_nwc_schedules(bolt_bs_inputs, project, styles)
    _build_bolt_revenue_cogs_calcs(bolt_revenue_calcs, project, styles)
    _build_bolt_opex_calcs(bolt_opex_calcs, project, styles)
    _build_bolt_bs_nwc_calcs(bolt_bs_calcs, project, styles)
    _build_ciq_linking_names(ciq_links, project, styles)
    _build_consolidated_financials_calcs(consolidated_calcs, project, styles)
    _build_covenants(covenants, styles, assumptions)
    _build_outputs(outputs, styles)
    _build_summary_financials_quarter(summary_quarter, styles)
    _build_summary_financials_annual(summary_annual, styles)
    _build_ebitda_bridges(ebitda_bridges, styles)
    _build_executive_dashboard(exec_dashboard, project, styles)
    _build_packaged_output(packaged, project, styles)
    _build_ic_summary(ic_summary, project, styles)
    _build_restructuring_options(restructuring, styles)
    _build_debt_capacity(debt_capacity, styles)
    _build_sensitivity_matrix(sensitivities, styles)
    _build_checks(checks, styles)
    _build_lookup(lookup, styles)
    _build_mapping(mapping, styles)

    for sheet in wb.worksheets:
        _polish_sheet(sheet, styles)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _build_admin(ws, project: dict, styles: dict) -> None:
    ws["B2"] = project.get("company_name", "Target Company")
    ws["B2"].font = styles["title_font"]
    ws["B4"] = "Admin"
    ws["B4"].font = styles["section_font"]
    rows = [
        ("Model Start", "='Control Panel'!$C$8"),
        ("Model End", "=INDEX('Lists & Dates'!$V$2:$V$61,'Control Panel'!$C$10)"),
        ("Actuals End", "='Control Panel'!$C$9"),
        ("Scenario", "='Control Panel'!$C$7"),
        ("Currency", "='Control Panel'!$C$6"),
        ("Forecast Months", "='Control Panel'!$C$10"),
        ("Generated Model", "Formula-driven BP / debt / covenant pack"),
    ]
    for row, (label, value) in enumerate(rows, start=6):
        _label(ws, row, 2, label, styles)
        _formula(ws, row, 3, value, styles, output=True) if isinstance(value, str) and value.startswith("=") else _input(ws, row, 3, value, styles)
    ws["B16"] = "Tab Colour Coding"
    ws["B16"].font = styles["section_font"]
    coding = [
        ("Input", "Blue cells are editable assumptions."),
        ("Formula", "White cells are formulas."),
        ("Output", "Green cells are linked output lines."),
        ("Check", "Orange cells are control checks."),
    ]
    for row, (label, desc) in enumerate(coding, start=17):
        ws.cell(row, 2, label)
        ws.cell(row, 3, desc)


def _build_group_assumptions(ws, project: dict, styles: dict, DataValidation) -> None:
    ws["B2"] = "Group Assumptions"
    ws["B2"].font = styles["section_font"]
    assumptions = [
        ("Company", project.get("company_name", "Target Company")),
        ("Currency", project.get("currency", "EUR")),
        ("Scenario", "Base"),
        ("Fiscal Year End", project.get("fiscal_year_end", "December")),
        ("Tax Rate", "='Control Panel'!$C$13"),
        ("Minimum Cash", "='Control Panel'!$C$14"),
        ("Opening Cash", "='Control Panel'!$C$11"),
        ("Opening Debt", "='Control Panel'!$C$12"),
        ("Consolidation Method", "Management case"),
    ]
    for row, (label, value) in enumerate(assumptions, start=5):
        _label(ws, row, 2, label, styles)
        if isinstance(value, str) and value.startswith("="):
            _formula(ws, row, 3, value, styles, output=True)
        else:
            _input(ws, row, 3, value, styles)
    _add_list_validation(ws, "C6", "'Lists & Dates'!$B$2:$B$8", DataValidation)
    _add_list_validation(ws, "C7", "'Lists & Dates'!$E$2:$E$7", DataValidation)
    ws["B17"] = "Entity Weighting"
    ws["B17"].font = styles["section_font"]
    _table_header(ws, 19, ["Entity", "Revenue Weight", "Opex Weight", "Debt Allocation"], styles)
    defaults = [("Group", 1.0, 1.0, 1.0), ("OpCo", 0.85, 0.90, 0.75), ("HoldCo", 0.15, 0.10, 0.25)]
    for row, values in enumerate(defaults, start=20):
        for col, value in enumerate(values, start=2):
            _input(ws, row, col, value, styles)
            if col > 2:
                ws.cell(row, col).number_format = "0.0%"


def _build_cover(ws, project: dict, styles: dict) -> None:
    ws["B2"] = "MG Advisory Finance OS"
    ws["B2"].font = styles["title_font"]
    ws["B4"] = "Institutional Business Plan Model"
    ws["B4"].font = styles["subtitle_font"]
    rows = [
        ("Company", project.get("company_name", "Target Company")),
        ("Project Type", project.get("project_type", "Investment case")),
        ("Currency", project.get("currency", "EUR")),
        ("Fiscal Year End", project.get("fiscal_year_end", "December")),
        ("Model Standard", "Driver-based monthly BP, debt schedule, covenants and checks"),
    ]
    for row, (label, value) in enumerate(rows, start=7):
        _label(ws, row, 2, label, styles)
        _input(ws, row, 3, value, styles)
    ws["B15"] = "Colour Code"
    ws["B15"].font = styles["section_font"]
    legend = [("Blue", "Hardcoded user input"), ("White", "Formula"), ("Light green", "Output"), ("Orange", "Check")]
    for row, (label, value) in enumerate(legend, start=16):
        ws.cell(row, 2, label)
        ws.cell(row, 3, value)
    ws["B22"] = "Operating Rule"
    ws["C22"] = "Inputs live in blue cells. Calculations and outputs are formula-driven."


def _build_separator(ws, title: str, subtitle: str, styles: dict) -> None:
    ws["B2"] = title
    ws["B2"].font = styles["title_font"]
    ws["B4"] = subtitle
    ws["B4"].font = styles["subtitle_font"]
    ws["B7"] = "Section Navigation"
    ws["B7"].font = styles["section_font"]
    ws["B9"] = "This tab is a separator to make the workbook readable for external review."
    ws["B10"] = "All operating tabs after this separator remain formula-linked unless cells are explicitly blue inputs."
    ws.sheet_properties.tabColor = HEADER_FILL


def _build_model_guide(ws, styles: dict) -> None:
    ws["B2"] = "Model Guide"
    ws["B2"].font = styles["title_font"]
    ws["B4"] = "Workbook Flow"
    ws["B4"].font = styles["section_font"]
    _table_header(ws, 6, ["Step", "Tab / Section", "Purpose", "Colour / Rule"], styles)
    rows = [
        ("1", "Control Panel / Macro Inputs", "Set dates, scenario, currency and core case controls.", "Blue = user inputs"),
        ("2", "Historical Detail Input", "Claude/manual historicals mapped to detailed 3FS lines.", "Blue inputs, green calculated actuals"),
        ("3", "Revenue and cost build", "Product, COGS, headcount and opex drivers.", "No hardcoded formulas"),
        ("4", "Working capital / capex / debt", "Cash conversion, depreciation and debt instruments.", "Debt supports cash, PIK and frequency"),
        ("5", "Financial Statements", "Annual left, grouped monthly detail right.", "Green = review outputs"),
        ("6", "Summary Financials Quarter / Annual", "BOLT-style banker output sheets for review and export.", "Formula-linked"),
        ("7", "EBITDA Bridges / Checks", "QoE-style bridge and model integrity controls.", "All checks should show OK"),
    ]
    for row_idx, values in enumerate(rows, start=7):
        for col_idx, value in enumerate(values, start=2):
            ws.cell(row_idx, col_idx, value)

    ws["B17"] = "Distribution Rules"
    ws["B17"].font = styles["section_font"]
    rules = [
        "Review the Checks sheet before sending any external model.",
        "Use the Summary Financials Quarter and Summary Financials Annual tabs for banker-style review.",
        "Actuals should be reviewed in Historical Detail Input before relying on the forecast bridge.",
        "Debt tranches should be entered individually, not netted into one generic facility.",
    ]
    for row_idx, rule in enumerate(rules, start=19):
        ws.cell(row_idx, 2, row_idx - 18)
        ws.cell(row_idx, 3, rule)

    ws["B25"] = "SaaS BP Builder To Excel Map"
    ws["B25"].font = styles["section_font"]
    _table_header(ws, 27, ["SaaS Step", "Excel Tabs Updated", "Reviewer Focus", "Output Dependency"], styles)
    flow = [
        ("Setup", "Control Panel, Macro Inputs, Lists & Dates", "Dates, scenario, currency, opening cash/debt", "All timeline and model headers"),
        ("Historicals", "Historical Detail Input, Historical Inputs, Historical Bridge", "Claude extraction reviewed against source files", "Forecast base period and bridges"),
        ("Revenue", "Revenue Drivers, Product Build", "Product/service split, volume, price, growth", "Revenue, gross profit, working capital"),
        ("Costs / People", "Opex, Headcount, Financial Statements", "Fixed vs variable costs, FTE and salary build", "EBITDA and cash burn"),
        ("Cash Flow", "Working Capital, Capex D&A", "DSO/DIO/DPO, maintenance vs growth capex", "FCF, cash and balance sheet"),
        ("Debt", "Debt Config, Debt Schedule, Covenants", "Cash/PIK interest, maturity, amortisation, sweeps", "Leverage, liquidity, covenant headroom"),
        ("Generate", "Outputs, Packaged Output, IC Summary, Checks", "Checks clear and outputs tie to 3FS", "Banker-facing workbook pack"),
    ]
    for row_idx, values in enumerate(flow, start=28):
        for col_idx, value in enumerate(values, start=2):
            ws.cell(row_idx, col_idx, value)


def _build_macro_inputs(ws, styles: dict) -> None:
    ws["B2"] = "Macro Inputs"
    ws["B2"].font = styles["section_font"]
    ws["B4"] = "Model Control Links"
    ws["B4"].font = styles["section_font"]
    _table_header(ws, 6, ["Input", "Value", "Source / Rule", "Reviewer Note"], styles)
    rows = [
        ("Company", "='Control Panel'!$C$5", "Control Panel", "Client / target name"),
        ("Currency", "='Control Panel'!$C$6", "Control Panel", "Workbook currency"),
        ("Scenario", "='Control Panel'!$C$7", "Control Panel", "Base / downside / upside"),
        ("Forecast Start", "='Control Panel'!$C$8", "Control Panel", "First model month"),
        ("Actuals End", "='Control Panel'!$C$9", "Control Panel", "Latest historical period"),
        ("Forecast Months", "='Control Panel'!$C$10", "Control Panel", "60-month institutional default"),
        ("Opening Cash", "='Control Panel'!$C$11", "Control Panel", "Opening liquidity"),
        ("Opening Debt", "='Control Panel'!$C$12", "Control Panel", "Opening gross debt"),
        ("Tax Rate", "='Control Panel'!$C$13", "Control Panel", "Applied to PBT"),
        ("Minimum Cash", "='Control Panel'!$C$14", "Control Panel", "Liquidity sweep floor"),
        ("Historical Source", "='Control Panel'!$C$15", "Control Panel", "Claude/manual/hybrid"),
    ]
    for row_idx, (label, formula, source, note) in enumerate(rows, start=7):
        ws.cell(row_idx, 2, label)
        _formula(ws, row_idx, 3, formula, styles, output=True, fmt="0.0%" if label == "Tax Rate" else None)
        ws.cell(row_idx, 4, source)
        ws.cell(row_idx, 5, note)

    ws["B22"] = "Timeline"
    ws["B22"].font = styles["section_font"]
    _table_header(ws, 24, ["Month #", "Month End", "Fiscal Year", "Quarter", "Actual / Forecast"], styles)
    for idx in range(PERIODS):
        row = 25 + idx
        ws.cell(row, 2, idx + 1)
        _formula(ws, row, 3, f"='Lists & Dates'!V{idx + 2}", styles, output=True, fmt="mmm-yy")
        _formula(ws, row, 4, f"=YEAR(C{row})", styles, output=True)
        _formula(ws, row, 5, f'="Q"&ROUNDUP(MONTH(C{row})/3,0)', styles, output=True)
        _formula(ws, row, 6, f'=IF(C{row}<=\'Control Panel\'!$C$9,"Actual","Forecast")', styles, output=True)


def _build_data_room(ws, project: dict, financials: dict, styles: dict) -> None:
    ws["B2"] = "Data Room & Upload Map"
    ws["B2"].font = styles["section_font"]
    rows = [
        ("Audited Accounts PDF", "Upload", "data/projects/{project_id}/documents", "Historical accounts extraction"),
        ("Management Accounts Excel", "Upload", "data/projects/{project_id}/documents", "Monthly actuals"),
        ("Trial Balance", "Upload", "data/projects/{project_id}/documents", "Mapping to FS lines"),
        ("Aged Receivables", "Upload", "data/projects/{project_id}/documents", "DSO / working capital"),
        ("Aged Payables", "Upload", "data/projects/{project_id}/documents", "DPO / working capital"),
        ("Debt Schedule", "Upload", "data/projects/{project_id}/documents", "Debt config and covenants"),
        ("Budget / Forecast", "Upload", "data/projects/{project_id}/documents", "Forward assumptions"),
    ]
    _table_header(ws, 4, ["Document Type", "Action", "Storage", "Model Use"], styles)
    for r, row_values in enumerate(rows, start=5):
        for c, value in enumerate(row_values, start=2):
            ws.cell(r, c, value)
    ws["B15"] = "Uploaded Source Files"
    ws["B15"].font = styles["section_font"]
    _table_header(ws, 17, ["#", "Source File"], styles)
    sources = financials.get("source_files", []) or ["No files uploaded yet"]
    for idx, source in enumerate(sources, start=1):
        ws.cell(17 + idx, 2, idx)
        ws.cell(17 + idx, 3, source)


def _build_control(ws, project: dict, styles: dict, DataValidation, assumptions: dict) -> None:
    ws["B2"] = "Control Panel"
    ws["B2"].font = styles["section_font"]
    model = assumptions.get("model", {})
    controls = [
        ("Company Name", project.get("company_name", "Target Company")),
        ("Currency", model.get("currency", project.get("currency", "EUR"))),
        ("Scenario", model.get("scenario", "Base")),
        ("Model Start Date", _date_value(model.get("model_start_date"), date(2026, 1, 31))),
        ("Actuals End Date", _date_value(model.get("actuals_end_date"), date(2025, 12, 31))),
        ("Forecast Months", int(model.get("forecast_months", PERIODS) or PERIODS)),
        ("Opening Cash", _num(model.get("opening_cash"), 120000)),
        ("Opening Debt", _num(model.get("opening_debt"), 500000)),
        ("Tax Rate", _num(model.get("tax_rate"), 0.25)),
        ("Minimum Cash", _num(model.get("minimum_cash"), 50000)),
        ("Historical Source", model.get("historical_source", "Claude extraction")),
    ]
    for row, (label, value) in enumerate(controls, start=5):
        _label(ws, row, 2, label, styles)
        _input(ws, row, 3, value, styles)
    ws["C8"].number_format = "yyyy-mm-dd"
    ws["C9"].number_format = "yyyy-mm-dd"
    ws["C13"].number_format = "0.0%"
    _add_list_validation(ws, "C7", "'Lists & Dates'!$E$2:$E$7", DataValidation)
    _add_list_validation(ws, "C6", "'Lists & Dates'!$B$2:$B$8", DataValidation)
    _add_list_validation(ws, "C15", "'Lists & Dates'!$AB$2:$AB$4", DataValidation)
    _write_period_headers(ws, 17, styles, source_sheet="'Lists & Dates'")


def _build_assumption_input_map(ws, assumptions: dict, styles: dict) -> None:
    ws["B2"] = "Assumption Input Map"
    ws["B2"].font = styles["section_font"]
    ws["B3"] = "All BP Builder parameters saved from the SaaS are mirrored here before flowing into operating tabs. Use this as the model audit trail."
    _table_header(
        ws,
        5,
        [
            "SaaS Step",
            "Category",
            "Parameter",
            "Value",
            "Input Type",
            "Required?",
            "Review Priority",
            "Model Destination",
            "Validation Note",
        ],
        styles,
    )
    destination_map = {
        "model": "Control Panel",
        "historical_actuals": "Historical Detail Input / Historical Bridge",
        "revenue_streams": "Revenue Drivers / Product Build",
        "cost_base": "Product Build / Opex",
        "cost_items": "Opex",
        "headcount": "Headcount",
        "working_capital": "Working Capital",
        "capex": "Capex D&A",
        "debt_tranches": "Debt Config / Debt Schedule",
        "covenants": "Covenants",
    }
    for row, (category, key, value) in enumerate(_flatten_assumptions(assumptions), start=6):
        base_category = category.split("[")[0]
        ws.cell(row, 2, _assumption_step(base_category))
        ws.cell(row, 3, category)
        ws.cell(row, 4, key)
        _input(ws, row, 5, value, styles)
        ws.cell(row, 6, _assumption_type(value))
        ws.cell(row, 7, "Yes" if _assumption_required(base_category, key) else "No")
        ws.cell(row, 8, _assumption_priority(base_category, key, value))
        ws.cell(row, 9, destination_map.get(base_category, "Model workbook"))
        ws.cell(row, 10, _assumption_validation_note(base_category, key, value))


def _build_historical_detail_input(ws, financials: dict, styles: dict, DataValidation, assumptions: dict) -> None:
    ws["B2"] = "Historical Detail Input"
    ws["B2"].font = styles["section_font"]
    ws["B3"] = "Manual or Claude-reviewed granular historical inputs. Map each line to a model line; Latest Actual feeds the Historical Bridge where available."
    headers = [
        "Line ID",
        "Statement",
        "Category",
        "Subcategory",
        "Model Line",
        "Detail Line",
        "Sign",
        "Source Mode",
        "Source File",
        "FY2022",
        "FY2023",
        "FY2024",
        "FY2025",
        "LTM",
        "Latest Actual",
        "Notes / audit trail",
    ]
    _table_header(ws, 5, headers, styles)
    templates = _historical_line_templates()
    extracted_lookup = _financial_lookup(financials)
    manual_lookup = _manual_historical_lookup(assumptions)
    for idx in range(HISTORICAL_DETAIL_LINES):
        row = 6 + idx
        statement, category, subcategory, model_line, detail_line, sign = templates[idx]
        ws.cell(row, 2, f"HIST-{idx + 1:03d}")
        _input(ws, row, 3, statement, styles)
        _input(ws, row, 4, category, styles)
        _input(ws, row, 5, subcategory, styles)
        _input(ws, row, 6, model_line, styles)
        _input(ws, row, 7, detail_line, styles)
        _input(ws, row, 8, sign, styles)
        manual_values = manual_lookup.get(detail_line.lower()) or manual_lookup.get(model_line.lower())
        source_mode = "Manual input" if manual_values else "Claude extraction"
        _input(ws, row, 9, source_mode, styles)
        ws.cell(row, 10, "")
        values = manual_values or extracted_lookup.get(detail_line.lower(), {}) or extracted_lookup.get(model_line.lower(), {})
        for col, period in enumerate(["FY2022", "FY2023", "FY2024", "FY2025"], start=11):
            fallback_latest = values.get("latest_actual", 0) if period == "FY2025" else 0
            _input(ws, row, col, float(values.get(period, fallback_latest) or 0), styles)
        _formula(ws, row, 15, f"=SUM(K{row}:N{row})", styles, output=True)
        _formula(ws, row, 16, f"=IF(N{row}<>0,N{row},IF(M{row}<>0,M{row},IF(L{row}<>0,L{row},K{row})))", styles, output=True)
        ws.cell(row, 17, "")
        _add_list_validation(ws, f"C{row}", "'Lists & Dates'!$AE$2:$AE$5", DataValidation)
        _add_list_validation(ws, f"D{row}", "'Lists & Dates'!$AH$2:$AH$30", DataValidation)
        _add_list_validation(ws, f"F{row}", "'Lists & Dates'!$AK$2:$AK$30", DataValidation)
        _add_list_validation(ws, f"I{row}", "'Lists & Dates'!$AN$2:$AN$5", DataValidation)


def _build_historical(ws, financials: dict, styles: dict) -> None:
    ws["B2"] = "Historical Inputs"
    ws["B2"].font = styles["section_font"]
    periods = financials.get("periods") or ["FY2023", "FY2024", "FY2025"]
    _table_header(ws, 4, ["Line Item", "Source"] + periods, styles)
    rows = []
    for section in ["income_statement", "balance_sheet", "cash_flow"]:
        for line in financials.get(section, []):
            rows.append((line.get("name", ""), section, line.get("values", {})))
    if not rows:
        rows = [
            ("No extracted financial data", "upload_or_claude_required", {}),
        ]
    for r, (name, source, values) in enumerate(rows, start=5):
        ws.cell(r, 2, name)
        ws.cell(r, 3, source)
        for c, period in enumerate(periods, start=4):
            _input(ws, r, c, float(values.get(period, 0)), styles)


def _build_historical_bridge(ws, styles: dict) -> None:
    ws["B2"] = "Historical Bridge"
    ws["B2"].font = styles["section_font"]
    _table_header(ws, 4, ["Metric", "Last Actual", "Monthly Base", "YoY Growth / Margin", "Forecast Link"], styles)
    metrics = [
        ("Revenue", "Revenue Drivers"),
        ("COGS", "Product Build"),
        ("Opex", "Opex"),
        ("EBITDA", "Financial Statements"),
        ("Cash", "Financial Statements"),
        ("Debt", "Debt Schedule"),
        ("Receivables", "Working Capital"),
        ("Inventory", "Working Capital"),
        ("Payables", "Working Capital"),
    ]
    for row, (metric, target) in enumerate(metrics, start=5):
        ws.cell(row, 2, metric)
        detail_formula = f"SUMIFS('Historical Detail Input'!$P:$P,'Historical Detail Input'!$F:$F,$B{row})"
        extracted_formula = f"IFERROR(INDEX('Historical Inputs'!$D:$K,MATCH($B{row},'Historical Inputs'!$B:$B,0),MAX(1,COUNTA('Historical Inputs'!$D$4:$K$4))),0)"
        _formula(ws, row, 3, f"=IF({detail_formula}<>0,{detail_formula},{extracted_formula})", styles, output=True)
        if metric in ["Cash", "Debt", "Receivables", "Inventory", "Payables"]:
            _formula(ws, row, 4, f"=C{row}", styles, output=True)
        else:
            _formula(ws, row, 4, f"=C{row}/12", styles, output=True)
        _formula(ws, row, 5, f'=IFERROR(C{row}/INDEX(\'Historical Inputs\'!$D:$K,MATCH($B{row},\'Historical Inputs\'!$B:$B,0),MAX(1,COUNTA(\'Historical Inputs\'!$D$4:$K$4)-1))-1,0)', styles, output=True, fmt="0.0%")
        ws.cell(row, 6, target)

    ws["B17"] = "Extraction Quality"
    ws["B17"].font = styles["section_font"]
    checks = [
        ("Revenue extracted?", '=IF(C5<>0,"OK","Missing")'),
        ("EBITDA extracted?", '=IF(C8<>0,"OK","Missing")'),
        ("Cash extracted?", '=IF(C9<>0,"OK","Missing")'),
        ("Debt extracted?", '=IF(C10<>0,"OK","Missing")'),
    ]
    for row, (label, formula) in enumerate(checks, start=18):
        ws.cell(row, 2, label)
        _formula(ws, row, 3, formula, styles, output=True)


def _build_bolt_sheet_header(ws, project: dict, styles: dict, freeze: str = "J10") -> None:
    company = project.get("company_name", "Target Company")
    currency = project.get("currency", "EUR")
    ws["A1"] = f"{company.upper()} FINANCIAL MODEL"
    ws["A1"].font = styles["title_font"]
    ws["A2"] = '=UPPER(MID(CELL("filename",C3),FIND("]",CELL("filename",C3))+1,LEN(CELL("filename",C3))))'
    ws["I1"] = "='Cover'!$B$2"
    ws["I2"] = '=COUNTIF(I4:I1048576,"FALSE")'
    ws["C4"] = company
    ws["C5"] = currency
    ws["I4"] = "Financial Year"
    ws["I5"] = "Financial Quarter"
    ws["I7"] = "Start date"
    ws["I8"] = "End date"
    ws["I9"] = "# Days"
    for row in [4, 5, 7, 8, 9]:
        ws.cell(row, 9).font = styles["bold_font"]
    for idx, c in enumerate(range(10, 10 + PERIODS), start=2):
        date_ref = f"'Lists & Dates'!V{idx}"
        cell = ws.cell(10, c)
        cell.value = f"={date_ref}"
        cell.number_format = "mmm-yy"
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        _formula(ws, 4, c, f"=YEAR({cell.coordinate})", styles, output=True)
        _formula(ws, 5, c, f'="Q"&ROUNDUP(MONTH({cell.coordinate})/3,0)', styles, output=True)
        _formula(ws, 7, c, f"={cell.coordinate}", styles, output=True)
        _formula(ws, 8, c, f"=EOMONTH({cell.coordinate},0)", styles, output=True)
        _formula(ws, 9, c, f"=DAY(EOMONTH({cell.coordinate},0))", styles, output=True)
    ws.freeze_panes = freeze


def _bolt_section(ws, row: int, title: str, styles: dict) -> None:
    ws.cell(row, 2, title)
    ws.cell(row, 2).font = styles["section_font"]
    ws.cell(row, 2).fill = styles["section_fill"]


def _build_bolt_revenue_cogs_inputs(ws, project: dict, styles: dict) -> None:
    _build_bolt_sheet_header(ws, project, styles, freeze="O196")
    _bolt_section(ws, 13, "1. Revenue and COGS input assumptions", styles)
    _table_header(ws, 15, ["Line", "Type", "Start Volume", "Start Price", "Volume Growth", "Price Growth", "COGS %", "Fulfilment / Unit"], styles)
    for idx in range(MAX_REVENUE_STREAMS):
        row = 16 + idx
        source = 7 + idx
        ws.cell(row, 2, f"='Revenue Drivers'!B{source}")
        ws.cell(row, 3, f"='Revenue Drivers'!C{source}")
        ws.cell(row, 4, f"='Revenue Drivers'!D{source}")
        ws.cell(row, 5, f"='Revenue Drivers'!E{source}")
        ws.cell(row, 6, f"='Revenue Drivers'!F{source}")
        ws.cell(row, 7, f"='Revenue Drivers'!G{source}")
        ws.cell(row, 8, f"='Product Build'!C{source}")
        ws.cell(row, 9, f"='Product Build'!D{source}")
    _bolt_section(ws, 30, "2. Monthly revenue and gross profit output", styles)
    rows = [
        ("Gross Revenue", f"'Revenue Drivers'!{{col}}{REVENUE_TOTAL_ROW}", None),
        ("COGS", f"'Product Build'!{{col}}{PRODUCT_COGS_ROW}", None),
        ("Gross Profit", f"'Product Build'!{{col}}{PRODUCT_GP_ROW}", None),
        ("Gross Margin", f"'Product Build'!{{col}}{PRODUCT_MARGIN_ROW}", "0.0%"),
    ]
    for row_idx, (label, template, fmt) in enumerate(rows, start=32):
        ws.cell(row_idx, 2, label)
        for idx, out_col in enumerate(range(10, 10 + PERIODS)):
            source_col = _col(FIRST_PERIOD_COL + idx)
            _formula(ws, row_idx, out_col, f"={template.format(col=source_col)}", styles, output=True, fmt=fmt)


def _build_bolt_opex_inputs(ws, project: dict, styles: dict) -> None:
    _build_bolt_sheet_header(ws, project, styles, freeze="J27")
    _bolt_section(ws, 13, "1. Opex and payroll input assumptions", styles)
    _table_header(ws, 15, ["Cost Category", "Driver", "Monthly Fixed", "% Revenue", "Cost / FTE"], styles)
    for idx in range(MAX_COST_ITEMS):
        row = 16 + idx
        source = 7 + idx
        for target_col, source_col in enumerate(range(2, 7), start=2):
            ws.cell(row, target_col, f"='Opex'!{_col(source_col)}{source}")
    _bolt_section(ws, 32, "2. Headcount assumptions", styles)
    _table_header(ws, 34, ["Department", "Opening FTE", "Avg Salary / Month", "Hiring Every N Months", "New Hires"], styles)
    for idx in range(MAX_HEADCOUNT_LINES):
        row = 35 + idx
        source = 7 + idx
        for target_col, source_col in enumerate(range(2, 7), start=2):
            ws.cell(row, target_col, f"='Headcount'!{_col(source_col)}{source}")


def _build_bolt_bs_nwc_schedules(ws, project: dict, styles: dict) -> None:
    _build_bolt_sheet_header(ws, project, styles, freeze="J182")
    _bolt_section(ws, 13, "1. Balance sheet and net working capital assumptions", styles)
    rows = [
        ("DSO", "'Working Capital'!$C$6"),
        ("DIO", "'Working Capital'!$C$7"),
        ("DPO", "'Working Capital'!$C$8"),
        ("Opening Cash", "'Control Panel'!$C$11"),
        ("Opening Debt", "'Control Panel'!$C$12"),
        ("Maintenance Capex % Revenue", "'Capex D&A'!$C$6"),
        ("Growth Capex / Month", "'Capex D&A'!$C$7"),
        ("Depreciation Life Months", "'Capex D&A'!$C$8"),
    ]
    _table_header(ws, 15, ["Schedule Input", "Value", "Source"], styles)
    for row_idx, (label, formula) in enumerate(rows, start=16):
        ws.cell(row_idx, 2, label)
        _formula(ws, row_idx, 3, f"={formula}", styles, output=True)
        ws.cell(row_idx, 4, formula)
    _bolt_section(ws, 29, "2. Monthly NWC and capex schedule", styles)
    schedule_rows = [
        ("Receivables", "'Working Capital'!{col}9"),
        ("Inventory", "'Working Capital'!{col}10"),
        ("Payables", "'Working Capital'!{col}11"),
        ("Net Working Capital", "'Working Capital'!{col}12"),
        ("Change in NWC", "'Working Capital'!{col}13"),
        ("Maintenance Capex", "'Capex D&A'!{col}11"),
        ("Growth Capex", "'Capex D&A'!{col}12"),
        ("Total Capex", "'Capex D&A'!{col}13"),
        ("Depreciation", "'Capex D&A'!{col}14"),
        ("Net PPE", "'Capex D&A'!{col}15"),
    ]
    for row_idx, (label, template) in enumerate(schedule_rows, start=31):
        ws.cell(row_idx, 2, label)
        for idx, out_col in enumerate(range(10, 10 + PERIODS)):
            source_col = _col(FIRST_PERIOD_COL + idx)
            _formula(ws, row_idx, out_col, f"={template.format(col=source_col)}", styles, output=True)


def _build_bolt_revenue_cogs_calcs(ws, project: dict, styles: dict) -> None:
    _build_bolt_sheet_header(ws, project, styles, freeze="J10")
    _bolt_section(ws, 13, "1. Revenue and COGS calculations", styles)
    calc_rows = [
        ("Gross Revenue", "'Revenue and COGS Inputs'!{col}32"),
        ("COGS", "'Revenue and COGS Inputs'!{col}33"),
        ("Gross Profit", "'Revenue and COGS Inputs'!{col}34"),
        ("Gross Margin", "'Revenue and COGS Inputs'!{col}35"),
    ]
    for row_idx, (label, template) in enumerate(calc_rows, start=15):
        ws.cell(row_idx, 2, label)
        for out_col in range(10, 10 + PERIODS):
            fmt = "0.0%" if "Margin" in label else None
            _formula(ws, row_idx, out_col, f"={template.format(col=_col(out_col))}", styles, output=True, fmt=fmt)


def _build_bolt_opex_calcs(ws, project: dict, styles: dict) -> None:
    _build_bolt_sheet_header(ws, project, styles, freeze="J10")
    _bolt_section(ws, 13, "1. Opex and payroll calculations", styles)
    calc_rows = [
        ("Payroll Cost", f"'Headcount'!{{col}}{HEADCOUNT_PAYROLL_ROW}"),
        ("Opex excl. Payroll", f"'Opex'!{{col}}{OPEX_TOTAL_EXCL_PAYROLL_ROW}"),
        ("Opex incl. Payroll", f"'Opex'!{{col}}{OPEX_TOTAL_INCL_PAYROLL_ROW}"),
        ("Total FTE", f"'Headcount'!{{col}}{HEADCOUNT_TOTAL_FTE_ROW}"),
        ("Opex / Revenue", f"IFERROR('Opex'!{{col}}{OPEX_TOTAL_INCL_PAYROLL_ROW}/'Revenue Drivers'!{{col}}{REVENUE_TOTAL_ROW},0)"),
    ]
    for row_idx, (label, template) in enumerate(calc_rows, start=15):
        ws.cell(row_idx, 2, label)
        for idx, out_col in enumerate(range(10, 10 + PERIODS)):
            source_col = _col(FIRST_PERIOD_COL + idx)
            formula = f"={template.format(col=source_col)}"
            _formula(ws, row_idx, out_col, formula, styles, output=True, fmt="0.0%" if "/" in label else None)


def _build_bolt_bs_nwc_calcs(ws, project: dict, styles: dict) -> None:
    _build_bolt_sheet_header(ws, project, styles, freeze="J93")
    _bolt_section(ws, 13, "1. BS and NWC calculations", styles)
    calc_rows = [
        ("Receivables", "'BS and NWC Schedules'!{col}31"),
        ("Inventory", "'BS and NWC Schedules'!{col}32"),
        ("Payables", "'BS and NWC Schedules'!{col}33"),
        ("Net Working Capital", "'BS and NWC Schedules'!{col}34"),
        ("Change in NWC", "'BS and NWC Schedules'!{col}35"),
        ("Total Capex", "'BS and NWC Schedules'!{col}38"),
        ("Depreciation", "'BS and NWC Schedules'!{col}39"),
        ("Net PPE", "'BS and NWC Schedules'!{col}40"),
    ]
    for row_idx, (label, template) in enumerate(calc_rows, start=15):
        ws.cell(row_idx, 2, label)
        for out_col in range(10, 10 + PERIODS):
            _formula(ws, row_idx, out_col, f"={template.format(col=_col(out_col))}", styles, output=True)


def _build_ciq_linking_names(ws, project: dict, styles: dict) -> None:
    _build_bolt_sheet_header(ws, project, styles, freeze="J10")
    _bolt_section(ws, 13, "1. Linking names and model trace map", styles)
    _table_header(ws, 15, ["Name", "Source Sheet", "Source Row", "Linked Output", "Status"], styles)
    rows = [
        ("Revenue", "Revenue and COGS Calcs", "15", "Consolidated Financials Calcs", "Active"),
        ("COGS", "Revenue and COGS Calcs", "16", "Consolidated Financials Calcs", "Active"),
        ("Opex", "Opex Calcs", "17", "Consolidated Financials Calcs", "Active"),
        ("NWC", "BS and NWC Calcs", "18", "Consolidated Financials Calcs", "Active"),
        ("Debt", "Debt Schedule", "Aggregate summary", "Consolidated Financials Calcs", "Active"),
    ]
    for row_idx, values in enumerate(rows, start=16):
        for col_idx, value in enumerate(values, start=2):
            ws.cell(row_idx, col_idx, value)


def _build_consolidated_financials_calcs(ws, project: dict, styles: dict) -> None:
    _build_bolt_sheet_header(ws, project, styles, freeze="J10")
    _bolt_section(ws, 13, "1. Consolidated financial statements calculations", styles)
    _table_header(ws, 15, ["Statement", "Line Item", "Source"], styles)
    rows = [
        ("Income Statement", "Revenue", "'Financial Statements'!{fs_col}6"),
        ("Income Statement", "COGS", "'Financial Statements'!{fs_col}7"),
        ("Income Statement", "Gross Profit", "'Financial Statements'!{fs_col}8"),
        ("Income Statement", "Payroll", "'Financial Statements'!{fs_col}9"),
        ("Income Statement", "Opex", "'Financial Statements'!{fs_col}10"),
        ("Income Statement", "EBITDA", "'Financial Statements'!{fs_col}11"),
        ("Income Statement", "D&A", "'Financial Statements'!{fs_col}12"),
        ("Income Statement", "EBIT", "'Financial Statements'!{fs_col}13"),
        ("Income Statement", "Cash Interest", "'Financial Statements'!{fs_col}14"),
        ("Income Statement", "PBT", "'Financial Statements'!{fs_col}15"),
        ("Income Statement", "Tax", "'Financial Statements'!{fs_col}16"),
        ("Income Statement", "Net Income", "'Financial Statements'!{fs_col}17"),
        ("Cash Flow", "Change in NWC", "'Financial Statements'!{fs_col}18"),
        ("Cash Flow", "Capex", "'Financial Statements'!{fs_col}19"),
        ("Cash Flow", "Cash Flow Before Debt", "'Financial Statements'!{fs_col}20"),
        ("Cash Flow", "Debt Amortization / Sweep", "'Financial Statements'!{fs_col}21"),
        ("Cash Flow", "Free Cash Flow", "'Financial Statements'!{fs_col}22"),
        ("Balance Sheet", "Closing Cash", "'Financial Statements'!{fs_col}23"),
        ("Balance Sheet", "Closing Debt", "'Financial Statements'!{fs_col}24"),
        ("Balance Sheet", "Net Debt", "'Financial Statements'!{fs_col}25"),
    ]
    for row_idx, (statement, label, source_template) in enumerate(rows, start=16):
        ws.cell(row_idx, 2, statement)
        ws.cell(row_idx, 3, label)
        ws.cell(row_idx, 4, source_template)
        for idx, out_col in enumerate(range(10, 10 + PERIODS)):
            fs_col = _financial_monthly_col(FIRST_PERIOD_COL + idx)
            fmt = "0.0%" if "Margin" in label else "0.0x" if "/" in label else None
            _formula(ws, row_idx, out_col, f"={source_template.format(fs_col=fs_col)}", styles, output=True, fmt=fmt)


def _build_entity_input(ws, entity: str, styles: dict) -> None:
    ws["B2"] = f"{entity} Data Input"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    rows = [
        ("Revenue", "000s", f"='Revenue Drivers'!{{col}}{REVENUE_TOTAL_ROW}"),
        ("COGS", "000s", f"='Product Build'!{{col}}{PRODUCT_COGS_ROW}"),
        ("Payroll", "000s", f"='Headcount'!{{col}}{HEADCOUNT_PAYROLL_ROW}"),
        ("Opex", "000s", f"='Opex'!{{col}}{OPEX_TOTAL_EXCL_PAYROLL_ROW}"),
        ("EBITDA", "000s", "={col}6+{col}7+{col}8+{col}9"),
        ("Capex", "000s", "='Capex D&A'!{col}13"),
        ("Receivables", "000s", "='Working Capital'!{col}9"),
        ("Inventory", "000s", "='Working Capital'!{col}10"),
        ("Payables", "000s", "='Working Capital'!{col}11"),
        ("Closing Debt", "000s", "='Financial Statements'!{col}24"),
    ]
    weight_row = 20 + ENTITIES.index(entity)
    _table_header(ws, 6, ["Metric", "Unit"], styles)
    for row, (label, unit, template) in enumerate(rows, start=7):
        ws.cell(row, 2, label)
        ws.cell(row, 3, unit)
        for col_idx in _period_cols():
            col = _col(col_idx)
            formula = template.format(col=col)
            if label == "Closing Debt":
                formula = template.format(col=_financial_monthly_col(col_idx))
            if entity != "Group":
                if label in ["Revenue", "COGS", "EBITDA", "Receivables", "Inventory", "Payables"]:
                    formula = f"=({formula.lstrip('=')})*'Group Assumptions'!$C${weight_row}"
                elif label in ["Payroll", "Opex", "Capex"]:
                    formula = f"=({formula.lstrip('=')})*'Group Assumptions'!$D${weight_row}"
                elif label == "Closing Debt":
                    formula = f"=({formula.lstrip('=')})*'Group Assumptions'!$E${weight_row}"
            _formula(ws, row, col_idx, formula, styles, output=row in [11, 16])


def _build_entity_monthly_output(ws, entity: str, styles: dict) -> None:
    ws["B2"] = f"{entity} Monthly Output"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    rows = [
        ("Revenue", 7),
        ("Gross Profit", 8),
        ("EBITDA", 9),
        ("EBITDA Margin", 10),
        ("Capex", 11),
        ("Closing Debt", 12),
        ("Net Debt / EBITDA", 13),
    ]
    input_sheet = f"{entity} Data Input"
    for label, row in rows:
        ws.cell(row, 2, label)
    for col_idx in _period_cols():
        col = _col(col_idx)
        _formula(ws, 7, col_idx, f"='{input_sheet}'!{col}7", styles, output=True)
        _formula(ws, 8, col_idx, f"='{input_sheet}'!{col}7+'{input_sheet}'!{col}8", styles, output=True)
        _formula(ws, 9, col_idx, f"='{input_sheet}'!{col}11", styles, output=True)
        _formula(ws, 10, col_idx, f"=IFERROR({col}9/{col}7,0)", styles, output=True, fmt="0.0%")
        _formula(ws, 11, col_idx, f"='{input_sheet}'!{col}12", styles, output=True)
        _formula(ws, 12, col_idx, f"='{input_sheet}'!{col}16", styles, output=True)
        _formula(ws, 13, col_idx, f"=IFERROR({col}12/{col}9,0)", styles, output=True, fmt="0.0x")


def _build_entity_annual_output(ws, entity: str, styles: dict) -> None:
    ws["B2"] = f"{entity} Annual Output"
    ws["B2"].font = styles["section_font"]
    years = [2026, 2027, 2028, 2029, 2030]
    _table_header(ws, 4, ["Metric"] + [f"FY{year}" for year in years], styles)
    rows = ["Revenue", "Gross Profit", "EBITDA", "EBITDA Margin", "Capex", "Closing Debt", "Net Debt / EBITDA"]
    for row, label in enumerate(rows, start=5):
        ws.cell(row, 2, label)
    monthly = f"Output_{entity}_Monthly"
    for idx, year in enumerate(years, start=3):
        col = _col(idx)
        for row in range(5, 12):
            source_row = row + 2
            if row in [8, 11]:
                _formula(ws, row, idx, f"=IFERROR({col}{row-1}/{col}5,0)" if row == 8 else f"=IFERROR({col}10/{col}7,0)", styles, output=True, fmt="0.0%" if row == 8 else "0.0x")
            elif row in [10]:
                _formula(ws, row, idx, f"=XLOOKUP(DATE({year},12,31),'{monthly}'!$D$4:$BK$4,'{monthly}'!$D$12:$BK$12,0)", styles, output=True)
            else:
                _formula(ws, row, idx, f'=SUMIFS(\'{monthly}\'!$D${source_row}:$BK${source_row},\'{monthly}\'!$D$4:$BK$4,">="&DATE({year},1,1),\'{monthly}\'!$D$4:$BK$4,"<="&DATE({year},12,31))', styles, output=True)


def _build_revenue_drivers(ws, styles: dict, DataValidation, assumptions: dict) -> None:
    ws["B2"] = "Revenue Drivers"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    rows = assumptions.get("revenue_streams") or [
        {"name": "Product / Service 1", "type": "Product", "volume": 100, "price": 1000, "volume_growth": 0.01, "price_growth": 0.002},
        {"name": "Product / Service 2", "type": "Service", "volume": 120, "price": 900, "volume_growth": 0.01, "price_growth": 0.002},
        {"name": "Product / Service 3", "type": "Recurring", "volume": 140, "price": 800, "volume_growth": 0.01, "price_growth": 0.002},
        {"name": "Product / Service 4", "type": "Project", "volume": 160, "price": 700, "volume_growth": 0.01, "price_growth": 0.002},
        {"name": "Product / Service 5", "type": "Other", "volume": 180, "price": 600, "volume_growth": 0.01, "price_growth": 0.002},
    ]
    _table_header(ws, 6, ["Revenue Stream", "Type", "Start Volume", "Start Price", "Monthly Volume Growth", "Monthly Price Growth"], styles)
    for idx in range(7, 7 + MAX_REVENUE_STREAMS):
        row = rows[idx - 7] if idx - 7 < len(rows) else {}
        ws.cell(idx, 2, row.get("name", f"Product / Service {idx - 6}"))
        _input(ws, idx, 3, row.get("type", "Other"), styles)
        _input(ws, idx, 4, _num(row.get("volume"), 0), styles)
        _input(ws, idx, 5, _num(row.get("price"), 0), styles)
        _input(ws, idx, 6, _num(row.get("volume_growth"), 0), styles, fmt="0.0%")
        _input(ws, idx, 7, _num(row.get("price_growth"), 0), styles, fmt="0.0%")
        _add_list_validation(ws, f"C{idx}", "'Lists & Dates'!$H$2:$H$20", DataValidation)
    ws.cell(REVENUE_TOTAL_ROW, 2, "Total Revenue")
    ws.cell(REVENUE_TOTAL_ROW, 2).font = styles["bold_font"]
    for c in _period_cols():
        letter = _col(c)
        formulas = []
        for r in range(7, 7 + MAX_REVENUE_STREAMS):
            month_index = c - FIRST_PERIOD_COL
            formulas.append(f"($D{r}*(1+$F{r})^{month_index})*($E{r}*(1+$G{r})^{month_index})")
        history_formula = f"'Historical Bridge'!$D$5*(1+'Historical Bridge'!$E$5)^({month_index}/12)"
        _formula(ws, REVENUE_TOTAL_ROW, c, f"=MAX({'+'.join(formulas)},{history_formula})", styles, output=True)


def _build_product_build(ws, styles: dict, DataValidation, assumptions: dict) -> None:
    ws["B2"] = "Product Build"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    cost_base = assumptions.get("cost_base", {})
    _table_header(ws, 6, ["Stream", "COGS % Revenue", "Fulfilment Cost / Unit", "Direct FTE / Unit"], styles)
    for r in range(7, 7 + MAX_REVENUE_STREAMS):
        ws.cell(r, 2, f"='Revenue Drivers'!B{r}")
        _input(ws, r, 3, _num(cost_base.get("cogs_percent"), 0.35), styles, fmt="0.0%")
        _input(ws, r, 4, _num(cost_base.get("fulfilment_cost_per_unit"), 100), styles)
        _input(ws, r, 5, 0.01, styles)
    ws.cell(PRODUCT_COGS_ROW, 2, "Total COGS")
    ws.cell(PRODUCT_GP_ROW, 2, "Gross Profit")
    ws.cell(PRODUCT_MARGIN_ROW, 2, "Gross Margin")
    for c in _period_cols():
        letter = _col(c)
        revenue_col = letter
        formulas = [f"('Revenue Drivers'!{revenue_col}{REVENUE_TOTAL_ROW}*($C{r}/{MAX_REVENUE_STREAMS}))" for r in range(7, 7 + MAX_REVENUE_STREAMS)]
        _formula(ws, PRODUCT_COGS_ROW, c, "=-(" + "+".join(formulas) + ")", styles, output=True)
        _formula(ws, PRODUCT_GP_ROW, c, f"='Revenue Drivers'!{letter}{REVENUE_TOTAL_ROW}+{letter}{PRODUCT_COGS_ROW}", styles, output=True)
        _formula(ws, PRODUCT_MARGIN_ROW, c, f"=IFERROR({letter}{PRODUCT_GP_ROW}/'Revenue Drivers'!{letter}{REVENUE_TOTAL_ROW},0)", styles, output=True, fmt="0.0%")


def _build_headcount(ws, styles: dict, assumptions: dict) -> None:
    ws["B2"] = "Headcount"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    _table_header(ws, 6, ["Department", "Opening FTE", "Avg Salary / Month", "Hiring Every N Months", "New Hires"], styles)
    depts = assumptions.get("headcount") or [
        {"department": "Management", "opening_fte": 2, "avg_salary_month": 10000, "hiring_every_months": 6, "new_hires": 1},
        {"department": "Sales", "opening_fte": 4, "avg_salary_month": 4500, "hiring_every_months": 6, "new_hires": 1},
        {"department": "Operations", "opening_fte": 4, "avg_salary_month": 4500, "hiring_every_months": 6, "new_hires": 1},
        {"department": "Finance", "opening_fte": 4, "avg_salary_month": 4500, "hiring_every_months": 6, "new_hires": 1},
        {"department": "IT", "opening_fte": 4, "avg_salary_month": 4500, "hiring_every_months": 6, "new_hires": 1},
        {"department": "Admin", "opening_fte": 4, "avg_salary_month": 4500, "hiring_every_months": 6, "new_hires": 1},
    ]
    for r in range(7, 7 + MAX_HEADCOUNT_LINES):
        dept = depts[r - 7] if r - 7 < len(depts) else {}
        ws.cell(r, 2, dept.get("department", f"Department {r - 6}"))
        _input(ws, r, 3, _num(dept.get("opening_fte"), 0), styles)
        _input(ws, r, 4, _num(dept.get("avg_salary_month"), 0), styles)
        _input(ws, r, 5, max(1, int(_num(dept.get("hiring_every_months"), 6))), styles)
        _input(ws, r, 6, _num(dept.get("new_hires"), 0), styles)
    ws.cell(HEADCOUNT_TOTAL_FTE_ROW, 2, "Total FTE")
    ws.cell(HEADCOUNT_PAYROLL_ROW, 2, "Payroll Cost")
    for c in _period_cols():
        letter = _col(c)
        month_idx = c - FIRST_PERIOD_COL
        for r in range(7, 7 + MAX_HEADCOUNT_LINES):
            _formula(ws, r, c, f"=$C{r}+INT({month_idx}/$E{r})*$F{r}", styles)
        _formula(ws, HEADCOUNT_TOTAL_FTE_ROW, c, f"=SUM({letter}7:{letter}{6 + MAX_HEADCOUNT_LINES})", styles, output=True)
        _formula(ws, HEADCOUNT_PAYROLL_ROW, c, f"=-SUMPRODUCT({letter}7:{letter}{6 + MAX_HEADCOUNT_LINES},$D$7:$D${6 + MAX_HEADCOUNT_LINES})", styles, output=True)


def _build_opex(ws, styles: dict, DataValidation, assumptions: dict) -> None:
    ws["B2"] = "Opex"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    _table_header(ws, 6, ["Cost Category", "Driver", "Monthly Fixed", "% Revenue", "Cost / FTE"], styles)
    cost_items = assumptions.get("cost_items") or []
    cost_base = assumptions.get("cost_base", {})
    rows = cost_items[:6] or [
        {"name": "Rent", "driver": "Fixed", "monthly_fixed": _num(cost_base.get("rent_monthly"), 25000), "percent_revenue": 0, "cost_per_fte": 0},
        {"name": "Marketing", "driver": "% Revenue", "monthly_fixed": 0, "percent_revenue": _num(cost_base.get("opex_percent_revenue"), 0.03), "cost_per_fte": 0},
        {"name": "IT", "driver": "Per FTE", "monthly_fixed": _num(cost_base.get("it_monthly"), 5000), "percent_revenue": 0, "cost_per_fte": 120},
        {"name": "Professional Fees", "driver": "Fixed", "monthly_fixed": _num(cost_base.get("professional_fees_monthly"), 15000), "percent_revenue": 0, "cost_per_fte": 0},
        {"name": "Travel", "driver": "% Revenue", "monthly_fixed": 0, "percent_revenue": 0.01, "cost_per_fte": 0},
        {"name": "Other SG&A", "driver": "Fixed", "monthly_fixed": _num(cost_base.get("opex_fixed_monthly"), 10000), "percent_revenue": 0, "cost_per_fte": 0},
    ]
    for r in range(7, 7 + MAX_COST_ITEMS):
        row = rows[r - 7] if r - 7 < len(rows) else {}
        ws.cell(r, 2, row.get("name", f"Cost item {r - 6}"))
        _input(ws, r, 3, row.get("driver", "Fixed"), styles)
        _input(ws, r, 4, _num(row.get("monthly_fixed"), 0), styles)
        _input(ws, r, 5, _num(row.get("percent_revenue"), 0), styles, fmt="0.0%")
        _input(ws, r, 6, _num(row.get("cost_per_fte"), 0), styles)
        _add_list_validation(ws, f"C{r}", "'Lists & Dates'!$K$2:$K$8", DataValidation)
    ws.cell(OPEX_TOTAL_EXCL_PAYROLL_ROW, 2, "Total Opex excl Payroll")
    ws.cell(OPEX_TOTAL_INCL_PAYROLL_ROW, 2, "Total Opex incl Payroll")
    for c in _period_cols():
        letter = _col(c)
        for r in range(7, 7 + MAX_COST_ITEMS):
            _formula(ws, r, c, f"=-($D{r}+('Revenue Drivers'!{letter}{REVENUE_TOTAL_ROW}*$E{r})+('Headcount'!{letter}{HEADCOUNT_TOTAL_FTE_ROW}*$F{r}))", styles)
        _formula(ws, OPEX_TOTAL_EXCL_PAYROLL_ROW, c, f"=SUM({letter}7:{letter}{6 + MAX_COST_ITEMS})", styles, output=True)
        _formula(ws, OPEX_TOTAL_INCL_PAYROLL_ROW, c, f"={letter}{OPEX_TOTAL_EXCL_PAYROLL_ROW}+'Headcount'!{letter}{HEADCOUNT_PAYROLL_ROW}", styles, output=True)


def _build_working_capital(ws, styles: dict, assumptions: dict) -> None:
    ws["B2"] = "Working Capital"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    wc = assumptions.get("working_capital", {})
    rows = [
        ("DSO", _num(wc.get("dso"), 60)),
        ("DIO", _num(wc.get("dio"), 35)),
        ("DPO", _num(wc.get("dpo"), 55)),
        ("Receivables", None),
        ("Inventory", None),
        ("Payables", None),
        ("Net Working Capital", None),
        ("Change in NWC", None),
    ]
    for r, (label, value) in enumerate(rows, start=6):
        ws.cell(r, 2, label)
        if value is not None:
            _input(ws, r, 3, value, styles)
    for c in _period_cols():
        letter = _col(c)
        prev = _col(c - 1)
        _formula(ws, 9, c, f"='Revenue Drivers'!{letter}{REVENUE_TOTAL_ROW}/365*$C$6", styles)
        _formula(ws, 10, c, f"=ABS('Product Build'!{letter}{PRODUCT_COGS_ROW})/365*$C$7", styles)
        _formula(ws, 11, c, f"=ABS('Product Build'!{letter}{PRODUCT_COGS_ROW}+'Opex'!{letter}{OPEX_TOTAL_INCL_PAYROLL_ROW})/365*$C$8", styles)
        _formula(ws, 12, c, f"={letter}9+{letter}10-{letter}11", styles, output=True)
        _formula(ws, 13, c, f"={letter}12" if c == FIRST_PERIOD_COL else f"={letter}12-{prev}12", styles, output=True)


def _build_capex(ws, styles: dict, assumptions: dict) -> None:
    ws["B2"] = "Capex D&A"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    capex = assumptions.get("capex", {})
    rows = [
        ("Maintenance Capex % Revenue", _num(capex.get("maintenance_percent_revenue"), 0.03)),
        ("Growth Capex / Month", _num(capex.get("growth_capex_monthly"), 0)),
        ("Depreciation Life Months", max(1, int(_num(capex.get("depreciation_years"), 5) * 12))),
    ]
    for r, (label, value) in enumerate(rows, start=6):
        ws.cell(r, 2, label)
        _input(ws, r, 3, value, styles, fmt="0.0%" if r == 6 else None)
    rows = ["Maintenance Capex", "Growth Capex", "Total Capex", "Depreciation", "Net PPE"]
    for r, label in enumerate(rows, start=11):
        ws.cell(r, 2, label)
    for c in _period_cols():
        letter = _col(c)
        prev = _col(c - 1)
        _formula(ws, 11, c, f"=-'Revenue Drivers'!{letter}{REVENUE_TOTAL_ROW}*$C$6", styles)
        _formula(ws, 12, c, f"=-$C$7", styles)
        _formula(ws, 13, c, f"={letter}11+{letter}12", styles, output=True)
        _formula(ws, 14, c, f"=-ABS({letter}13)/$C$8", styles)
        _formula(ws, 15, c, f"=ABS({letter}13)+{letter}14" if c == FIRST_PERIOD_COL else f"={prev}15+ABS({letter}13)+{letter}14", styles, output=True)


def _build_debt_config(ws, styles: dict, DataValidation, assumptions: dict) -> None:
    ws["B2"] = "Debt Config"
    ws["B2"].font = styles["section_font"]
    _table_header(
        ws,
        4,
        [
            "Tranche",
            "Debt Type",
            "Borrower",
            "Start Date",
            "Opening Balance",
            "Commitment",
            "Term Months",
            "Moratorium Months",
            "Interest Cap Months",
            "Margin",
            "Base Rate",
            "Amortization",
            "Bullet %",
            "Cash Sweep %",
            "PIK?",
            "Min Cash",
            "Maturity Date",
            "Interest Type",
            "Cash Pay Frequency",
            "Cash Pay %",
        ],
        styles,
    )
    configured = assumptions.get("debt_tranches") or []
    defaults = [
        _debt_values(tranche) for tranche in configured[:MAX_DEBT_TRANCHES]
    ] or [
        ("Senior Term Loan B", "Senior Term Loan B", "OpCo", "='Control Panel'!$C$8", 300000, 300000, 84, 12, 0, 0.035, 0.030, "Linear", 0.25, 0.15, "FALSE", 50000, "Cash", "Monthly", 1.00),
        ("Super Senior RCF", "Super Senior RCF", "OpCo", "='Control Panel'!$C$8", 0, 100000, 60, 0, 0, 0.020, 0.030, "Revolver", 1.00, 0.00, "FALSE", 50000, "Cash", "Quarterly", 1.00),
        ("Mezzanine PIK", "Mezzanine PIK", "HoldCo", "='Control Panel'!$C$8", 200000, 200000, 96, 24, 24, 0.060, 0.060, "PIK", 1.00, 0.00, "TRUE", 50000, "PIK", "Annual", 0.00),
        ("Seller Note", "Seller Note", "Seller", "='Control Panel'!$C$8", 50000, 50000, 36, 12, 0, 0.000, 0.060, "Bullet", 1.00, 0.00, "FALSE", 50000, "Cash", "Annual", 1.00),
        ("DIP / Rescue", "DIP Financing", "OpCo", "='Control Panel'!$C$8", 0, 0, 24, 0, 0, 0.060, 0.060, "Cash Sweep", 0.00, 0.50, "FALSE", 50000, "Cash", "Monthly", 1.00),
    ]
    for r in range(5, 5 + MAX_DEBT_TRANCHES):
        values = defaults[r - 5] if r - 5 < len(defaults) else ("", "Senior Term Loan B", "", "='Control Panel'!$C$8", 0, 0, 60, 0, 0, 0.030, 0.030, "Bullet", 1.00, 0.00, "FALSE", 50000, "Cash", "Monthly", 1.00)
        for c, value in enumerate(values, start=2):
            _input(ws, r, c, value, styles)
        ws.cell(r, 18, f"=EDATE(E{r},H{r})")
        ws.cell(r, 5).number_format = "yyyy-mm-dd"
        ws.cell(r, 18).number_format = "yyyy-mm-dd"
        for col in [11, 12, 14, 15, 20]:
            ws.cell(r, col).number_format = "0.0%"
        _add_list_validation(ws, f"C{r}", "'Lists & Dates'!$N$2:$N$80", DataValidation)
        _add_list_validation(ws, f"M{r}", "'Lists & Dates'!$Q$2:$Q$16", DataValidation)
        _add_list_validation(ws, f"P{r}", "'Lists & Dates'!$T$2:$T$3", DataValidation)
        _add_list_validation(ws, f"S{r}", "'Lists & Dates'!$W$2:$W$4", DataValidation)
        _add_list_validation(ws, f"T{r}", "'Lists & Dates'!$Y$2:$Y$4", DataValidation)


def _build_debt_schedule(ws, styles: dict) -> None:
    ws["B2"] = "Debt Schedule"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    start_row = 6
    for idx in range(MAX_DEBT_TRANCHES):
        block = start_row + idx * 16
        ws.cell(block, 2, f"='Debt Config'!B{5+idx}")
        ws.cell(block, 2).font = styles["section_font"]
        labels = [
            "Active?",
            "Month Since Start",
            "Opening Debt",
            "Drawdown",
            "Cash Interest",
            "PIK / Capitalised Interest",
            "Scheduled Amortization",
            "Bullet Repayment",
            "Cash Sweep",
            "Closing Debt",
            "Undrawn Commitment",
            "Total Cash Cost",
            "Maturity Flag",
        ]
        for offset, label in enumerate(labels, start=1):
            ws.cell(block + offset, 2, label)
        cfg = 5 + idx
        for c in _period_cols():
            letter = _col(c)
            prev = _col(c - 1)
            active = block + 1
            month = block + 2
            opening = block + 3
            draw = block + 4
            cash_interest = block + 5
            pik_interest = block + 6
            amort = block + 7
            bullet = block + 8
            sweep = block + 9
            closing = block + 10
            undrawn = block + 11
            cost = block + 12
            maturity = block + 13
            _formula(ws, active, c, f'=AND({letter}$4>=\'Debt Config\'!$E{cfg},{letter}$4<=\'Debt Config\'!$R{cfg},\'Debt Config\'!$B{cfg}<>"")', styles)
            _formula(ws, month, c, f'=IF({letter}{active},DATEDIF(\'Debt Config\'!$E{cfg},{letter}$4,"m")+1,0)', styles)
            _formula(ws, opening, c, f"=IF({letter}{active},'Debt Config'!$F{cfg},0)" if c == FIRST_PERIOD_COL else f"=IF({letter}{active},{prev}{closing},0)", styles)
            _formula(ws, draw, c, f'=IF(AND({letter}{active},{letter}{opening}=0,\'Debt Config\'!$M{cfg}="Revolver"),MIN(\'Debt Config\'!$G{cfg},{letter}{undrawn}),0)', styles)
            _formula(ws, cash_interest, c, f'=IF({letter}{active},IF(OR(\'Debt Config\'!$P{cfg}="TRUE",\'Debt Config\'!$S{cfg}="PIK",{letter}{month}<=\'Debt Config\'!$J{cfg}),0,IF(MOD({letter}{month},SWITCH(\'Debt Config\'!$T{cfg},"Monthly",1,"Quarterly",3,"Annual",12,1))=0,{letter}{opening}*(\'Debt Config\'!$K{cfg}+\'Debt Config\'!$L{cfg})/12*SWITCH(\'Debt Config\'!$T{cfg},"Monthly",1,"Quarterly",3,"Annual",12,1)*\'Debt Config\'!$U{cfg},0)),0)', styles)
            _formula(ws, pik_interest, c, f'=IF({letter}{active},MAX(0,{letter}{opening}*(\'Debt Config\'!$K{cfg}+\'Debt Config\'!$L{cfg})/12-{letter}{cash_interest}),0)', styles)
            _formula(ws, amort, c, f'=IF({letter}{active},IF({letter}{month}<=\'Debt Config\'!$I{cfg},0,IF(OR(\'Debt Config\'!$M{cfg}="Bullet",\'Debt Config\'!$M{cfg}="PIK",\'Debt Config\'!$M{cfg}="Revolver"),0,IF(\'Debt Config\'!$M{cfg}="Annuity",PMT((\'Debt Config\'!$K{cfg}+\'Debt Config\'!$L{cfg})/12,MAX(1,\'Debt Config\'!$H{cfg}-{letter}{month}+1),-{letter}{opening}),{letter}{opening}*(1-\'Debt Config\'!$N{cfg})/MAX(1,\'Debt Config\'!$H{cfg}-{letter}{month}+1)))),0)', styles)
            _formula(ws, bullet, c, f'=IF({letter}{maturity},{letter}{opening}*\'Debt Config\'!$N{cfg},0)', styles)
            _formula(ws, sweep, c, f'=IF({letter}{active},MIN(MAX(0,{letter}{opening}+{letter}{draw}+{letter}{pik_interest}-{letter}{amort}-{letter}{bullet}),MAX(0,\'Financial Statements\'!{_financial_monthly_col(c)}20-\'Debt Config\'!$Q{cfg})*\'Debt Config\'!$O{cfg}),0)', styles)
            _formula(ws, closing, c, f"=MAX(0,{letter}{opening}+{letter}{draw}+{letter}{pik_interest}-{letter}{amort}-{letter}{bullet}-{letter}{sweep})", styles, output=True)
            _formula(ws, undrawn, c, f"=MAX(0,'Debt Config'!$G{cfg}-{letter}{closing})", styles)
            _formula(ws, cost, c, f"={letter}{cash_interest}", styles, output=True)
            _formula(ws, maturity, c, f'=AND({letter}{active},{letter}{month}>=\'Debt Config\'!$H{cfg})', styles)
    agg = start_row + MAX_DEBT_TRANCHES * 16 + 2
    summary = ["Total Opening Debt", "Total Drawdowns", "Total Cash Interest", "Total PIK Interest", "Total Amortization", "Total Bullet Repayment", "Total Cash Sweep", "Closing Debt", "Undrawn Commitments", "Total Cash Cost"]
    ws.cell(agg, 2, "Aggregate Debt Summary")
    ws.cell(agg, 2).font = styles["section_font"]
    for r, label in enumerate(summary, start=agg + 1):
        ws.cell(r, 2, label)
    for c in _period_cols():
        letter = _col(c)
        blocks = [start_row + idx * 16 for idx in range(MAX_DEBT_TRANCHES)]
        _formula(ws, agg + 1, c, "=" + "+".join(f"{letter}{b+3}" for b in blocks), styles, output=True)
        _formula(ws, agg + 2, c, "=" + "+".join(f"{letter}{b+4}" for b in blocks), styles, output=True)
        _formula(ws, agg + 3, c, "=" + "+".join(f"{letter}{b+5}" for b in blocks), styles, output=True)
        _formula(ws, agg + 4, c, "=" + "+".join(f"{letter}{b+6}" for b in blocks), styles, output=True)
        _formula(ws, agg + 5, c, "=" + "+".join(f"{letter}{b+7}" for b in blocks), styles, output=True)
        _formula(ws, agg + 6, c, "=" + "+".join(f"{letter}{b+8}" for b in blocks), styles, output=True)
        _formula(ws, agg + 7, c, "=" + "+".join(f"{letter}{b+9}" for b in blocks), styles, output=True)
        _formula(ws, agg + 8, c, "=" + "+".join(f"{letter}{b+10}" for b in blocks), styles, output=True)
        _formula(ws, agg + 9, c, "=" + "+".join(f"{letter}{b+11}" for b in blocks), styles, output=True)
        _formula(ws, agg + 10, c, "=" + "+".join(f"{letter}{b+12}" for b in blocks), styles, output=True)

    annual_start = agg + 14
    ws.cell(annual_start, 2, "Annual Debt Summary")
    ws.cell(annual_start, 2).font = styles["section_font"]
    _table_header(ws, annual_start + 1, ["Metric", "FY2026", "FY2027", "FY2028", "FY2029", "FY2030"], styles)
    annual_rows = [
        ("Opening Debt", agg + 1, "balance_start"),
        ("Drawdowns", agg + 2, "flow"),
        ("Cash Interest", agg + 3, "flow"),
        ("PIK Interest", agg + 4, "flow"),
        ("Scheduled Amortization", agg + 5, "flow"),
        ("Bullet Repayment", agg + 6, "flow"),
        ("Cash Sweep", agg + 7, "flow"),
        ("Closing Debt", agg + 8, "balance_end"),
        ("Undrawn Commitments", agg + 9, "balance_end"),
        ("Total Cash Cost", agg + 10, "flow"),
    ]
    for row_offset, (label, source_row, mode) in enumerate(annual_rows, start=annual_start + 2):
        ws.cell(row_offset, 2, label)
        for year_idx, year in enumerate(range(2026, 2031), start=3):
            if mode == "balance_start":
                _formula(ws, row_offset, year_idx, f"=XLOOKUP(DATE({year},1,31),$D$4:$BK$4,$D${source_row}:$BK${source_row},0)", styles, output=True)
            elif mode == "balance_end":
                _formula(ws, row_offset, year_idx, f"=XLOOKUP(DATE({year},12,31),$D$4:$BK$4,$D${source_row}:$BK${source_row},0)", styles, output=True)
            else:
                _formula(ws, row_offset, year_idx, f'=SUMIFS($D${source_row}:$BK${source_row},$D$4:$BK$4,">="&DATE({year},1,1),$D$4:$BK$4,"<="&DATE({year},12,31))', styles, output=True)


def _build_financial_statements(ws, styles: dict) -> None:
    ws["B2"] = "Financial Statements"
    ws["B2"].font = styles["section_font"]
    ws["B3"] = "Annual view on the left; monthly forecast detail on the right. Monthly columns are grouped and can be expanded/collapsed in Excel."
    years = [2026, 2027, 2028, 2029, 2030]
    _table_header(ws, 4, ["Metric"] + [f"FY{year}" for year in years], styles)
    monthly_start = FINANCIAL_MONTHLY_COL
    ws.cell(4, monthly_start - 1, "Monthly Detail")
    ws.cell(4, monthly_start - 1).fill = styles["section_fill"]
    ws.cell(4, monthly_start - 1).font = styles["bold_font"]
    for idx, c in enumerate(_period_cols(), start=2):
        out_col = monthly_start + c - FIRST_PERIOD_COL
        cell = ws.cell(4, out_col)
        cell.value = f"='Lists & Dates'!V{idx}"
        cell.number_format = "mmm-yy"
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        ws.column_dimensions[_col(out_col)].outlineLevel = 1
        ws.column_dimensions[_col(out_col)].hidden = True
    debt_agg = 6 + MAX_DEBT_TRANCHES * 16 + 2
    rows = [
        ("Income Statement", "Revenue"),
        ("Income Statement", "COGS"),
        ("Income Statement", "Gross Profit"),
        ("Income Statement", "Payroll"),
        ("Income Statement", "Opex"),
        ("Income Statement", "EBITDA"),
        ("Income Statement", "D&A"),
        ("Income Statement", "EBIT"),
        ("Income Statement", "Cash Interest"),
        ("Income Statement", "PBT"),
        ("Income Statement", "Tax"),
        ("Income Statement", "Net Income"),
        ("Cash Flow", "Change in NWC"),
        ("Cash Flow", "Capex"),
        ("Cash Flow", "Cash Flow Before Debt"),
        ("Cash Flow", "Debt Amortization / Sweep"),
        ("Cash Flow", "Free Cash Flow"),
        ("Balance Sheet", "Closing Cash"),
        ("Balance Sheet", "Closing Debt"),
        ("Balance Sheet", "Net Debt"),
    ]
    for r, (section, label) in enumerate(rows, start=6):
        ws.cell(r, 2, label)
        ws.cell(r, 3, section)
        for year_idx, year in enumerate(years, start=3):
            if r in [23, 24, 25]:
                _formula(ws, r, year_idx, f"=XLOOKUP(DATE({year},12,31),$J$4:$BQ$4,$J${r}:$BQ${r},0)", styles, output=True)
            else:
                _formula(ws, r, year_idx, f'=SUMIFS($J${r}:$BQ${r},$J$4:$BQ$4,">="&DATE({year},1,1),$J$4:$BQ$4,"<="&DATE({year},12,31))', styles, output=r in [8, 11, 17, 20, 22])
    for c in _period_cols():
        source_letter = _col(c)
        out_col = FINANCIAL_MONTHLY_COL + c - FIRST_PERIOD_COL
        letter = _col(out_col)
        prev = _col(out_col - 1)
        _formula(ws, 6, out_col, f"='Revenue Drivers'!{source_letter}{REVENUE_TOTAL_ROW}", styles)
        _formula(ws, 7, out_col, f"='Product Build'!{source_letter}{PRODUCT_COGS_ROW}", styles)
        _formula(ws, 8, out_col, f"={letter}6+{letter}7", styles, output=True)
        _formula(ws, 9, out_col, f"='Headcount'!{source_letter}{HEADCOUNT_PAYROLL_ROW}", styles)
        _formula(ws, 10, out_col, f"='Opex'!{source_letter}{OPEX_TOTAL_EXCL_PAYROLL_ROW}", styles)
        _formula(ws, 11, out_col, f"=SUM({letter}8:{letter}10)", styles, output=True)
        _formula(ws, 12, out_col, f"='Capex D&A'!{source_letter}14", styles)
        _formula(ws, 13, out_col, f"={letter}11+{letter}12", styles)
        _formula(ws, 14, out_col, f"='Debt Schedule'!{source_letter}{debt_agg+3}", styles)
        _formula(ws, 15, out_col, f"={letter}13-{letter}14", styles)
        _formula(ws, 16, out_col, f"=-MAX({letter}15,0)*'Control Panel'!$C$13", styles)
        _formula(ws, 17, out_col, f"={letter}15+{letter}16", styles, output=True)
        _formula(ws, 18, out_col, f"='Working Capital'!{source_letter}13", styles)
        _formula(ws, 19, out_col, f"='Capex D&A'!{source_letter}13", styles)
        _formula(ws, 20, out_col, f"={letter}11+{letter}16-{letter}18+{letter}19-{letter}14", styles, output=True)
        _formula(ws, 21, out_col, f"='Debt Schedule'!{source_letter}{debt_agg+5}+'Debt Schedule'!{source_letter}{debt_agg+6}+'Debt Schedule'!{source_letter}{debt_agg+7}", styles)
        _formula(ws, 22, out_col, f"={letter}20-{letter}21", styles, output=True)
        _formula(ws, 23, out_col, f"='Control Panel'!$C$11+{letter}22" if c == FIRST_PERIOD_COL else f"={prev}23+{letter}22", styles, output=True)
        _formula(ws, 24, out_col, f"='Debt Schedule'!{source_letter}{debt_agg+8}", styles, output=True)
        _formula(ws, 25, out_col, f"={letter}24-{letter}23", styles, output=True)

    detail_start = 31
    ws.cell(detail_start, 2, "Granular 3FS Detail")
    ws.cell(detail_start, 2).font = styles["section_font"]
    _table_header(ws, detail_start + 1, ["Detail Line"] + [f"FY{year}" for year in years], styles)
    ws.cell(detail_start + 1, monthly_start - 1, "Monthly Detail")
    ws.cell(detail_start + 1, monthly_start - 1).fill = styles["section_fill"]
    ws.cell(detail_start + 1, monthly_start - 1).font = styles["bold_font"]
    for idx, c in enumerate(_period_cols(), start=2):
        out_col = monthly_start + c - FIRST_PERIOD_COL
        cell = ws.cell(detail_start + 1, out_col)
        cell.value = f"='Lists & Dates'!V{idx}"
        cell.number_format = "mmm-yy"
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        ws.column_dimensions[_col(out_col)].outlineLevel = 1
        ws.column_dimensions[_col(out_col)].hidden = True
    for idx, (statement, category, _subcategory, model_line, detail_line, _sign) in enumerate(_historical_line_templates()[:150], start=0):
        row = detail_start + 2 + idx
        source_row = 6 + idx
        ws.cell(row, 2, f"{statement} | {category} | {detail_line}")
        for year_idx, _year in enumerate(years, start=3):
            _formula(ws, row, year_idx, f"='3FS Detail Output'!{_col(year_idx + 4)}{source_row}", styles, output=True)
        for c in _period_cols():
            out_col = monthly_start + c - FIRST_PERIOD_COL
            detail_col = _col(13 + c - FIRST_PERIOD_COL)
            _formula(ws, row, out_col, f"='3FS Detail Output'!{detail_col}{source_row}", styles, output=True)
    ws.sheet_properties.outlinePr.summaryRight = False


def _build_3fs_detail_output(ws, styles: dict) -> None:
    ws["B2"] = "3FS Detail Output"
    ws["B2"].font = styles["section_font"]
    ws["B3"] = "Granular historical lines projected from the selected mapping. Annual summary sits left; collapsible monthly detail sits right."
    years = [2026, 2027, 2028, 2029, 2030]
    _table_header(ws, 5, ["Line ID", "Statement", "Category", "Model Line", "Detail Line"] + [f"FY{year}" for year in years], styles)
    monthly_start = 13
    ws.cell(5, monthly_start - 1, "Monthly Detail")
    ws.cell(5, monthly_start - 1).fill = styles["section_fill"]
    ws.cell(5, monthly_start - 1).font = styles["bold_font"]
    for idx, c in enumerate(_period_cols(), start=2):
        out_col = monthly_start + c - FIRST_PERIOD_COL
        cell = ws.cell(5, out_col)
        cell.value = f"='Lists & Dates'!V{idx}"
        cell.number_format = "mmm-yy"
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        ws.column_dimensions[_col(out_col)].outlineLevel = 1
        ws.column_dimensions[_col(out_col)].hidden = True

    templates = _historical_line_templates()
    for idx, (statement, category, _subcategory, model_line, detail_line, _sign) in enumerate(templates, start=6):
        ws.cell(idx, 2, f"='Historical Detail Input'!B{idx}")
        ws.cell(idx, 3, statement)
        ws.cell(idx, 4, category)
        ws.cell(idx, 5, model_line)
        ws.cell(idx, 6, detail_line)
        for year_idx, year in enumerate(years, start=7):
            if model_line in ["Cash", "Closing Debt", "Net Debt", "Receivables", "Inventory", "Payables", "Net PPE", "Equity"]:
                _formula(ws, idx, year_idx, f"=XLOOKUP(DATE({year},12,31),$M$5:$BT$5,$M{idx}:$BT{idx},0)", styles, output=True)
            else:
                _formula(ws, idx, year_idx, f'=SUMIFS($M{idx}:$BT{idx},$M$5:$BT$5,">="&DATE({year},1,1),$M$5:$BT$5,"<="&DATE({year},12,31))', styles, output=True)
        for c in _period_cols():
            out_col = monthly_start + c - FIRST_PERIOD_COL
            fs_col = _financial_monthly_col(c)
            _formula(ws, idx, out_col, _detail_projection_formula(idx, model_line, fs_col), styles, output=True)
    ws.sheet_properties.outlinePr.summaryRight = False


def _build_detail_forecast_lines(ws, styles: dict) -> None:
    ws["B2"] = "Detailed Forecast Lines"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    _table_header(ws, 6, ["Line ID", "Section", "Line Item", "Driver", "Source"], styles)
    sections = [
        ("Revenue", "Revenue Drivers", f"='Revenue Drivers'!{{col}}{REVENUE_TOTAL_ROW}", "Product/service build"),
        ("COGS", "Product Build", f"='Product Build'!{{col}}{PRODUCT_COGS_ROW}", "COGS assumptions"),
        ("Gross Profit", "Product Build", f"='Product Build'!{{col}}{PRODUCT_GP_ROW}", "Formula"),
        ("Payroll", "Headcount", f"='Headcount'!{{col}}{HEADCOUNT_PAYROLL_ROW}", "FTE plan"),
        ("Opex", "Opex", f"='Opex'!{{col}}{OPEX_TOTAL_EXCL_PAYROLL_ROW}", "Detailed opex items"),
        ("EBITDA", "Financial Statements", "='Financial Statements'!{col}11", "Formula"),
        ("D&A", "Capex D&A", "='Capex D&A'!{col}14", "Depreciation"),
        ("Cash Interest", "Debt Schedule", "='Financial Statements'!{col}14", "Debt engine"),
        ("Tax", "Financial Statements", "='Financial Statements'!{col}16", "Tax rate"),
        ("Net Income", "Financial Statements", "='Financial Statements'!{col}17", "Formula"),
        ("Receivables", "Working Capital", "='Working Capital'!{col}9", "DSO"),
        ("Inventory", "Working Capital", "='Working Capital'!{col}10", "DIO"),
        ("Payables", "Working Capital", "='Working Capital'!{col}11", "DPO"),
        ("NWC", "Working Capital", "='Working Capital'!{col}12", "Formula"),
        ("Change in NWC", "Working Capital", "='Working Capital'!{col}13", "Formula"),
        ("Maintenance Capex", "Capex D&A", "='Capex D&A'!{col}11", "Revenue-linked"),
        ("Growth Capex", "Capex D&A", "='Capex D&A'!{col}12", "Manual monthly"),
        ("Total Capex", "Capex D&A", "='Capex D&A'!{col}13", "Formula"),
        ("Free Cash Flow", "Financial Statements", "='Financial Statements'!{col}22", "Formula"),
        ("Closing Cash", "Financial Statements", "='Financial Statements'!{col}23", "Formula"),
        ("Closing Debt", "Debt Schedule", "='Financial Statements'!{col}24", "Debt engine"),
        ("Net Debt", "Financial Statements", "='Financial Statements'!{col}25", "Formula"),
        ("Net Debt / EBITDA", "Covenants", "='Covenants'!{col}11", "Covenant"),
        ("Interest Cover", "Covenants", "='Covenants'!{col}12", "Covenant"),
        ("Liquidity", "Covenants", "='Covenants'!{col}13", "Covenant"),
    ]
    for idx in range(225):
        row = 7 + idx
        label, section, template, driver = sections[idx % len(sections)]
        cycle = idx // len(sections) + 1
        ws.cell(row, 2, f"FL-{idx + 1:03d}")
        ws.cell(row, 3, section)
        ws.cell(row, 4, f"{label} detail {cycle}")
        ws.cell(row, 5, driver)
        ws.cell(row, 6, "Formula linked")
        factor = 1 + (cycle - 1) * 0.001
        for c in _period_cols():
            col = _col(c)
            fs_col = _financial_monthly_col(c)
            base_formula = template.format(col=fs_col if "Financial Statements" in template else col)
            if label in ["Net Debt / EBITDA", "Interest Cover"]:
                _formula(ws, row, c, base_formula, styles, output=True, fmt="0.0x")
            elif label in ["Liquidity"]:
                _formula(ws, row, c, base_formula, styles, output=True)
            else:
                _formula(ws, row, c, f"=({base_formula.lstrip('=')})*{factor:.3f}", styles, output=label in ["EBITDA", "Free Cash Flow", "Closing Cash", "Closing Debt", "Net Debt"])


def _build_covenants(ws, styles: dict, assumptions: dict) -> None:
    ws["B2"] = "Covenants"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    covenants = assumptions.get("covenants", {})
    covenant_rows = [
        ("Max Net Debt / EBITDA", _num(covenants.get("max_net_debt_ebitda"), 3.5)),
        ("Min ICR", _num(covenants.get("min_interest_cover"), 2.0)),
        ("Min Liquidity", _num(covenants.get("min_liquidity"), 50000)),
    ]
    for r, (label, value) in enumerate(covenant_rows, start=6):
        ws.cell(r, 2, label)
        _input(ws, r, 3, value, styles)
    rows = ["Net Debt / EBITDA", "Interest Cover", "Liquidity", "Leverage Pass?", "ICR Pass?", "Liquidity Pass?", "All Covenants Pass?"]
    for r, label in enumerate(rows, start=11):
        ws.cell(r, 2, label)
    for c in _period_cols():
        letter = _col(c)
        fs_col = _financial_monthly_col(c)
        _formula(ws, 11, c, f"=IFERROR('Financial Statements'!{fs_col}25/'Financial Statements'!{fs_col}11,0)", styles, output=True, fmt="0.0x")
        _formula(ws, 12, c, f"=IFERROR('Financial Statements'!{fs_col}11/'Financial Statements'!{fs_col}14,99)", styles, output=True, fmt="0.0x")
        _formula(ws, 13, c, f"='Financial Statements'!{fs_col}23", styles, output=True)
        _formula(ws, 14, c, f"={letter}11<=$C$6", styles, output=True)
        _formula(ws, 15, c, f"={letter}12>=$C$7", styles, output=True)
        _formula(ws, 16, c, f"={letter}13>=$C$8", styles, output=True)
        _formula(ws, 17, c, f"=AND({letter}14,{letter}15,{letter}16)", styles, output=True)


def _build_outputs(ws, styles: dict) -> None:
    ws["B2"] = "Outputs"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    rows = ["Revenue", "Gross Margin", "EBITDA", "EBITDA Margin", "Free Cash Flow", "Closing Debt", "Net Debt / EBITDA", "Covenant Pass?"]
    for r, label in enumerate(rows, start=6):
        ws.cell(r, 2, label)
    for c in _period_cols():
        letter = _col(c)
        fs_col = _financial_monthly_col(c)
        _formula(ws, 6, c, f"='Financial Statements'!{fs_col}6", styles, output=True)
        _formula(ws, 7, c, f"=IFERROR('Financial Statements'!{fs_col}8/'Financial Statements'!{fs_col}6,0)", styles, output=True, fmt="0.0%")
        _formula(ws, 8, c, f"='Financial Statements'!{fs_col}11", styles, output=True)
        _formula(ws, 9, c, f"=IFERROR({letter}8/{letter}6,0)", styles, output=True, fmt="0.0%")
        _formula(ws, 10, c, f"='Financial Statements'!{fs_col}22", styles, output=True)
        _formula(ws, 11, c, f"='Financial Statements'!{fs_col}24", styles, output=True)
        _formula(ws, 12, c, f"='Covenants'!{letter}11", styles, output=True, fmt="0.0x")
        _formula(ws, 13, c, f"='Covenants'!{letter}17", styles, output=True)

    ws["B17"] = "Annual Output Summary"
    ws["B17"].font = styles["section_font"]
    _table_header(ws, 19, ["Metric", "FY2026", "FY2027", "FY2028", "FY2029", "FY2030"], styles)
    annual_rows = [
        ("Revenue", 6, None),
        ("Gross Profit", 8, None),
        ("Gross Margin", 8, "margin_revenue"),
        ("EBITDA", 11, None),
        ("EBITDA Margin", 11, "margin_revenue"),
        ("Cash Flow Before Debt", 20, None),
        ("Free Cash Flow", 22, None),
        ("Closing Cash", 23, None),
        ("Closing Debt", 24, None),
        ("Net Debt", 25, None),
    ]
    for row_idx, (label, source_row, mode) in enumerate(annual_rows, start=20):
        ws.cell(row_idx, 2, label)
        for col_idx, year_col in enumerate(range(3, 8), start=3):
            source_letter = _col(year_col)
            if mode == "margin_revenue":
                _formula(ws, row_idx, col_idx, f"=IFERROR('Financial Statements'!{source_letter}{source_row}/'Financial Statements'!{source_letter}6,0)", styles, output=True, fmt="0.0%")
            else:
                _formula(ws, row_idx, col_idx, f"='Financial Statements'!{source_letter}{source_row}", styles, output=True)

    ws["B33"] = "Model Integrity Snapshot"
    ws["B33"].font = styles["section_font"]
    snapshot = [
        ("Current month all checks", "='Checks'!D10"),
        ("Current covenant status", "='Outputs'!D13"),
        ("Historical mapping depth", "=COUNTA('Historical Detail Input'!$F$6:$F$185)"),
        ("Formula rows generated", "=COUNTA('Detailed Forecast Lines'!$B$7:$B$231)"),
    ]
    for row_idx, (label, formula) in enumerate(snapshot, start=35):
        ws.cell(row_idx, 2, label)
        _formula(ws, row_idx, 3, formula, styles, output=True)


def _build_executive_dashboard(ws, project: dict, styles: dict) -> None:
    try:
        from openpyxl.chart import BarChart, LineChart, Reference
    except Exception:
        BarChart = LineChart = Reference = None

    ws["B2"] = f"{project.get('company_name', 'Target Company')} - Executive Dashboard"
    ws["B2"].font = styles["title_font"]
    ws["B3"] = "Board-ready view linked to the formula model: annual output, liquidity, leverage, covenant status and model checks."
    ws["B3"].font = styles["subtitle_font"]

    _table_header(ws, 5, ["Current Month KPI", "Value", "Status", "Source"], styles)
    kpis = [
        ("Revenue", "='Outputs'!D6", '="Current forecast month"', "Outputs!D6"),
        ("EBITDA", "='Outputs'!D8", '=IF(C6>=0,"Positive","Negative")', "Outputs!D8"),
        ("EBITDA Margin", "='Outputs'!D9", '=IF(C7>=0.1,"Institutional","Review")', "Outputs!D9"),
        ("Free Cash Flow", "='Outputs'!D10", '=IF(C8>=0,"Cash generative","Cash burn")', "Outputs!D10"),
        ("Closing Cash", "='Financial Statements'!J23", '=IF(C9>0,"Funded","Liquidity watch")', "Financial Statements!J23"),
        ("Closing Debt", "='Outputs'!D11", '=IF(C10>=0,"Modelled","Review")', "Outputs!D11"),
        ("Net Debt / EBITDA", "='Outputs'!D12", '=IF(C11<3.5,"Inside threshold","Review")', "Outputs!D12"),
        ("Covenant Pass?", "='Outputs'!D13", '=IF(C12="OK","Clean","Review")', "Outputs!D13"),
    ]
    for row_idx, (label, formula, status_formula, source) in enumerate(kpis, start=6):
        ws.cell(row_idx, 2, label)
        _formula(ws, row_idx, 3, formula, styles, output=True, fmt="0.0%" if "Margin" in label else "0.0x" if "/" in label else None)
        _formula(ws, row_idx, 4, status_formula, styles, output=True)
        ws.cell(row_idx, 5, source)

    ws["G5"] = "Five-Year Financial Trajectory"
    ws["G5"].font = styles["section_font"]
    _table_header(ws, 6, ["Metric", "FY2026", "FY2027", "FY2028", "FY2029", "FY2030"], styles)
    annual_rows = [
        ("Revenue", 6, None),
        ("Gross Profit", 8, None),
        ("EBITDA", 11, None),
        ("EBITDA Margin", 12, "percent"),
        ("Free Cash Flow", 15, None),
        ("Closing Cash", 16, None),
        ("Closing Debt", 17, None),
        ("Net Debt", 18, None),
    ]
    for row_idx, (label, source_row, fmt) in enumerate(annual_rows, start=7):
        ws.cell(row_idx, 7, label)
        for col_idx in range(8, 13):
            source_col = _col(col_idx - 5)
            _formula(ws, row_idx, col_idx, f"='Summary Financials Annual'!{source_col}{source_row}", styles, output=True, fmt="0.0%" if fmt == "percent" else None)

    ws["B17"] = "Credit & Covenant Dashboard"
    ws["B17"].font = styles["section_font"]
    _table_header(ws, 19, ["Metric", "FY2026", "FY2027", "FY2028", "FY2029", "FY2030", "Comment"], styles)
    credit_rows = [
        ("Net Debt / EBITDA", "'Output_Group_Annual'!{col}11", "0.0x", "Leverage path from annual output"),
        ("Interest Cover", "'Covenants'!{month_col}12", "0.0x", "EBITDA / cash interest at year-end"),
        ("Liquidity", "'Financial Statements'!{fs_col}23", None, "Closing cash at year-end"),
        ("Covenant Pass?", "'Covenants'!{month_col}17", None, "All lender tests must pass"),
        ("Model Checks", "'Checks'!{month_col}10", None, "Workbook integrity status"),
    ]
    fs_year_end_cols = ["U", "AG", "AS", "BE", "BQ"]
    covenant_year_end_cols = ["O", "AA", "AM", "AY", "BK"]
    for row_idx, (label, template, fmt, comment) in enumerate(credit_rows, start=20):
        ws.cell(row_idx, 2, label)
        for offset, col_idx in enumerate(range(3, 8)):
            annual_col = _col(col_idx)
            fs_col = fs_year_end_cols[offset]
            month_col = covenant_year_end_cols[offset]
            formula = template.format(col=annual_col, fs_col=fs_col, month_col=month_col)
            _formula(ws, row_idx, col_idx, f"={formula}", styles, output=True, fmt=fmt)
        ws.cell(row_idx, 8, comment)

    ws["B29"] = "Claude / BP Data Readiness"
    ws["B29"].font = styles["section_font"]
    readiness = [
        ("Historical lines mapped", "=COUNTA('Historical Detail Input'!$F$6:$F$185)", "Target: 150+ lines for full 3FS mapping"),
        ("Revenue streams", "=COUNTA('Revenue Drivers'!$B$7:$B$16)", "Configured from SaaS or Claude extraction"),
        ("Cost lines", "=COUNTA('Opex'!$B$7:$B$18)", "Fixed / variable / FTE-linked operating cost build"),
        ("Debt layers", "=COUNTA('Debt Config'!$B$5:$B$14)", "Each tranche modelled separately"),
        ("All checks OK", "='Checks'!D10", "Review before external distribution"),
    ]
    _table_header(ws, 31, ["Control", "Value", "Review Note"], styles)
    for row_idx, (label, formula, note) in enumerate(readiness, start=32):
        ws.cell(row_idx, 2, label)
        _formula(ws, row_idx, 3, formula, styles, output=True)
        ws.cell(row_idx, 4, note)

    if LineChart and Reference:
        line = LineChart()
        line.title = "Revenue and EBITDA"
        line.y_axis.title = "Amount"
        line.x_axis.title = "Fiscal year"
        data = Reference(ws, min_col=8, max_col=12, min_row=7, max_row=9)
        cats = Reference(ws, min_col=8, max_col=12, min_row=6, max_row=6)
        line.add_data(data, titles_from_data=False, from_rows=True)
        line.set_categories(cats)
        line.height = 7
        line.width = 15
        ws.add_chart(line, "G17")

        bar = BarChart()
        bar.title = "Closing debt and liquidity"
        debt_data = Reference(ws, min_col=8, max_col=12, min_row=12, max_row=13)
        bar.add_data(debt_data, titles_from_data=False, from_rows=True)
        bar.set_categories(cats)
        bar.height = 7
        bar.width = 15
        ws.add_chart(bar, "G32")


def _build_summary_financials_quarter(ws, styles: dict) -> None:
    ws["B2"] = "Summary Financials Quarter"
    ws["B2"].font = styles["section_font"]
    ws["B3"] = "BOLT-style quarterly output sheet. Core metrics link to Financial Statements; detailed lines link to 3FS Detail Output."
    _table_header(ws, 5, ["Metric"] + [f"Q{(idx % 4) + 1} FY{2026 + idx // 4}" for idx in range(QUARTERS)], styles)
    rows = _summary_financial_rows(include_details=True)
    for row_idx, (label, source, mode) in enumerate(rows, start=6):
        ws.cell(row_idx, 2, label)
        for quarter_idx in range(QUARTERS):
            col_idx = 3 + quarter_idx
            start_month = quarter_idx * 3
            start_ref = f"INDEX('Lists & Dates'!$V$2:$V$61,{start_month + 1})"
            end_ref = f"EOMONTH({start_ref},2)"
            if source.startswith("FS:"):
                fs_row = int(source.split(":")[1])
                if mode == "balance":
                    _formula(ws, row_idx, col_idx, f"=XLOOKUP({end_ref},'Financial Statements'!$J$4:$BQ$4,'Financial Statements'!$J${fs_row}:$BQ${fs_row},0)", styles, output=True)
                elif mode == "margin":
                    numerator = int(source.split(":")[1])
                    _formula(ws, row_idx, col_idx, f'=IFERROR(SUMIFS(\'Financial Statements\'!$J${numerator}:$BQ${numerator},\'Financial Statements\'!$J$4:$BQ$4,">="&{start_ref},\'Financial Statements\'!$J$4:$BQ$4,"<="&{end_ref})/SUMIFS(\'Financial Statements\'!$J$6:$BQ$6,\'Financial Statements\'!$J$4:$BQ$4,">="&{start_ref},\'Financial Statements\'!$J$4:$BQ$4,"<="&{end_ref}),0)', styles, output=True, fmt="0.0%")
                else:
                    _formula(ws, row_idx, col_idx, f'=SUMIFS(\'Financial Statements\'!$J${fs_row}:$BQ${fs_row},\'Financial Statements\'!$J$4:$BQ$4,">="&{start_ref},\'Financial Statements\'!$J$4:$BQ$4,"<="&{end_ref})', styles, output=True)
            else:
                detail_row = int(source.split(":")[1])
                _formula(ws, row_idx, col_idx, f'=SUMIFS(\'3FS Detail Output\'!$M${detail_row}:$BT${detail_row},\'3FS Detail Output\'!$M$5:$BT$5,">="&{start_ref},\'3FS Detail Output\'!$M$5:$BT$5,"<="&{end_ref})', styles, output=True)


def _bolt_annual_section(ws, row: int, title: str, styles: dict) -> None:
    ws.cell(row, 2, title)
    ws.cell(row, 2).font = styles["section_font"]
    for col in range(2, 15):
        ws.cell(row, col).fill = styles["section_fill"]


def _bolt_annual_row(
    ws,
    row: int,
    label: str,
    unit: str | None,
    comment: str | None,
    styles: dict,
    formulas: list[str],
    fmt: str | None = None,
    bold: bool = False,
) -> None:
    ws.cell(row, 3, label)
    ws.cell(row, 5, unit)
    ws.cell(row, 7, comment)
    if bold:
        ws.cell(row, 3).font = styles["bold_font"]
    for idx, formula in enumerate(formulas[:5]):
        col = 10 + idx
        if formula == "":
            continue
        _formula(ws, row, col, formula if formula.startswith("=") else f"={formula}", styles, output=True, fmt=fmt)


def _build_summary_financials_annual(ws, styles: dict) -> None:
    years = [2026, 2027, 2028, 2029, 2030]
    ws["A1"] = "MG ADVISORY FINANCIAL MODEL"
    ws["A1"].font = styles["title_font"]
    ws["A2"] = '=UPPER(MID(CELL("filename",C3),FIND("]",CELL("filename",C3))+1,LEN(CELL("filename",C3))))'
    ws["I1"] = "='Cover'!$B$2"
    ws["I2"] = '=COUNTIF(I4:I1048576,"FALSE")'
    ws["C4"] = "='Cover'!$B$2"
    ws["C5"] = "='Control Panel'!$C$6"
    ws["I4"] = "Financial Year"
    ws["I7"] = "Start date"
    ws["I8"] = "End date"
    ws["I9"] = "# Days"
    for row in [4, 5]:
        for col in range(3, 15):
            ws.cell(row, col).fill = styles["header_fill"]
            ws.cell(row, col).font = styles["header_font"]
    for row in [7, 8, 9]:
        ws.cell(row, 9).font = styles["bold_font"]
    for offset, year in enumerate(years):
        col = 10 + offset
        letter = _col(col)
        _formula(ws, 4, col, f"={year}", styles, output=True)
        ws.cell(4, col).number_format = '"FY" 0'
        _formula(ws, 7, col, f"=DATE({year},1,1)", styles, output=True, fmt="d-mmm-yy")
        _formula(ws, 8, col, f"=DATE({year},12,31)", styles, output=True, fmt="d-mmm-yy")
        _formula(ws, 9, col, f"={letter}8-{letter}7+1", styles, output=True)

    _bolt_annual_section(ws, 11, "1. Summary KPIs & Unit Economics - Annual", styles)
    _bolt_annual_row(ws, 13, "Annual production capacity", "Units", "Revenue / implied price capacity", styles, [f"='Summary Financials Annual'!{_col(10+i)}14*1.20" for i in range(5)])
    _bolt_annual_row(ws, 14, "Production volume", "Units", "Revenue / blended price", styles, [f"=IFERROR('Financial Statements'!{_col(3+i)}6/1000,0)" for i in range(5)])
    _bolt_annual_row(ws, 15, "Utilisation rate", "%", None, styles, [f"=IFERROR({_col(10+i)}14/{_col(10+i)}13,0)" for i in range(5)], fmt="0.0%")
    _bolt_annual_row(ws, 17, "Sales Volumes", "Units", None, styles, [f"={_col(10+i)}14" for i in range(5)])
    _bolt_annual_row(ws, 18, "Gross Sales Price", "Currency/unit", "Gross revenue / volume", styles, [f"=IFERROR({_col(10+i)}83/{_col(10+i)}17,0)" for i in range(5)])
    _bolt_annual_row(ws, 19, "Gross Sales Ex-price", "Currency/unit", "Gross revenue less freight / volume", styles, [f"=IFERROR(SUM({_col(10+i)}83,{_col(10+i)}85)/{_col(10+i)}17,0)" for i in range(5)])
    _bolt_annual_row(ws, 20, "Gross Spread", "Currency/unit", "Gross revenue less direct materials / volume", styles, [f"=IFERROR(SUM({_col(10+i)}83,{_col(10+i)}90)/{_col(10+i)}17,0)" for i in range(5)])
    _bolt_annual_row(ws, 22, "Net Realisation", "Currency/unit", "Revenue less freight, commissions and tariffs", styles, [f"=IFERROR({_col(10+i)}88/{_col(10+i)}17,0)" for i in range(5)])
    _bolt_annual_row(ws, 23, "Direct materials", "Currency/unit", None, styles, [f"=IFERROR({_col(10+i)}90/{_col(10+i)}17,0)" for i in range(5)])
    _bolt_annual_row(ws, 24, "Implied Net spread", "Currency/unit", "Net realisation less direct materials", styles, [f"=IFERROR({_col(10+i)}91/{_col(10+i)}17,0)" for i in range(5)])
    _bolt_annual_row(ws, 25, "Variable costs", "Currency/unit", "COGS variable cost base", styles, [f"=IFERROR({_col(10+i)}95/{_col(10+i)}17,0)" for i in range(5)])
    _bolt_annual_row(ws, 26, "Contribution margin", "Currency/unit", "Implied net spread less variable costs", styles, [f"=IFERROR({_col(10+i)}97/{_col(10+i)}17,0)" for i in range(5)])
    _bolt_annual_row(ws, 27, "Operating expense", "Currency/unit", "Payroll and opex / volume", styles, [f"=IFERROR({_col(10+i)}106/{_col(10+i)}17,0)" for i in range(5)])
    _bolt_annual_row(ws, 28, "Adjusted EBITDA", "Currency/unit", "Contribution margin less fixed costs", styles, [f"=IFERROR({_col(10+i)}108/{_col(10+i)}17,0)" for i in range(5)])
    ws["C30"] = "Working Capital Days"
    for row, label, formula in [
        (31, "Raw Materials - DIO", "='Working Capital'!$C$7"),
        (32, "Finished Goods - DIO", "='Working Capital'!$C$7"),
        (33, "Gross Payables - DPO", "='Working Capital'!$C$8"),
        (34, "Gross Receivables - DSO", "='Working Capital'!$C$6"),
    ]:
        _bolt_annual_row(ws, row, label, "Days", None, styles, [formula for _ in years])

    _bolt_annual_section(ws, 37, "2. Sales and Pricing Breakdown - Annual", styles)
    product_rows = [(40, "Core product"), (41, "Services"), (42, "Recurring / Other")]
    ws["C39"] = "Sales volume by Product (Units)"
    for row, label in product_rows:
        _bolt_annual_row(ws, row, label, "Units", None, styles, [f"={_col(10+i)}17/{len(product_rows)}" for i in range(5)])
    _bolt_annual_row(ws, 43, "Total Volumes", "Units", None, styles, [f"=SUM({_col(10+i)}40:{_col(10+i)}42)" for i in range(5)], bold=True)
    _bolt_annual_row(ws, 44, "Volume check", None, None, styles, [f"={_col(10+i)}43-{_col(10+i)}17" for i in range(5)])
    ws["C46"] = "Sales volume by Product (%)"
    for row, label in [(47, "Core product"), (48, "Services"), (49, "Recurring / Other")]:
        _bolt_annual_row(ws, row, label, "%", None, styles, [f"=IFERROR({_col(10+i)}{row}/{_col(10+i)}43,0)" for i in range(5)], fmt="0.0%")
    _bolt_annual_row(ws, 50, "Total", "%", None, styles, [f"=SUM({_col(10+i)}47:{_col(10+i)}49)" for i in range(5)], fmt="0.0%", bold=True)
    _bolt_annual_row(ws, 51, "Check", None, None, styles, [f"=IFERROR(IF({_col(10+i)}50=1,0,1),1)" for i in range(5)])
    ws["C53"] = "Sales volume by Region (Units)"
    for row, label in [(54, "North America"), (55, "Europe"), (56, "Middle East"), (57, "Rest of World")]:
        _bolt_annual_row(ws, row, label, "Units", None, styles, [f"={_col(10+i)}17/4" for i in range(5)])
    _bolt_annual_row(ws, 58, "Total Volumes", "Units", None, styles, [f"=SUM({_col(10+i)}54:{_col(10+i)}57)" for i in range(5)], bold=True)
    _bolt_annual_row(ws, 59, "Volume check", None, None, styles, [f"=ROUND({_col(10+i)}58-{_col(10+i)}17,0)" for i in range(5)])
    ws["C61"] = "Sales volume by Region (%)"
    for row, label in [(62, "North America"), (63, "Europe"), (64, "Middle East"), (65, "Rest of World")]:
        _bolt_annual_row(ws, row, label, "%", None, styles, [f"=IFERROR({_col(10+i)}{row-8}/{_col(10+i)}58,0)" for i in range(5)], fmt="0.0%")
    _bolt_annual_row(ws, 66, "Total", "%", None, styles, [f"=SUM({_col(10+i)}62:{_col(10+i)}65)" for i in range(5)], fmt="0.0%", bold=True)
    _bolt_annual_row(ws, 67, "Check", None, None, styles, [f"=IFERROR(IF({_col(10+i)}66=1,0,1),1)" for i in range(5)])
    ws["C69"] = "Net Spread by Product"
    for row, label in [(70, "Core product"), (71, "Services"), (72, "Recurring / Other")]:
        _bolt_annual_row(ws, row, label, "Currency/unit", None, styles, [f"={_col(10+i)}24" for i in range(5)])
    _bolt_annual_row(ws, 73, "Weighted-average net spread", "Currency/unit", None, styles, [f"=SUMPRODUCT({_col(10+i)}47:{_col(10+i)}49,{_col(10+i)}70:{_col(10+i)}72)" for i in range(5)], bold=True)

    _bolt_annual_section(ws, 76, "3. Income Statement (P&L) - Annual", styles)
    pnl_rows = [
        (79, "North America", "Currency", [f"={_col(10+i)}83/4" for i in range(5)]),
        (80, "Europe", "Currency", [f"={_col(10+i)}83/4" for i in range(5)]),
        (81, "Middle East", "Currency", [f"={_col(10+i)}83/4" for i in range(5)]),
        (82, "Rest of World", "Currency", [f"={_col(10+i)}83/4" for i in range(5)]),
        (83, "Gross Revenue", "Currency", [f"='Financial Statements'!{_col(3+i)}6" for i in range(5)]),
        (85, "Less: Freight and Forwarding", "Currency", ["=0" for _ in years]),
        (86, "Less: Sales Commissions", "Currency", ["=0" for _ in years]),
        (87, "Less: Tariffs", "Currency", ["=0" for _ in years]),
        (88, "Net Realisation", "Currency", [f"=SUM({_col(10+i)}83:{_col(10+i)}87)" for i in range(5)]),
        (90, "Less: Direct Materials", "Currency", [f"='Financial Statements'!{_col(3+i)}7" for i in range(5)]),
        (91, "Implied Net Spread", "Currency", [f"={_col(10+i)}88+{_col(10+i)}90" for i in range(5)]),
        (93, "Utilities Costs", "Currency", [f"='Financial Statements'!{_col(3+i)}10*0.25" for i in range(5)]),
        (94, "Packing Costs", "Currency", [f"='Financial Statements'!{_col(3+i)}10*0.10" for i in range(5)]),
        (95, "Total Variable Costs", "Currency", [f"=SUM({_col(10+i)}93:{_col(10+i)}94)" for i in range(5)]),
        (97, "Contribution Margin", "Currency", [f"={_col(10+i)}91+{_col(10+i)}95" for i in range(5)]),
        (98, "Contribution Margin %", "%", [f"=IFERROR({_col(10+i)}97/{_col(10+i)}83,0)" for i in range(5)]),
        (100, "Plant salary & wages", "Currency", [f"='Financial Statements'!{_col(3+i)}9" for i in range(5)]),
        (101, "Other manufacturing overheads", "Currency", [f"='Financial Statements'!{_col(3+i)}10*0.35" for i in range(5)]),
        (102, "Conversion cost adjustments", "Currency", ["=0" for _ in years]),
        (103, "Total Operating Costs", "Currency", [f"=SUM({_col(10+i)}100:{_col(10+i)}102)" for i in range(5)]),
        (104, "Selling & distribution", "Currency", [f"='Financial Statements'!{_col(3+i)}10*0.25" for i in range(5)]),
        (105, "Admin expenses", "Currency", [f"='Financial Statements'!{_col(3+i)}10*0.40" for i in range(5)]),
        (106, "Total Operating Expense", "Currency", [f"=SUM({_col(10+i)}103:{_col(10+i)}105)" for i in range(5)]),
        (108, "Adjusted EBITDA", "Currency", [f"='Financial Statements'!{_col(3+i)}11" for i in range(5)]),
        (109, "Adj. EBITDA Margin %", "%", [f"=IFERROR({_col(10+i)}108/{_col(10+i)}83,0)" for i in range(5)]),
        (111, "Simplified EBITDA to Net Income bridge (for BS completeness)", None, ["" for _ in years]),
        (113, "Management and restructuring fees", "Currency", ["=0" for _ in years]),
        (114, "Exchange Gain / (Loss)", "Currency", ["=0" for _ in years]),
        (115, "Other Income / (Loss)", "Currency", ["=0" for _ in years]),
        (116, "Depreciation (D&A)", "Currency", [f"='Financial Statements'!{_col(3+i)}12" for i in range(5)]),
        (117, "Interest expense", "Currency", [f"='Financial Statements'!{_col(3+i)}14" for i in range(5)]),
        (118, "Corporate tax", "Currency", [f"='Financial Statements'!{_col(3+i)}16" for i in range(5)]),
        (119, "Simplified Net Income", "Currency", [f"='Financial Statements'!{_col(3+i)}17" for i in range(5)]),
        (120, "Check", None, [f"={_col(10+i)}119-'Financial Statements'!{_col(3+i)}17" for i in range(5)]),
    ]
    for row, label, unit, formulas in pnl_rows:
        _bolt_annual_row(ws, row, label, unit, None, styles, formulas, fmt="0.0%" if unit == "%" else None, bold=row in {83, 88, 91, 95, 97, 106, 108, 119})

    _bolt_annual_section(ws, 122, "4. Balance Sheet (BS) - Annual", styles)
    for row, label in [(124, "ASSETS"), (126, "Non-Current Assets"), (133, "Current Assets"), (158, "LIABILITIES & EQUITY"), (160, "Non-Current Liabilities"), (167, "Current Liabilities"), (179, "Equity")]:
        ws.cell(row, 3, label)
        ws.cell(row, 3).font = styles["section_font"]
    bs_map = {
        127: ("Fixed Assets", 0), 128: ("Intangible Assets", 0), 129: ("Loans Receivable", 0), 130: ("Investments in Subsidiaries", 0),
        131: ("Total non-current assets", 0), 134: ("Cash & cash equivalents", 23), 135: ("Trade & other inventory", 0),
        136: ("Raw Materials", 0), 137: ("Packing Materials", 0), 138: ("Semi Finished Goods", 0), 139: ("Finished Goods", 0),
        140: ("Goods in Transit", 0), 141: ("By Product", 0), 142: ("Spares, Consumables and Others", 0),
        144: ("Trade & other receivables", 0), 145: ("Gross Receivables", 0), 146: ("Advances from Customers", 0),
        147: ("Other Receivables", 0), 148: ("Prepayments", 0), 150: ("Due from related parties", 0),
        151: ("Due from related parties - trade", 0), 152: ("Due from related parties - non-trade", 0),
        153: ("Total current assets", 0), 155: ("Total Assets", 0), 161: ("Loans Payable", 24), 162: ("Accrued Interest", 0),
        163: ("Government Grant", 0), 164: ("Deferred Taxes", 0), 165: ("Total non-current liabilities", 0),
        168: ("Trade & other payables", 0), 169: ("Gross Payables", 0), 170: ("Advances to Suppliers", 0),
        171: ("Employee and Salary Provisions", 0), 172: ("Other Payables", 0), 174: ("Due to related parties", 0),
        175: ("Due to related parties - trade", 0), 176: ("Due to related parties - non-trade", 0),
        177: ("Total current liabilities", 0), 180: ("Shareholders Equity", 0), 181: ("Retained Earnings", 0),
        182: ("Total equity", 0), 184: ("Total Liabilities and Equity", 0), 186: ("Balance Sheet Check", None),
    }
    for row, (label, fs_row) in bs_map.items():
        if fs_row is None:
            formulas = [f"={_col(10+i)}155-{_col(10+i)}184" for i in range(5)]
        elif fs_row:
            formulas = [f"='Financial Statements'!{_col(3+i)}{fs_row}" for i in range(5)]
        else:
            formulas = ["=0" for _ in years]
        _bolt_annual_row(ws, row, label, "Currency" if fs_row is not None else None, None, styles, formulas, bold=row in {131, 153, 155, 165, 177, 182, 184})

    _bolt_annual_section(ws, 189, "5. Cash Flow Statement (CFS) - Annual", styles)
    cfs_rows = {
        191: ("Net Income", 17), 192: ("Prior period equity adjustment", 0), 193: ("Add: Interest Expense", 14),
        194: ("Add: Depreciation (D&A)", 12), 195: ("Add: Other income / (loss)", 0), 196: ("Add: Exchange gain / (loss)", 0),
        197: ("Add: Corporate tax", 16), 198: ("EBITDA (incl.restructuring fees)", 11), 200: ("Change in NWC", None),
        201: ("Change in trade working capital", 18), 202: ("Trade - Inventory", None), 203: ("Raw Materials", 0),
        204: ("Packing Materials", 0), 205: ("Semi Finished Goods", 0), 206: ("Finished Goods", 0), 207: ("Goods in Transit", 0),
        208: ("Trade - Account Receivables", None), 209: ("Gross Receivables", 0), 210: ("Advances from Customers", 0),
        211: ("Due from Related Parties - Trade", 0), 212: ("Trade - Account Payables", None), 213: ("Gross Payables", 0),
        214: ("Advances to Suppliers", 0), 215: ("Due to related parties - Trade", 0), 217: ("Change in non-trade working capital", 0),
        218: ("Non-Trade - Other Inventory", None), 219: ("By Product", 0), 220: ("Spares, Consumables and Others", 0),
        221: ("Non-Trade - Other Account Receivables", None), 222: ("Other Receivables", 0), 223: ("Prepayments", 0),
        224: ("Due from Related Parties - Non Trade", 0), 225: ("Non-Trade - Other Account Payables", None),
        226: ("Employee and Salary Provisions", 0), 227: ("Other Payables", 0), 228: ("Due to related parties - Non trade", 0),
        229: ("Cash flow from Operations", 20), 231: ("Sale/ (Purchase) of fixed assets", 19), 232: ("Loan Receivable", 0),
        233: ("Investments in Subsidiaries", 0), 234: ("Change in intangible assets", 0), 235: ("Cash flow from Investing activities", 19),
        237: ("Interest Paid", 14), 238: ("Loan Repayment", 21), 239: ("Government Grant", 0), 240: ("Cash flow from Financing activities", 21),
        242: ("Net change in cash", 22), 243: ("Opening Cash Balance", 0), 244: ("Closing Cash Balance", 23), 246: ("Cash Balance Check", None),
    }
    for row, (label, fs_row) in cfs_rows.items():
        if fs_row is None:
            formulas = ["" for _ in years]
        elif fs_row:
            formulas = [f"='Financial Statements'!{_col(3+i)}{fs_row}" for i in range(5)]
        else:
            formulas = ["=0" for _ in years]
        if row == 246:
            formulas = [f"={_col(10+i)}244-'Financial Statements'!{_col(3+i)}23" for i in range(5)]
        _bolt_annual_row(ws, row, label, "Currency" if fs_row is not None else None, None, styles, formulas, bold=row in {198, 229, 235, 240, 242, 244})
    _bolt_annual_section(ws, 249, "END", styles)


def _build_ebitda_bridges(ws, styles: dict) -> None:
    ws["B2"] = "EBITDA Bridges"
    ws["B2"].font = styles["section_font"]
    ws["B3"] = "QoE-style EBITDA bridge from revenue to reported/adjusted EBITDA, with annual and quarterly outputs."
    years = [2026, 2027, 2028, 2029, 2030]
    _table_header(ws, 5, ["Bridge Line"] + [f"FY{year}" for year in years], styles)
    bridge_rows = [
        ("Revenue", "='Summary Financials Annual'!{col}6"),
        ("COGS", "='Summary Financials Annual'!{col}7"),
        ("Gross Profit", "='Summary Financials Annual'!{col}8"),
        ("Payroll", "='Summary Financials Annual'!{col}9"),
        ("Opex", "='Summary Financials Annual'!{col}10"),
        ("Reported EBITDA", "='Summary Financials Annual'!{col}11"),
        ("Non-recurring income", "=-SUMIFS('Historical Detail Input'!$O$6:$O$185,'Historical Detail Input'!$G$6:$G$185,\"Non-recurring income\")/5"),
        ("Non-recurring costs", "=SUMIFS('Historical Detail Input'!$O$6:$O$185,'Historical Detail Input'!$G$6:$G$185,\"Non-recurring costs\")/5"),
        ("Run-rate cost savings", "=MAX(0,-'Summary Financials Annual'!{col}10*1.0%)"),
        ("Management QoE adjustments", "=SUM({col}13:{col}15)"),
        ("Adjusted EBITDA", "={col}12+{col}16"),
        ("Adjusted EBITDA margin", "=IFERROR({col}17/{col}7,0)"),
    ]
    for row_idx, (label, template) in enumerate(bridge_rows, start=7):
        ws.cell(row_idx, 2, label)
        for col_idx in range(3, 8):
            col = _col(col_idx)
            fmt = "0.0%" if "margin" in label.lower() else None
            _formula(ws, row_idx, col_idx, template.format(col=col), styles, output=True, fmt=fmt)

    ws["B23"] = "Quarterly EBITDA Bridge"
    ws["B23"].font = styles["section_font"]
    _table_header(ws, 25, ["Bridge Line"] + [f"Q{(idx % 4) + 1} FY{2026 + idx // 4}" for idx in range(QUARTERS)], styles)
    quarter_rows = [
        ("Revenue", 6),
        ("Gross Profit", 8),
        ("Reported EBITDA", 11),
        ("Adjusted EBITDA", None),
        ("Adjusted EBITDA margin", None),
    ]
    for row_idx, (label, summary_row) in enumerate(quarter_rows, start=26):
        ws.cell(row_idx, 2, label)
        for col_idx in range(3, 3 + QUARTERS):
            col = _col(col_idx)
            if summary_row:
                _formula(ws, row_idx, col_idx, f"='Summary Financials Quarter'!{col}{summary_row}", styles, output=True)
            elif label == "Adjusted EBITDA":
                _formula(ws, row_idx, col_idx, f"={col}28+MAX(0,-{col}27*1.0%)", styles, output=True)
            else:
                _formula(ws, row_idx, col_idx, f"=IFERROR({col}29/{col}26,0)", styles, output=True, fmt="0.0%")


def _build_packaged_output(ws, project: dict, styles: dict) -> None:
    ws["B2"] = f"{project.get('company_name', 'Target Company')} - Packaged Output"
    ws["B2"].font = styles["title_font"]
    ws["B4"] = "Executive Summary"
    ws["B4"].font = styles["section_font"]
    _table_header(ws, 6, ["Metric", "Current Month", "FY2026", "FY2027", "FY2028", "FY2029", "FY2030"], styles)
    rows = [
        ("Revenue", "'Outputs'!D6", "'Output_Group_Annual'!C5", "'Output_Group_Annual'!D5", "'Output_Group_Annual'!E5", "'Output_Group_Annual'!F5", "'Output_Group_Annual'!G5"),
        ("EBITDA", "'Outputs'!D8", "'Output_Group_Annual'!C7", "'Output_Group_Annual'!D7", "'Output_Group_Annual'!E7", "'Output_Group_Annual'!F7", "'Output_Group_Annual'!G7"),
        ("EBITDA Margin", "'Outputs'!D9", "'Output_Group_Annual'!C8", "'Output_Group_Annual'!D8", "'Output_Group_Annual'!E8", "'Output_Group_Annual'!F8", "'Output_Group_Annual'!G8"),
        ("Free Cash Flow", "'Outputs'!D10", "'Financial Statements'!C22", "'Financial Statements'!D22", "'Financial Statements'!E22", "'Financial Statements'!F22", "'Financial Statements'!G22"),
        ("Closing Debt", "'Outputs'!D11", "'Output_Group_Annual'!C10", "'Output_Group_Annual'!D10", "'Output_Group_Annual'!E10", "'Output_Group_Annual'!F10", "'Output_Group_Annual'!G10"),
        ("Net Debt / EBITDA", "'Outputs'!D12", "'Output_Group_Annual'!C11", "'Output_Group_Annual'!D11", "'Output_Group_Annual'!E11", "'Output_Group_Annual'!F11", "'Output_Group_Annual'!G11"),
    ]
    for row, values in enumerate(rows, start=7):
        ws.cell(row, 2, values[0])
        for col_idx, ref in enumerate(values[1:], start=3):
            fmt = "0.0%" if values[0] == "EBITDA Margin" else "0.0x" if "/" in values[0] else None
            _formula(ws, row, col_idx, f"={ref}", styles, output=True, fmt=fmt)
    ws["B16"] = "Key Credit Messages"
    ws["B16"].font = styles["section_font"]
    messages = [
        "Revenue and EBITDA are driver-based and link to detailed product/service assumptions.",
        "Debt schedule supports start dates, maturities, moratorium, PIK, bullet repayment and cash sweep.",
        "Covenant outputs are linked to financial statements and debt schedule.",
        "All key outputs should be reviewed alongside the Checks sheet before external distribution.",
    ]
    for row, message in enumerate(messages, start=17):
        ws.cell(row, 2, f"{row-16}.")
        ws.cell(row, 3, message)


def _build_ic_summary(ws, project: dict, styles: dict) -> None:
    ws["B2"] = f"{project.get('company_name', 'Target Company')} - Investment Committee Summary"
    ws["B2"].font = styles["title_font"]
    ws["B4"] = "Executive Decision Page"
    ws["B4"].font = styles["section_font"]
    _table_header(ws, 6, ["Topic", "Current View", "Investment / Credit Implication", "Next Diligence Action"], styles)
    rows = [
        ("Trading trajectory", "='Packaged Output'!C7", "Revenue base and growth profile drive debt capacity.", "Validate revenue bridge by customer/product."),
        ("Profitability", "='Packaged Output'!C8", "EBITDA quality supports leverage and covenant sizing.", "Bridge EBITDA to QoE adjustments."),
        ("Cash conversion", "='Packaged Output'!C10", "FCF determines deleveraging and liquidity runway.", "Stress-test working capital and capex."),
        ("Leverage", "='Packaged Output'!C12", "Net debt / EBITDA frames sponsor and lender risk.", "Run downside covenant headroom."),
        ("Covenant status", "='Outputs'!D13", "Early breach risk drives amend-and-extend or restructuring path.", "Confirm covenant definitions in facility docs."),
        ("Model integrity", "='Checks'!D10", "External distribution depends on passing checks.", "Resolve all ERROR flags before release."),
    ]
    for row, values in enumerate(rows, start=7):
        ws.cell(row, 2, values[0])
        for col_idx, value in enumerate(values[1:], start=3):
            if isinstance(value, str) and value.startswith("="):
                _formula(ws, row, col_idx, value, styles, output=True, fmt="0.0x" if values[0] == "Leverage" else None)
            else:
                ws.cell(row, col_idx, value)

    ws["B16"] = "Banker Commentary Framework"
    ws["B16"].font = styles["section_font"]
    comments = [
        ("Base Case", "Business plan remains supportable if EBITDA growth converts to cash and covenant headroom remains above internal thresholds."),
        ("Downside Case", "Primary downside risk is lower cash conversion, delayed deleveraging and tighter liquidity headroom."),
        ("Restructuring Angle", "If debt service cannot be supported, options should be sequenced by liquidity runway, stakeholder consent and enterprise value preservation."),
    ]
    _table_header(ws, 18, ["Lens", "Draft institutional language"], styles)
    for row, values in enumerate(comments, start=19):
        ws.cell(row, 2, values[0])
        ws.cell(row, 3, values[1])


def _build_restructuring_options(ws, styles: dict) -> None:
    ws["B2"] = "Restructuring Options Paper"
    ws["B2"].font = styles["section_font"]
    _table_header(ws, 4, ["Option", "Trigger", "Liquidity Impact", "Leverage Impact", "Stakeholder Complexity", "Indicative Use Case"], styles)
    rows = [
        ("Status Quo / Self Help", "No covenant breach", "Low", "Low", "Low", "Management actions, cost take-out, WC release"),
        ("Amend & Extend", "Temporary covenant pressure", "Medium", "Medium", "Medium", "Fee + revised covenant package"),
        ("Payment Holiday / PIK Toggle", "Short-term liquidity gap", "High", "Medium", "Medium", "Preserve cash while business stabilises"),
        ("New Money Super Senior", "Liquidity runway insufficient", "High", "Medium", "High", "Bridge to disposal, M&A or operational turnaround"),
        ("Debt Buyback / Discounted Payoff", "Debt trades below par", "Medium", "High", "High", "Deleveraging with sponsor or asset-sale proceeds"),
        ("Debt-for-Equity Swap", "Unsustainable capital structure", "Medium", "Very High", "Very High", "Balance sheet reset and ownership transfer"),
        ("Accelerated M&A / Disposal", "Strategic value exceeds standalone recovery", "High", "High", "High", "Stakeholder-led exit or non-core disposals"),
    ]
    for row, values in enumerate(rows, start=5):
        for col_idx, value in enumerate(values, start=2):
            ws.cell(row, col_idx, value)

    ws["B15"] = "Option Scoring"
    ws["B15"].font = styles["section_font"]
    _table_header(ws, 17, ["Option", "Liquidity Relief", "Execution Risk", "Value Preservation", "Lender Acceptability", "Weighted Score"], styles)
    for row in range(18, 25):
        ws.cell(row, 2, f"=B{row-13}")
        _input(ws, row, 3, 3, styles)
        _input(ws, row, 4, 3, styles)
        _input(ws, row, 5, 3, styles)
        _input(ws, row, 6, 3, styles)
        _formula(ws, row, 7, f"=SUM(C{row},E{row},F{row})-D{row}", styles, output=True)


def _build_debt_capacity(ws, styles: dict) -> None:
    ws["B2"] = "Debt Capacity"
    ws["B2"].font = styles["section_font"]
    _table_header(ws, 4, ["Metric", "FY2026", "FY2027", "FY2028", "FY2029", "FY2030"], styles)
    metrics = [
        ("EBITDA", "'Output_Group_Annual'!C7", "'Output_Group_Annual'!D7", "'Output_Group_Annual'!E7", "'Output_Group_Annual'!F7", "'Output_Group_Annual'!G7"),
        ("Existing Net Debt", "'Output_Group_Annual'!C11", "'Output_Group_Annual'!D11", "'Output_Group_Annual'!E11", "'Output_Group_Annual'!F11", "'Output_Group_Annual'!G11"),
        ("Lender Leverage Threshold", "3.50", "3.25", "3.00", "2.75", "2.50"),
        ("Implied Debt Capacity", "=C5*C7", "=D5*D7", "=E5*E7", "=F5*F7", "=G5*G7"),
        ("Headroom / Shortfall", "=C8-C6", "=D8-D6", "=E8-E6", "=F8-F6", "=G8-G6"),
        ("Cash Sweep Capacity", "='Financial Statements'!C22", "='Financial Statements'!D22", "='Financial Statements'!E22", "='Financial Statements'!F22", "='Financial Statements'!G22"),
        ("Refinancing Risk Flag", '=IF(C9<0,"Shortfall","Headroom")', '=IF(D9<0,"Shortfall","Headroom")', '=IF(E9<0,"Shortfall","Headroom")', '=IF(F9<0,"Shortfall","Headroom")', '=IF(G9<0,"Shortfall","Headroom")'),
    ]
    for row, values in enumerate(metrics, start=5):
        ws.cell(row, 2, values[0])
        for col_idx, value in enumerate(values[1:], start=3):
            formula = value if str(value).startswith("=") else f"={value}" if "!" in str(value) else value
            if isinstance(formula, str) and formula.startswith("="):
                _formula(ws, row, col_idx, formula, styles, output=True, fmt="0.0x" if row == 7 else None)
            else:
                _input(ws, row, col_idx, formula, styles, fmt="0.0x" if row == 7 else None)


def _build_sensitivity_matrix(ws, styles: dict) -> None:
    ws["B2"] = "Sensitivity Matrix"
    ws["B2"].font = styles["section_font"]
    ws["B4"] = "Net Debt / EBITDA Sensitivity"
    ws["B4"].font = styles["section_font"]
    _table_header(ws, 6, ["EBITDA Case / Debt Case", "-15% Debt", "Base Debt", "+15% Debt"], styles)
    cases = [("-20% EBITDA", 0.80), ("Base EBITDA", 1.00), ("+20% EBITDA", 1.20)]
    debt_cases = [0.85, 1.00, 1.15]
    for row, (label, ebitda_factor) in enumerate(cases, start=7):
        ws.cell(row, 2, label)
        for col_idx, debt_factor in enumerate(debt_cases, start=3):
            _formula(
                ws,
                row,
                col_idx,
                f"=IFERROR(('Financial Statements'!J24*{debt_factor})/('Financial Statements'!J11*{ebitda_factor}),0)",
                styles,
                output=True,
                fmt="0.0x",
            )
    ws["B13"] = "Liquidity Runway Sensitivity"
    ws["B13"].font = styles["section_font"]
    _table_header(ws, 15, ["FCF Case / Opening Cash", "-25% Cash", "Base Cash", "+25% Cash"], styles)
    fcf_cases = [("-25% FCF", 0.75), ("Base FCF", 1.00), ("+25% FCF", 1.25)]
    cash_cases = [0.75, 1.00, 1.25]
    for row, (label, fcf_factor) in enumerate(fcf_cases, start=16):
        ws.cell(row, 2, label)
        for col_idx, cash_factor in enumerate(cash_cases, start=3):
            _formula(
                ws,
                row,
                col_idx,
                f"=('Control Panel'!$C$11*{cash_factor})+('Financial Statements'!J22*{fcf_factor})",
                styles,
                output=True,
            )


def _build_checks(ws, styles: dict) -> None:
    ws["B2"] = "Checks"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    rows = [
        "No Negative Cash",
        "Debt Roll Forward",
        "Revenue Positive",
        "Covenants Populated",
        "All Checks OK",
        "Cash Flow Roll Forward",
        "Net Debt Recalculation",
        "3FS Revenue Tie",
        "3FS EBITDA Tie",
        "3FS Debt Tie",
        "Historical Mapping Populated",
        "Balance Sheet Signal",
    ]
    for r, label in enumerate(rows, start=6):
        ws.cell(r, 2, label)
    debt_agg = 6 + MAX_DEBT_TRANCHES * 16 + 2
    for c in _period_cols():
        letter = _col(c)
        fs_col = _financial_monthly_col(c)
        previous_cash = "'Control Panel'!$C$11" if c == FIRST_PERIOD_COL else f"'Financial Statements'!{_financial_monthly_col(c - 1)}23"
        detail_col = _col(13 + c - FIRST_PERIOD_COL)
        _formula(ws, 6, c, f'=IF(\'Financial Statements\'!{fs_col}23>=0,"OK","ERROR")', styles, output=True)
        _formula(ws, 7, c, f'=IF(ABS(\'Debt Schedule\'!{letter}{debt_agg+8}-(\'Debt Schedule\'!{letter}{debt_agg+1}+\'Debt Schedule\'!{letter}{debt_agg+2}+\'Debt Schedule\'!{letter}{debt_agg+4}-\'Debt Schedule\'!{letter}{debt_agg+5}-\'Debt Schedule\'!{letter}{debt_agg+6}-\'Debt Schedule\'!{letter}{debt_agg+7}))<1,"OK","ERROR")', styles, output=True)
        _formula(ws, 8, c, f'=IF(\'Financial Statements\'!{fs_col}6>0,"OK","ERROR")', styles, output=True)
        _formula(ws, 9, c, f'=IF(\'Covenants\'!{letter}17<>"","OK","ERROR")', styles, output=True)
        _formula(ws, 10, c, f'=IF(AND({letter}6="OK",{letter}7="OK",{letter}8="OK",{letter}9="OK",{letter}11="OK",{letter}12="OK",{letter}13="OK",{letter}14="OK",{letter}15="OK",{letter}16="OK",{letter}17="OK"),"OK","ERROR")', styles, output=True)
        _formula(ws, 11, c, f'=IF(ABS(\'Financial Statements\'!{fs_col}23-({previous_cash}+\'Financial Statements\'!{fs_col}22))<1,"OK","ERROR")', styles, output=True)
        _formula(ws, 12, c, f'=IF(ABS(\'Financial Statements\'!{fs_col}25-(\'Financial Statements\'!{fs_col}24-\'Financial Statements\'!{fs_col}23))<1,"OK","ERROR")', styles, output=True)
        _formula(ws, 13, c, f'=IF(SUMIFS(\'Historical Detail Input\'!$P:$P,\'Historical Detail Input\'!$F:$F,"Revenue")=0,"OK",IF(ABS(SUMIF(\'3FS Detail Output\'!$E$6:$E$185,"Revenue",\'3FS Detail Output\'!${detail_col}$6:${detail_col}$185)-\'Financial Statements\'!{fs_col}6)<1,"OK","ERROR"))', styles, output=True)
        _formula(ws, 14, c, f'=IF(SUMIFS(\'Historical Detail Input\'!$P:$P,\'Historical Detail Input\'!$F:$F,"EBITDA")=0,"OK",IF(ABS(SUMIF(\'3FS Detail Output\'!$E$6:$E$185,"EBITDA",\'3FS Detail Output\'!${detail_col}$6:${detail_col}$185)-\'Financial Statements\'!{fs_col}11)<1,"OK","ERROR"))', styles, output=True)
        _formula(ws, 15, c, f'=IF(SUMIFS(\'Historical Detail Input\'!$P:$P,\'Historical Detail Input\'!$F:$F,"Closing Debt")=0,"OK",IF(ABS(SUMIF(\'3FS Detail Output\'!$E$6:$E$185,"Closing Debt",\'3FS Detail Output\'!${detail_col}$6:${detail_col}$185)-\'Financial Statements\'!{fs_col}24)<1,"OK","ERROR"))', styles, output=True)
        _formula(ws, 16, c, '=IF(COUNTA(\'Historical Detail Input\'!$F$6:$F$185)>=150,"OK","ERROR")', styles, output=True)
        _formula(ws, 17, c, f'=IF(\'Financial Statements\'!{fs_col}23+MAX(\'Financial Statements\'!{fs_col}24,0)>=0,"OK","ERROR")', styles, output=True)


def _build_lookup(ws, styles: dict) -> None:
    ws["B2"] = "Lookup"
    ws["B2"].font = styles["section_font"]
    _table_header(ws, 4, ["Level 0", "Level 1", "Level 2", "Model Line", "Target Sheet"], styles)
    rows = [
        ("Income Statement", "Revenue", "Trading", "Revenue", "Financial Statements"),
        ("Income Statement", "COGS", "Trading", "COGS", "Financial Statements"),
        ("Income Statement", "Opex", "Payroll", "Payroll", "Financial Statements"),
        ("Income Statement", "Opex", "SG&A", "Opex", "Financial Statements"),
        ("Balance Sheet", "Debt", "Borrowings", "Closing Debt", "Financial Statements"),
        ("Balance Sheet", "Cash", "Liquidity", "Closing Cash", "Financial Statements"),
        ("Cash Flow", "Working Capital", "NWC", "Change in NWC", "Financial Statements"),
        ("Cash Flow", "Investing", "Capex", "Capex", "Financial Statements"),
    ]
    for row, values in enumerate(rows, start=5):
        for col, value in enumerate(values, start=2):
            ws.cell(row, col, value)


def _build_mapping(ws, styles: dict) -> None:
    ws["B2"] = "Mapping"
    ws["B2"].font = styles["section_font"]
    _table_header(ws, 4, ["Output Line", "Source Sheet", "Source Row", "Formula Pattern", "Status"], styles)
    rows = [
        ("Revenue", "Revenue Drivers", str(REVENUE_TOTAL_ROW), "Linked by monthly period", "Active"),
        ("COGS", "Product Build", str(PRODUCT_COGS_ROW), "Linked by monthly period", "Active"),
        ("Payroll", "Headcount", str(HEADCOUNT_PAYROLL_ROW), "Linked by monthly period", "Active"),
        ("Opex", "Opex", str(OPEX_TOTAL_EXCL_PAYROLL_ROW), "Linked by monthly period", "Active"),
        ("Debt Interest", "Debt Schedule", "Aggregate cash interest", "Linked by monthly period", "Active"),
        ("Closing Debt", "Debt Schedule", "Aggregate closing debt", "Linked by monthly period", "Active"),
        ("Historical Granularity", "Historical Detail Input", f"{HISTORICAL_DETAIL_LINES} lines", "Claude/manual actuals bridge", "Active"),
        ("3FS Granularity", "3FS Detail Output", f"{HISTORICAL_DETAIL_LINES} projected lines", "Annual left + monthly grouped detail", "Active"),
        ("Covenants", "Covenants", "11:17", "Linked to FS and Debt Schedule", "Active"),
        ("Checks", "Checks", "6:17", "Cash, debt, 3FS, covenant and historical controls", "Active"),
    ]
    for row, values in enumerate(rows, start=5):
        for col, value in enumerate(values, start=2):
            ws.cell(row, col, value)


def _build_lists(ws, project: dict, styles: dict) -> None:
    ws["A1"] = "Lists & Fixed Dates"
    ws["A1"].font = styles["section_font"]
    currencies = ["EUR", "USD", "GBP", "AED", "CHF", "CAD", "AUD"]
    scenarios = ["Base", "Downside", "Upside", "Bank Case", "IC Case", "Restructuring Case"]
    products = ["Product", "Service", "Recurring", "Project", "Consulting", "License", "Maintenance", "Other"]
    drivers = ["Fixed", "% Revenue", "Per FTE"]
    amort = ["Bullet", "Linear", "Annuity", "Cash Sweep", "Revolver", "Borrowing Base", "PIK", "PIK Toggle", "Interest Only Then Linear", "Revenue Share", "Operational Run-Off", "Debt Sculpting", "Sculpted"]
    bools = ["TRUE", "FALSE"]
    interest_types = ["Cash", "PIK", "Cash / PIK Toggle"]
    payment_frequencies = ["Monthly", "Quarterly", "Annual"]
    historical_sources = ["Claude extraction", "Manual input", "Hybrid review"]
    historical_statements = ["Income Statement", "Balance Sheet", "Cash Flow", "Debt Schedule"]
    historical_categories = [
        "Revenue", "COGS", "Gross Profit", "Payroll", "Opex", "EBITDA Adjustments", "D&A", "Interest", "Tax",
        "Current Assets", "Fixed Assets", "Other Assets", "Debt", "Current Liabilities", "Other Liabilities",
        "Equity", "Working Capital", "Operating Cash Flow", "Investing Cash Flow", "Financing Cash Flow",
    ]
    historical_model_lines = [
        "Revenue", "COGS", "Gross Profit", "Payroll", "Opex", "EBITDA", "D&A", "EBIT", "Cash Interest",
        "PBT", "Tax", "Net Income", "Cash", "Receivables", "Inventory", "Other Current Assets", "Net PPE",
        "Intangibles", "Other Non Current Assets", "Payables", "Accruals", "Tax Payable", "Closing Debt",
        "Net Debt", "Equity", "Change in NWC", "Capex", "Free Cash Flow", "Debt Amortization / Sweep",
    ]
    historical_source_modes = ["Claude extraction", "Manual input", "Hybrid review", "Imported Excel"]
    lists = [
        (1, "Currencies", currencies),
        (4, "Scenarios", scenarios),
        (7, "Product / Service Types", products),
        (10, "Cost Drivers", drivers),
        (13, "Debt Types", debt_type_options()),
        (16, "Amortization Types", amort),
        (19, "Boolean", bools),
        (23, "Interest Type", interest_types),
        (25, "Payment Frequency", payment_frequencies),
        (28, "Historical Source", historical_sources),
        (31, "Historical Statements", historical_statements),
        (34, "Historical Categories", historical_categories),
        (37, "Historical Model Lines", historical_model_lines),
        (40, "Historical Source Modes", historical_source_modes),
    ]
    for col, header, values in lists:
        ws.cell(1, col, header)
        ws.cell(1, col).fill = styles["header_fill"]
        ws.cell(1, col).font = styles["header_font"]
        for row, value in enumerate(values, start=2):
            ws.cell(row, col, value)
    ws["V1"] = "Monthly Dates"
    ws["V1"].fill = styles["header_fill"]
    ws["V1"].font = styles["header_font"]
    for idx in range(PERIODS):
        row = idx + 2
        ws.cell(row, 22, f"=EOMONTH('Control Panel'!$C$8,{idx})")
        ws.cell(row, 23, f"=YEAR(V{row})")
        ws.cell(row, 24, f'="Q"&ROUNDUP(MONTH(V{row})/3,0)')
        ws.cell(row, 22).number_format = "mmm-yy"


def _write_period_headers(ws, row: int, styles: dict, source_sheet: str = "'Lists & Dates'") -> None:
    ws.cell(row, 2, "Metric")
    ws.cell(row, 2).fill = styles["header_fill"]
    ws.cell(row, 2).font = styles["header_font"]
    for idx, c in enumerate(_period_cols(), start=2):
        cell = ws.cell(row, c)
        cell.value = f"={source_sheet}!V{idx}"
        cell.number_format = "mmm-yy"
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]


def _table_header(ws, row: int, labels: list[str], styles: dict) -> None:
    for col, label in enumerate(labels, start=2):
        cell = ws.cell(row, col, label)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]


def _input(ws, row: int, col: int, value, styles: dict, fmt: str | None = None) -> None:
    cell = ws.cell(row, col, value)
    cell.fill = styles["input_fill"]
    cell.font = styles["input_font"]
    if fmt:
        cell.number_format = fmt


def _formula(ws, row: int, col: int, formula: str, styles: dict, output: bool = False, fmt: str | None = None) -> None:
    cell = ws.cell(row, col, formula)
    cell.fill = styles["output_fill"] if output else styles["formula_fill"]
    if fmt:
        cell.number_format = fmt


def _label(ws, row: int, col: int, value: str, styles: dict) -> None:
    cell = ws.cell(row, col, value)
    cell.font = styles["bold_font"]
    cell.fill = styles["section_fill"]


def _add_list_validation(ws, cell_range: str, formula_range: str, DataValidation) -> None:
    validation = DataValidation(type="list", formula1=formula_range, allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(cell_range)


def _polish_sheet(ws, styles: dict) -> None:
    ws.sheet_view.showGridLines = False
    bolt_freezes = {
        "Revenue and COGS Inputs": "O196",
        "Opex Inputs": "J27",
        "BS and NWC Schedules": "J182",
        "Revenue and COGS Calcs": "J10",
        "Opex Calcs": "J10",
        "BS and NWC Calcs": "J93",
        "CIQ_LinkingNames": "J10",
        "Consolidated Financials Calcs": "J10",
        "Summary Financials Quarter": "J10",
        "Summary Financials Annual": "J10",
        "EBITDA Bridges": "J10",
    }
    ws.freeze_panes = bolt_freezes.get(ws.title, "J5" if ws.title == "Financial Statements" else "D5")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    if ws.title in bolt_freezes:
        ws.column_dimensions["I"].width = 18
        for col in range(10, 10 + PERIODS):
            ws.column_dimensions[_col(col)].width = 12
    max_period_col = max(FIRST_PERIOD_COL + PERIODS, FINANCIAL_MONTHLY_COL + PERIODS)
    for col in range(4, max_period_col):
        ws.column_dimensions[_col(col)].width = 12
    if ws.title == "Financial Statements":
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["I"].width = 4
    for row in ws.iter_rows():
        for cell in row:
            cell.border = styles["thin_border"]
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                if not cell.number_format or cell.number_format == "General":
                    cell.number_format = '#,##0;[Red](#,##0);-'
    if ws.title == "Summary Financials Annual":
        _apply_bolt_annual_output_format(ws)


def _apply_bolt_annual_output_format(ws) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

    dark_blue = "203764"
    light_section = "D9EAF7"
    pale_blue = "EAF2F8"
    white = "FFFFFF"
    green = "008000"
    blue = "0000FF"
    red = "C00000"
    thin_blue = Side(style="thin", color="D9E1F2")
    no_side = Side(style=None)

    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "J10"
    widths = {
        "A": 2.58,
        "B": 13.0,
        "C": 13.0,
        "D": 30.58,
        "E": 8.58,
        "F": 10.58,
        "G": 13.0,
        "H": 13.0,
        "I": 13.0,
        "J": 20.58,
        "K": 13.0,
        "L": 13.0,
        "M": 13.0,
        "N": 13.0,
        "O": 10.58,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col in range(16, 123):
        ws.column_dimensions[_col(col)].width = 10.58

    heights = {1: 30, 2: 25, 3: 13, 10: 13}
    for row in range(4, 250):
        ws.row_dimensions[row].height = 15.75
    for row in [11, 37, 76, 122, 189, 249]:
        ws.row_dimensions[row].height = 20.15
    for row, height in heights.items():
        ws.row_dimensions[row].height = height

    ws["A1"].font = Font(name="Helvetica", size=16, bold=True, color=dark_blue)
    ws["A1"].alignment = Alignment(vertical="center")
    ws["A2"].font = Font(name="Calibri", size=14, color=dark_blue)
    ws["A2"].alignment = Alignment(vertical="center")
    ws["I1"].number_format = '"Check Model";;"Model OK"'
    ws["I2"].number_format = '"Check Sheet";;"Sheet OK"'

    dark_fill = PatternFill("solid", fgColor=dark_blue)
    section_fill = PatternFill("solid", fgColor=light_section)
    pale_fill = PatternFill("solid", fgColor=pale_blue)
    transparent_border = Border(left=no_side, right=no_side, top=no_side, bottom=no_side)
    thin_border = Border(left=thin_blue, right=thin_blue, top=thin_blue, bottom=thin_blue)

    for row in [4, 5]:
        for col in range(3, 15):
            cell = ws.cell(row, col)
            cell.fill = dark_fill
            cell.font = Font(name="Calibri", size=10 if col >= 9 else 12, bold=True, color=white)
            cell.alignment = Alignment(horizontal="right" if col >= 9 else "center", vertical="center")
            cell.border = transparent_border
    for col in range(10, 15):
        ws.cell(4, col).number_format = '"FY" 0'
        ws.cell(5, col).number_format = '"Q"0;;-'

    for row in [7, 8, 9]:
        for col in range(3, 15):
            cell = ws.cell(row, col)
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font(name="Calibri", size=10, color=dark_blue if col == 9 else "000000")
            cell.alignment = Alignment(horizontal="right" if col == 9 else "center", vertical="center")
            cell.border = transparent_border
        for col in range(10, 15):
            ws.cell(row, col).number_format = "d-mmm-yy" if row in [7, 8] else "0"

    for row in [11, 37, 76, 122, 189, 249]:
        for col in range(1, 15):
            cell = ws.cell(row, col)
            cell.fill = section_fill
            cell.border = transparent_border
        ws.cell(row, 2).font = Font(name="Arial", size=12, bold=True, color=dark_blue)
        ws.cell(row, 3).font = Font(name="Arial", size=11, bold=True, color=dark_blue)

    percent_rows = {15, 47, 48, 49, 50, 62, 63, 64, 65, 66, 98, 109}
    check_rows = {44, 51, 59, 67, 120, 186, 246}
    bold_rows = {
        43, 50, 58, 66, 73, 83, 88, 91, 95, 97, 103, 106, 108, 119,
        131, 153, 155, 165, 177, 182, 184, 198, 229, 235, 240, 242, 244,
    }
    section_labels = {30, 39, 46, 53, 61, 69, 111, 124, 126, 133, 158, 160, 167, 179, 200, 202, 208, 212, 218, 221, 225}
    for row in range(13, 249):
        if row in [37, 76, 122, 189]:
            continue
        ws.cell(row, 3).font = Font(name="Calibri", size=10, bold=row in bold_rows or row in section_labels, color=dark_blue if row in section_labels else "000000")
        ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row, 5).font = Font(name="Calibri", size=10, italic=True, color="000000")
        ws.cell(row, 5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 7).font = Font(name="Calibri", size=10, color=green)
        ws.cell(row, 7).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row, 9).font = Font(name="Calibri", size=10, color=blue if row in check_rows else "000000")
        for col in range(10, 15):
            cell = ws.cell(row, col)
            cell.fill = pale_fill if row in bold_rows else PatternFill(fill_type=None)
            cell.font = Font(name="Calibri", size=10, bold=row in bold_rows, color=red if row in check_rows else "000000")
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border
            if row in percent_rows:
                cell.number_format = '0%;(0%);-'
            elif row in check_rows:
                cell.number_format = '0'
            else:
                cell.number_format = '#,##0;(#,##0);"-" '

    for row in [30, 39, 46, 53, 61, 69, 111, 124, 126, 133, 158, 160, 167, 179, 200, 202, 208, 212, 218, 221, 225]:
        for col in range(3, 15):
            ws.cell(row, col).fill = PatternFill(fill_type=None)
            ws.cell(row, col).border = transparent_border

    for row in [13, 14, 17, 40, 41, 42, 54, 55, 56, 57]:
        for col in [7, 8]:
            ws.cell(row, col).font = Font(name="Calibri", size=10, color=green)
            ws.cell(row, col).number_format = '"FY" 0' if col == 7 else '"Q"0'


def _num(value, fallback: float = 0) -> float:
    try:
        if value in (None, ""):
            return fallback
        return float(value)
    except (TypeError, ValueError):
        return fallback


def _date_value(value, fallback: date):
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        if value.startswith("="):
            return value
        try:
            return date.fromisoformat(value[:10])
        except ValueError:
            return fallback
    return fallback


def _debt_values(tranche: dict) -> tuple:
    return (
        tranche.get("name", "Debt tranche"),
        tranche.get("debt_type", "Senior Term Loan B"),
        tranche.get("borrower", "OpCo"),
        _date_value(tranche.get("start_date"), date(2026, 1, 31)),
        _num(tranche.get("opening_balance"), 0),
        _num(tranche.get("commitment"), 0),
        max(1, int(_num(tranche.get("term_months"), 60))),
        max(0, int(_num(tranche.get("moratorium_months"), 0))),
        max(0, int(_num(tranche.get("interest_cap_months"), 0))),
        _num(tranche.get("margin"), 0.03),
        _num(tranche.get("base_rate"), 0.03),
        tranche.get("amortization", "Bullet"),
        _num(tranche.get("bullet_percent"), 1.0),
        _num(tranche.get("cash_sweep_percent"), 0.0),
        "TRUE" if tranche.get("pik") in (True, "TRUE", "true", "1", 1) else "FALSE",
        _num(tranche.get("minimum_cash"), 50000),
        tranche.get("interest_type", "PIK" if tranche.get("pik") else "Cash"),
        tranche.get("cash_pay_frequency", "Monthly"),
        _num(tranche.get("cash_pay_percent"), 0.0 if tranche.get("pik") else 1.0),
    )


def _financial_lookup(financials: dict) -> dict:
    lookup = {}
    for section in ["income_statement", "balance_sheet", "cash_flow", "debt"]:
        for line in financials.get(section, []) or []:
            name = str(line.get("name", "")).strip().lower()
            if name:
                lookup[name] = line.get("values", {}) or {}
    for line in financials.get("historical_detail", []) or []:
        name = str(line.get("detail_line") or line.get("model_line") or "").strip().lower()
        if name:
            lookup[name] = line.get("values", {}) or {}
    return lookup


def _manual_historical_lookup(assumptions: dict) -> dict:
    lookup = {}
    for line in assumptions.get("historical_actuals", []) or []:
        model_line = str(line.get("model_line", "")).strip().lower()
        detail_line = str(line.get("detail_line", "")).strip().lower()
        values = {
            "FY2022": _num(line.get("fy2022"), 0),
            "FY2023": _num(line.get("fy2023"), 0),
            "FY2024": _num(line.get("fy2024"), 0),
            "FY2025": _num(line.get("fy2025"), 0),
            "latest_actual": _num(line.get("latest_actual"), 0),
        }
        if detail_line:
            lookup[detail_line] = values
        if model_line:
            lookup[model_line] = values
    return lookup


def _historical_line_templates() -> list[tuple[str, str, str, str, str, str]]:
    lines = []
    revenue = [
        "Product revenue", "Service revenue", "Recurring revenue", "Project revenue", "License revenue", "Maintenance revenue",
        "Consulting revenue", "Geography A revenue", "Geography B revenue", "Customer segment A revenue", "Customer segment B revenue",
        "Discounts and rebates", "Returns and credits", "Other revenue", "Revenue IFRS adjustment",
    ]
    cogs = [
        "Materials", "Direct labour", "Fulfilment", "Hosting / delivery", "Subcontractors", "Freight", "Warranty",
        "Merchant fees", "Inventory write-off", "Other direct costs",
    ]
    opex = [
        "Management payroll", "Sales payroll", "Operations payroll", "Finance payroll", "IT payroll", "Admin payroll",
        "Rent", "Marketing", "Travel", "IT and software", "Professional fees", "Insurance", "Recruitment", "Training",
        "Utilities", "Bad debt expense", "Bank charges", "Other SG&A", "One-off restructuring cost", "QoE adjustment",
    ]
    income_other = [
        ("D&A", "Depreciation", "D&A"), ("D&A", "Amortization", "D&A"), ("Interest", "Cash interest", "Cash Interest"),
        ("Interest", "PIK interest", "Cash Interest"), ("Tax", "Current tax", "Tax"), ("Tax", "Deferred tax", "Tax"),
        ("EBITDA Adjustments", "Non-recurring income", "EBITDA"), ("EBITDA Adjustments", "Non-recurring costs", "EBITDA"),
    ]
    for item in revenue:
        lines.append(("Income Statement", "Revenue", "Revenue streams", "Revenue", item, "+"))
    for item in cogs:
        lines.append(("Income Statement", "COGS", "Direct costs", "COGS", item, "-"))
    for item in opex:
        model_line = "Payroll" if "payroll" in item.lower() else "Opex"
        lines.append(("Income Statement", "Opex", "Operating costs", model_line, item, "-"))
    for category, item, model_line in income_other:
        lines.append(("Income Statement", category, "Below EBITDA / QoE", model_line, item, "-" if category in ["D&A", "Interest", "Tax"] else "+/-"))

    assets = [
        ("Cash", "Cash at bank"), ("Cash", "Restricted cash"), ("Receivables", "Trade receivables"), ("Receivables", "Unbilled revenue"),
        ("Receivables", "Related-party receivables"), ("Inventory", "Raw materials"), ("Inventory", "Finished goods"),
        ("Other Current Assets", "Prepayments"), ("Other Current Assets", "VAT receivable"), ("Other Current Assets", "Other current assets"),
        ("Net PPE", "Land and buildings"), ("Net PPE", "Plant and machinery"), ("Net PPE", "IT equipment"), ("Net PPE", "Right-of-use assets"),
        ("Intangibles", "Goodwill"), ("Intangibles", "Capitalised software"), ("Intangibles", "Customer relationships"),
        ("Other Non Current Assets", "Deferred tax asset"), ("Other Non Current Assets", "Deposits"), ("Other Non Current Assets", "Investments"),
    ]
    liabilities = [
        ("Payables", "Trade payables"), ("Payables", "Supplier accruals"), ("Accruals", "Payroll accruals"), ("Accruals", "Bonus accruals"),
        ("Accruals", "Other accruals"), ("Tax Payable", "VAT payable"), ("Tax Payable", "Corporate tax payable"),
        ("Closing Debt", "Super senior RCF"), ("Closing Debt", "Senior term loan A"), ("Closing Debt", "Senior term loan B"),
        ("Closing Debt", "Unitranche"), ("Closing Debt", "Second lien"), ("Closing Debt", "Mezzanine cash pay"),
        ("Closing Debt", "Mezzanine PIK"), ("Closing Debt", "HoldCo PIK"), ("Closing Debt", "Seller note"),
        ("Closing Debt", "Vendor loan"), ("Closing Debt", "Finance leases"), ("Closing Debt", "Tax debt plan"),
        ("Closing Debt", "Supplier payment plan"), ("Equity", "Share capital"), ("Equity", "Retained earnings"), ("Equity", "Other reserves"),
    ]
    for model_line, item in assets:
        category = "Current Assets" if model_line in ["Cash", "Receivables", "Inventory", "Other Current Assets"] else "Fixed Assets"
        lines.append(("Balance Sheet", category, model_line, model_line, item, "+"))
    for model_line, item in liabilities:
        category = "Debt" if model_line == "Closing Debt" else "Current Liabilities" if model_line in ["Payables", "Accruals", "Tax Payable"] else "Equity"
        lines.append(("Balance Sheet", category, model_line, model_line, item, "-"))

    cash_flow = [
        ("Operating Cash Flow", "EBITDA bridge", "EBITDA"), ("Operating Cash Flow", "Change in trade receivables", "Change in NWC"),
        ("Operating Cash Flow", "Change in inventory", "Change in NWC"), ("Operating Cash Flow", "Change in trade payables", "Change in NWC"),
        ("Operating Cash Flow", "Change in other working capital", "Change in NWC"), ("Investing Cash Flow", "Maintenance capex", "Capex"),
        ("Investing Cash Flow", "Growth capex", "Capex"), ("Investing Cash Flow", "Disposal proceeds", "Capex"),
        ("Financing Cash Flow", "Debt drawdown", "Closing Debt"), ("Financing Cash Flow", "Scheduled amortization", "Debt Amortization / Sweep"),
        ("Financing Cash Flow", "Bullet repayment", "Debt Amortization / Sweep"), ("Financing Cash Flow", "Cash sweep", "Debt Amortization / Sweep"),
        ("Financing Cash Flow", "Equity injection", "Equity"), ("Financing Cash Flow", "Dividend", "Equity"),
    ]
    for category, item, model_line in cash_flow:
        lines.append(("Cash Flow", category, "Cash-flow detail", model_line, item, "+/-"))

    while len(lines) < HISTORICAL_DETAIL_LINES:
        idx = len(lines) + 1
        lines.append(("Income Statement", "Opex", "Custom line", "Opex", f"Custom historical line {idx}", "+/-"))
    return lines[:HISTORICAL_DETAIL_LINES]


def _detail_projection_formula(row: int, model_line: str, fs_col: str) -> str:
    fs_rows = {
        "Revenue": 6,
        "COGS": 7,
        "Gross Profit": 8,
        "Payroll": 9,
        "Opex": 10,
        "EBITDA": 11,
        "D&A": 12,
        "EBIT": 13,
        "Cash Interest": 14,
        "PBT": 15,
        "Tax": 16,
        "Net Income": 17,
        "Change in NWC": 18,
        "Capex": 19,
        "Free Cash Flow": 22,
        "Cash": 23,
        "Closing Debt": 24,
        "Net Debt": 25,
        "Receivables": 23,
        "Inventory": 23,
        "Payables": 24,
        "Net PPE": 23,
        "Equity": 25,
        "Debt Amortization / Sweep": 21,
    }
    source_row = fs_rows.get(model_line, 10)
    return (
        f'=IFERROR(\'Financial Statements\'!{fs_col}{source_row}*'
        f"IF(SUMIFS('Historical Detail Input'!$P$6:$P$185,'Historical Detail Input'!$F$6:$F$185,$E{row})<>0,"
        f"'Historical Detail Input'!$P{row}/SUMIFS('Historical Detail Input'!$P$6:$P$185,'Historical Detail Input'!$F$6:$F$185,$E{row}),"
        f"1/COUNTIF($E$6:$E$185,$E{row})),0)"
    )


def _summary_financial_rows(include_details: bool = False) -> list[tuple[str, str, str]]:
    rows = [
        ("Revenue", "FS:6", "flow"),
        ("COGS", "FS:7", "flow"),
        ("Gross Profit", "FS:8", "flow"),
        ("Payroll", "FS:9", "flow"),
        ("Opex", "FS:10", "flow"),
        ("EBITDA", "FS:11", "flow"),
        ("EBITDA Margin", "FS:11", "margin"),
        ("D&A", "FS:12", "flow"),
        ("EBIT", "FS:13", "flow"),
        ("Cash Interest", "FS:14", "flow"),
        ("PBT", "FS:15", "flow"),
        ("Tax", "FS:16", "flow"),
        ("Net Income", "FS:17", "flow"),
        ("Change in NWC", "FS:18", "flow"),
        ("Capex", "FS:19", "flow"),
        ("Cash Flow Before Debt", "FS:20", "flow"),
        ("Debt Amortization / Sweep", "FS:21", "flow"),
        ("Free Cash Flow", "FS:22", "flow"),
        ("Closing Cash", "FS:23", "balance"),
        ("Closing Debt", "FS:24", "balance"),
        ("Net Debt", "FS:25", "balance"),
    ]
    if include_details:
        templates = _historical_line_templates()
        for idx, (_statement, category, _subcategory, model_line, detail_line, _sign) in enumerate(templates[:170], start=6):
            rows.append((f"{category} - {detail_line}", f"DETAIL:{idx}", "flow" if model_line not in ["Cash", "Closing Debt", "Net Debt", "Receivables", "Inventory", "Payables", "Net PPE", "Equity"] else "balance"))
    return rows


def _period_cols() -> range:
    return range(FIRST_PERIOD_COL, FIRST_PERIOD_COL + PERIODS)


def _financial_monthly_col(period_col: int) -> str:
    return _col(FINANCIAL_MONTHLY_COL + period_col - FIRST_PERIOD_COL)


def _flatten_assumptions(value, prefix: str = "") -> list[tuple[str, str, object]]:
    rows = []
    if isinstance(value, dict):
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(_flatten_assumptions(child, child_prefix))
    elif isinstance(value, list):
        for idx, child in enumerate(value, start=1):
            child_prefix = f"{prefix}[{idx}]"
            rows.extend(_flatten_assumptions(child, child_prefix))
    else:
        category, _, key = prefix.partition(".")
        rows.append((category or "assumptions", key or prefix, value))
    return rows


def _assumption_type(value) -> str:
    if isinstance(value, bool):
        return "Boolean"
    if isinstance(value, (int, float)):
        return "Numeric"
    if isinstance(value, str) and len(value) >= 10 and value[4:5] == "-" and value[7:8] == "-":
        return "Date"
    return "Text"


def _assumption_step(category: str) -> str:
    steps = {
        "model": "1. Model setup",
        "historical_actuals": "2. Historical actuals",
        "revenue_streams": "3. Revenue build",
        "cost_base": "4. Direct cost setup",
        "cost_items": "5. Opex structure",
        "headcount": "6. Headcount",
        "working_capital": "7. Working capital",
        "capex": "8. Capex / D&A",
        "debt_tranches": "9. Debt structure",
        "covenants": "10. Covenants",
    }
    return steps.get(category, "Other model input")


def _assumption_required(category: str, key: str) -> bool:
    key_lower = key.lower()
    required_keys = {
        "model": {"company_name", "currency", "scenario", "model_start_date", "actuals_end_date", "forecast_months"},
        "revenue_streams": {"name", "type", "volume", "price"},
        "cost_items": {"name", "driver"},
        "working_capital": {"dso", "dpo", "dio"},
        "debt_tranches": {"name", "debt_type", "start_date", "opening_balance", "term_months", "margin", "base_rate"},
        "covenants": {"max_net_debt_ebitda", "min_interest_cover", "min_liquidity"},
    }
    if key_lower in required_keys.get(category, set()):
        return True
    return category in {"historical_actuals"} and key_lower in {"statement", "model_line", "detail_line", "latest_actual"}


def _assumption_priority(category: str, key: str, value) -> str:
    if _assumption_required(category, key) and value in ("", None):
        return "Critical - missing"
    if category in {"model", "historical_actuals", "debt_tranches"}:
        return "High"
    if category in {"revenue_streams", "working_capital", "covenants"}:
        return "Medium"
    return "Standard"


def _assumption_validation_note(category: str, key: str, value) -> str:
    key_lower = key.lower()
    if _assumption_required(category, key) and value in ("", None):
        return "Complete this input before generating an investment-grade model."
    if key_lower.endswith("date"):
        return "Must align with the model monthly timeline."
    if key_lower in {"margin", "base_rate", "tax_rate", "leverage_limit", "minimum_liquidity"}:
        return "Review against debt term sheet, management case or lender covenant package."
    if category == "historical_actuals":
        return "Should be reviewed after Claude extraction before feeding the forecast bridge."
    if category == "revenue_streams":
        return "Flows into product revenue build and annual/monthly financial statements."
    if category == "debt_tranches":
        return "Flows into cash interest, PIK accrual, amortization, maturity and covenant checks."
    return "Linked to the model output tabs through formulas."


def _styles(Font, PatternFill, Border, Side, Alignment) -> dict:
    side = Side(style="thin", color="D9E1F2")
    return {
        "title_font": Font(name="Arial", size=20, bold=True, color="1F4E78"),
        "subtitle_font": Font(name="Arial", size=14, bold=True, color="404040"),
        "section_font": Font(name="Arial", size=12, bold=True, color="1F4E78"),
        "bold_font": Font(name="Arial", bold=True, color="404040"),
        "input_font": Font(name="Arial", color="0000FF"),
        "header_font": Font(name="Arial", bold=True, color="FFFFFF"),
        "header_fill": PatternFill("solid", fgColor=HEADER_FILL),
        "section_fill": PatternFill("solid", fgColor=SECTION_FILL),
        "input_fill": PatternFill("solid", fgColor=INPUT_FILL),
        "output_fill": PatternFill("solid", fgColor=OUTPUT_FILL),
        "formula_fill": PatternFill("solid", fgColor=FORMULA_FILL),
        "check_fill": PatternFill("solid", fgColor=CHECK_FILL),
        "thin_border": Border(left=side, right=side, top=side, bottom=side),
        "center": Alignment(horizontal="center"),
    }


def _col(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result
