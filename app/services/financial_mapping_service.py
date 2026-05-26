import csv
import json
from pathlib import Path
from datetime import date, datetime

from app.engines.historical_accounts_engine import build_basic_historical_pack
from app.services.ai_service import extract_financial_historicals
from app.services.extraction_service import list_project_documents
from app.services.project_service import get_project, project_dir


def normalize_project_financials(project_id: str, use_ai: bool = True) -> dict:
    documents = list_project_documents(project_id)
    extracted_rows = []
    document_summaries = []

    for document in documents:
        if document.suffix.lower() == ".csv":
            rows = _read_csv_rows(document)
            extracted_rows.extend(rows)
            document_summaries.append(_document_summary(document, rows))
        elif document.suffix.lower() in {".xlsx", ".xlsm"}:
            rows = _read_xlsx_rows(document)
            extracted_rows.extend(rows)
            document_summaries.append(_document_summary(document, rows))
        elif document.suffix.lower() == ".pdf":
            text = _read_pdf_text(document)
            document_summaries.append(_document_summary(document, [], text=text))

    normalized = build_basic_historical_pack(extracted_rows, [path.name for path in documents])
    normalized["extraction"] = {
        "mode": "local",
        "confidence": "low" if not extracted_rows else "medium",
        "documents": document_summaries,
    }

    if use_ai and documents:
        ai_result = _run_ai_historical_extraction(project_id, documents, extracted_rows, document_summaries)
        if ai_result:
            normalized = _merge_ai_extraction(normalized, ai_result)

    output_path = project_dir(project_id) / "normalized" / "financials.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(normalized, indent=2), encoding="utf-8")
    return normalized


def load_normalized_financials(project_id: str) -> dict:
    output_path = project_dir(project_id) / "normalized" / "financials.json"
    if not output_path.exists():
        return normalize_project_financials(project_id)
    try:
        return json.loads(output_path.read_text(encoding="utf-8"))
    except Exception:
        return normalize_project_financials(project_id)


def save_normalized_financials(project_id: str, financials: dict) -> dict:
    output_path = project_dir(project_id) / "normalized" / "financials.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(financials, indent=2), encoding="utf-8")
    return financials


def merge_ai_extraction_into_financials(normalized: dict, ai_result: dict) -> dict:
    return _merge_ai_extraction(normalized, ai_result)


def _read_csv_rows(path: Path) -> list[dict]:
    rows = []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                rows.append({str(key or "").strip(): value for key, value in row.items()})
    except Exception:
        return []
    return rows


def _read_xlsx_rows(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    rows = []
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
        for sheet in workbook.worksheets[:12]:
            values = list(sheet.iter_rows(values_only=True, max_row=250, max_col=80))
            rows.extend(_extract_sheet_rows(sheet.title, values))
    except Exception:
        return []
    return rows


def _extract_sheet_rows(sheet_name: str, values: list[tuple]) -> list[dict]:
    extracted = []
    if not values:
        return extracted

    header_candidates = []
    for idx, row in enumerate(values[:80]):
        period_cells = [_period_label(cell) for cell in row]
        period_count = sum(1 for period in period_cells if period)
        if period_count >= 2:
            header_candidates.append((idx, period_cells))

    if header_candidates:
        header_idx, period_cells = header_candidates[0]
        period_columns = [(col_idx, period) for col_idx, period in enumerate(period_cells) if period]
        for row_idx, row in enumerate(values[header_idx + 1 :], start=header_idx + 2):
            label = _best_row_label(row)
            if not label:
                continue
            item = {"Source Sheet": sheet_name, "Source Row": row_idx, "Line Item": label}
            has_value = False
            for col_idx, period in period_columns:
                value = row[col_idx] if col_idx < len(row) else None
                number = _coerce_number(value)
                if number is not None:
                    item[period] = number
                    has_value = True
            if has_value:
                extracted.append(item)
        return extracted

    # Fallback for flat files where first row is a conventional header.
    headers = [str(cell or "").strip() for cell in values[0]]
    for row_idx, values_row in enumerate(values[1:], start=2):
        row = {"Source Sheet": sheet_name, "Source Row": row_idx}
        has_value = False
        for index, header in enumerate(headers):
            if not header:
                continue
            value = values_row[index] if index < len(values_row) else None
            row[_period_label(header) or header] = value
            if value not in (None, ""):
                has_value = True
        if has_value:
            extracted.append(row)
    return extracted


def _read_pdf_text(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        return ""

    try:
        reader = PdfReader(str(path))
        pages = []
        for page in reader.pages[:25]:
            pages.append(page.extract_text() or "")
        return "\n".join(pages)[:45000]
    except Exception:
        return ""


def _document_summary(path: Path, rows: list[dict], text: str = "") -> dict:
    preview_rows = rows[:80]
    return {
        "filename": path.name,
        "extension": path.suffix.lower(),
        "row_count": len(rows),
        "columns": list(preview_rows[0].keys())[:40] if preview_rows else [],
        "preview_rows": preview_rows[:25],
        "text_preview": text[:12000],
    }


def _best_row_label(row: tuple) -> str:
    for cell in row[:12]:
        if isinstance(cell, str):
            text = cell.strip()
            if text and not _period_label(text) and not _looks_numeric(text):
                return text[:180]
    return ""


def _period_label(value) -> str | None:
    if isinstance(value, (date, datetime)):
        return f"FY{value.year}"
    text = str(value or "").strip()
    if not text:
        return None
    upper = text.upper().replace(" ", "")
    if upper.startswith("FY") and any(char.isdigit() for char in upper):
        digits = "".join(char for char in upper if char.isdigit())
        if len(digits) == 2:
            return f"FY20{digits}"
        if len(digits) >= 4:
            return f"FY{digits[:4]}"
    if text.isdigit() and len(text) == 4 and 1990 <= int(text) <= 2100:
        return f"FY{text}"
    for token in ["2021", "2022", "2023", "2024", "2025", "2026", "2027", "2028", "2029", "2030"]:
        if token in text:
            return f"FY{token}"
    return None


def _coerce_number(value):
    if isinstance(value, (int, float)):
        return float(value)
    if value in (None, ""):
        return None
    text = str(value).strip().replace(",", "").replace(" ", "").replace("(", "-").replace(")", "")
    try:
        return float(text)
    except ValueError:
        return None


def _looks_numeric(value: str) -> bool:
    return _coerce_number(value) is not None


def _run_ai_historical_extraction(
    project_id: str,
    documents: list[Path],
    rows: list[dict],
    document_summaries: list[dict],
) -> dict | None:
    context = {
        "source_files": [path.name for path in documents],
        "structured_rows": rows[:1200],
        "documents": document_summaries,
    }
    project = get_project(project_id) or {"id": project_id}
    result = extract_financial_historicals(project, context)

    trace_path = project_dir(project_id) / "normalized" / "claude_extraction.json"
    trace_path.parent.mkdir(parents=True, exist_ok=True)
    trace_path.write_text(json.dumps(result, indent=2), encoding="utf-8")
    return result


def _merge_ai_extraction(normalized: dict, ai_result: dict) -> dict:
    extraction = ai_result.get("extraction") or {}
    merged = dict(normalized)

    for key in ["periods", "income_statement", "balance_sheet", "cash_flow", "historical_detail", "source_files"]:
        value = extraction.get(key)
        if value:
            merged[key] = value

    if extraction.get("working_capital"):
        merged["working_capital"] = extraction["working_capital"]
    if extraction.get("debt"):
        merged["debt"] = extraction["debt"]

    merged["currency"] = extraction.get("currency") or merged.get("currency", "unknown")
    merged["unit"] = extraction.get("unit") or merged.get("unit", "unknown")
    merged["extraction"] = {
        "mode": ai_result.get("source", "unknown"),
        "claude_configured": ai_result.get("configured", False),
        "confidence": extraction.get("confidence", "low"),
        "issues": extraction.get("issues", []),
        "error": ai_result.get("error"),
        "documents": normalized.get("extraction", {}).get("documents", []),
    }
    return merged
