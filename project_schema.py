from datetime import date
from pathlib import Path

from app.engines.debt_engine import debt_type_options


PERIODS = 60
MAX_DEBT_TRANCHES = 10
FIRST_PERIOD_COL = 4
ENTITIES = ["Group", "OpCo", "HoldCo"]
INPUT_FILL = "D9EAF7"
HEADER_FILL = "1F4E78"
SECTION_FILL = "D9E2F3"
OUTPUT_FILL = "E2F0D9"
CHECK_FILL = "FCE4D6"
FORMULA_FILL = "FFFFFF"
LINK_FILL = "EAF2F8"


def build_business_plan_workbook(project: dict, financials: dict, output_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except Exception as exc:
        raise RuntimeError(f"Excel generation dependency unavailable: {exc}") from exc

    wb = Workbook()
    wb.remove(wb.active)
    styles = _styles(Font, PatternFill, Border, Side, Alignment)

    admin = wb.create_sheet("Admin")
    group_assumptions = wb.create_sheet("Group Assumptions")
    cover = wb.create_sheet("Cover")
    data_room = wb.create_sheet("Data Room")
    control = wb.create_sheet("Control Panel")
    historical = wb.create_sheet("Historical Inputs")
    entity_input_sheets = [wb.create_sheet(f"{entity} Data Input") for entity in ENTITIES]
    entity_output_sheets = [wb.create_sheet(f"Output_{entity}_Monthly") for entity in ENTITIES]
    entity_annual_sheets = [wb.create_sheet(f"Output_{entity}_Annual") for entity in ENTITIES]
    revenue = wb.create_sheet("Revenue Drivers")
    products = wb.create_sheet("Product Build")
    headcount = wb.create_sheet("Headcount")
    opex = wb.create_sheet("Opex")
    wc = wb.create_sheet("Working Capital")
    capex = wb.create_sheet("Capex D&A")
    debt_config = wb.create_sheet("Debt Config")
    debt_schedule = wb.create_sheet("Debt Schedule")
    statements = wb.create_sheet("Financial Statements")
    covenants = wb.create_sheet("Covenants")
    outputs = wb.create_sheet("Outputs")
    packaged = wb.create_sheet("Packaged Output")
    ic_summary = wb.create_sheet("IC Summary")
    restructuring = wb.create_sheet("Restructuring Options")
    debt_capacity = wb.create_sheet("Debt Capacity")
    sensitivities = wb.create_sheet("Sensitivity Matrix")
    checks = wb.create_sheet("Checks")
    lookup = wb.create_sheet("Lookup")
    mapping = wb.create_sheet("Mapping")
    lists = wb.create_sheet("Lists & Dates")

    _build_lists(lists, project, styles)
    _build_admin(admin, project, styles)
    _build_group_assumptions(group_assumptions, project, styles, DataValidation)
    _build_cover(cover, project, styles)
    _build_data_room(data_room, project, financials, styles)
    _build_control(control, project, styles, DataValidation)
    _build_historical(historical, financials, styles)
    for idx, sheet in enumerate(entity_input_sheets):
        _build_entity_input(sheet, ENTITIES[idx], styles)
    for idx, sheet in enumerate(entity_output_sheets):
        _build_entity_monthly_output(sheet, ENTITIES[idx], styles)
    for idx, sheet in enumerate(entity_annual_sheets):
        _build_entity_annual_output(sheet, ENTITIES[idx], styles)
    _build_revenue_drivers(revenue, styles, DataValidation)
    _build_product_build(products, styles, DataValidation)
    _build_headcount(headcount, styles)
    _build_opex(opex, styles, DataValidation)
    _build_working_capital(wc, styles)
    _build_capex(capex, styles)
    _build_debt_config(debt_config, styles, DataValidation)
    _build_debt_schedule(debt_schedule, styles)
    _build_financial_statements(statements, styles)
    _build_covenants(covenants, styles)
    _build_outputs(outputs, styles)
    _build_packaged_output(packaged, project, styles)
    _build_ic_summary(ic_summary, project, styles)
    _build_restructuring_options(restructuring, styles)
    _build_debt_capacity(debt_capacity, styles)
    _build_sensitivity_matrix(sensitivities, styles)
    _build_checks(checks, styles)
    _build_lookup(lookup, styles)
    _build_mapping(mapping, styles)

    for sheet in wb.worksheets:
        _polish_sheet(sheet, styles)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _build_admin(ws, project: dict, styles: dict) -> None:
    ws["B2"] = project.get("company_name", "Target Company")
    ws["B2"].font = styles["title_font"]
    ws["B4"] = "Admin"
    ws["B4"].font = styles["section_font"]
    rows = [
        ("Model Start", "='Control Panel'!$C$8"),
        ("Model End", "=INDEX('Lists & Dates'!$V$2:$V$61,'Control Panel'!$C$10)"),
        ("Actuals End", "='Control Panel'!$C$9"),
        ("Scenario", "='Control Panel'!$C$7"),
        ("Currency", "='Control Panel'!$C$6"),
        ("Forecast Months", "='Control Panel'!$C$10"),
        ("Generated Model", "Formula-driven BP / debt / covenant pack"),
    ]
    for row, (label, value) in enumerate(rows, start=6):
        _label(ws, row, 2, label, styles)
        _formula(ws, row, 3, value, styles, output=True) if isinstance(value, str) and value.startswith("=") else _input(ws, row, 3, value, styles)
    ws["B16"] = "Tab Colour Coding"
    ws["B16"].font = styles["section_font"]
    coding = [
        ("Input", "Blue cells are editable assumptions."),
        ("Formula", "White cells are formulas."),
        ("Output", "Green cells are linked output lines."),
        ("Check", "Orange cells are control checks."),
    ]
    for row, (label, desc) in enumerate(coding, start=17):
        ws.cell(row, 2, label)
        ws.cell(row, 3, desc)


def _build_group_assumptions(ws, project: dict, styles: dict, DataValidation) -> None:
    ws["B2"] = "Group Assumptions"
    ws["B2"].font = styles["section_font"]
    assumptions = [
        ("Company", project.get("company_name", "Target Company")),
        ("Currency", project.get("currency", "EUR")),
        ("Scenario", "Base"),
        ("Fiscal Year End", project.get("fiscal_year_end", "December")),
        ("Tax Rate", "='Control Panel'!$C$13"),
        ("Minimum Cash", "='Control Panel'!$C$14"),
        ("Opening Cash", "='Control Panel'!$C$11"),
        ("Opening Debt", "='Control Panel'!$C$12"),
        ("Consolidation Method", "Management case"),
    ]
    for row, (label, value) in enumerate(assumptions, start=5):
        _label(ws, row, 2, label, styles)
        if isinstance(value, str) and value.startswith("="):
            _formula(ws, row, 3, value, styles, output=True)
        else:
            _input(ws, row, 3, value, styles)
    _add_list_validation(ws, "C6", "'Lists & Dates'!$B$2:$B$8", DataValidation)
    _add_list_validation(ws, "C7", "'Lists & Dates'!$E$2:$E$7", DataValidation)
    ws["B17"] = "Entity Weighting"
    ws["B17"].font = styles["section_font"]
    _table_header(ws, 19, ["Entity", "Revenue Weight", "Opex Weight", "Debt Allocation"], styles)
    defaults = [("Group", 1.0, 1.0, 1.0), ("OpCo", 0.85, 0.90, 0.75), ("HoldCo", 0.15, 0.10, 0.25)]
    for row, values in enumerate(defaults, start=20):
        for col, value in enumerate(values, start=2):
            _input(ws, row, col, value, styles)
            if col > 2:
                ws.cell(row, col).number_format = "0.0%"


def _build_cover(ws, project: dict, styles: dict) -> None:
    ws["B2"] = "MG Advisory Finance OS"
    ws["B2"].font = styles["title_font"]
    ws["B4"] = "Institutional Business Plan Model"
    ws["B4"].font = styles["subtitle_font"]
    rows = [
        ("Company", project.get("company_name", "Target Company")),
        ("Project Type", project.get("project_type", "Investment case")),
        ("Currency", project.get("currency", "EUR")),
        ("Fiscal Year End", project.get("fiscal_year_end", "December")),
        ("Model Standard", "Driver-based monthly BP, debt schedule, covenants and checks"),
    ]
    for row, (label, value) in enumerate(rows, start=7):
        _label(ws, row, 2, label, styles)
        _input(ws, row, 3, value, styles)
    ws["B15"] = "Colour Code"
    ws["B15"].font = styles["section_font"]
    legend = [("Blue", "Hardcoded user input"), ("White", "Formula"), ("Light green", "Output"), ("Orange", "Check")]
    for row, (label, value) in enumerate(legend, start=16):
        ws.cell(row, 2, label)
        ws.cell(row, 3, value)
    ws["B22"] = "Operating Rule"
    ws["C22"] = "Inputs live in blue cells. Calculations and outputs are formula-driven."


def _build_data_room(ws, project: dict, financials: dict, styles: dict) -> None:
    ws["B2"] = "Data Room & Upload Map"
    ws["B2"].font = styles["section_font"]
    rows = [
        ("Audited Accounts PDF", "Upload", "data/projects/{project_id}/documents", "Historical accounts extraction"),
        ("Management Accounts Excel", "Upload", "data/projects/{project_id}/documents", "Monthly actuals"),
        ("Trial Balance", "Upload", "data/projects/{project_id}/documents", "Mapping to FS lines"),
        ("Aged Receivables", "Upload", "data/projects/{project_id}/documents", "DSO / working capital"),
        ("Aged Payables", "Upload", "data/projects/{project_id}/documents", "DPO / working capital"),
        ("Debt Schedule", "Upload", "data/projects/{project_id}/documents", "Debt config and covenants"),
        ("Budget / Forecast", "Upload", "data/projects/{project_id}/documents", "Forward assumptions"),
    ]
    _table_header(ws, 4, ["Document Type", "Action", "Storage", "Model Use"], styles)
    for r, row_values in enumerate(rows, start=5):
        for c, value in enumerate(row_values, start=2):
            ws.cell(r, c, value)
    ws["B15"] = "Uploaded Source Files"
    ws["B15"].font = styles["section_font"]
    _table_header(ws, 17, ["#", "Source File"], styles)
    sources = financials.get("source_files", []) or ["No files uploaded yet"]
    for idx, source in enumerate(sources, start=1):
        ws.cell(17 + idx, 2, idx)
        ws.cell(17 + idx, 3, source)


def _build_control(ws, project: dict, styles: dict, DataValidation) -> None:
    ws["B2"] = "Control Panel"
    ws["B2"].font = styles["section_font"]
    controls = [
        ("Company Name", project.get("company_name", "Target Company")),
        ("Currency", project.get("currency", "EUR")),
        ("Scenario", "Base"),
        ("Model Start Date", date(2026, 1, 31)),
        ("Actuals End Date", date(2025, 12, 31)),
        ("Forecast Months", PERIODS),
        ("Opening Cash", 120000),
        ("Opening Debt", 500000),
        ("Tax Rate", 0.25),
        ("Minimum Cash", 50000),
    ]
    for row, (label, value) in enumerate(controls, start=5):
        _label(ws, row, 2, label, styles)
        _input(ws, row, 3, value, styles)
    ws["C8"].number_format = "yyyy-mm-dd"
    ws["C9"].number_format = "yyyy-mm-dd"
    ws["C13"].number_format = "0.0%"
    _add_list_validation(ws, "C7", "'Lists & Dates'!$E$2:$E$7", DataValidation)
    _add_list_validation(ws, "C6", "'Lists & Dates'!$B$2:$B$8", DataValidation)
    _write_period_headers(ws, 17, styles, source_sheet="'Lists & Dates'")


def _build_historical(ws, financials: dict, styles: dict) -> None:
    ws["B2"] = "Historical Inputs"
    ws["B2"].font = styles["section_font"]
    periods = financials.get("periods") or ["FY2023", "FY2024", "FY2025"]
    _table_header(ws, 4, ["Line Item", "Source"] + periods, styles)
    rows = []
    for section in ["income_statement", "balance_sheet", "cash_flow"]:
        for line in financials.get(section, []):
            rows.append((line.get("name", ""), section, line.get("values", {})))
    if not rows:
        rows = [
            ("Revenue", "fallback", {"FY2023": 900000, "FY2024": 1000000, "FY2025": 1100000}),
            ("COGS", "fallback", {"FY2023": -400000, "FY2024": -450000, "FY2025": -500000}),
            ("Opex", "fallback", {"FY2023": -220000, "FY2024": -250000, "FY2025": -280000}),
            ("EBITDA", "fallback", {"FY2023": 280000, "FY2024": 300000, "FY2025": 320000}),
            ("Cash", "fallback", {"FY2023": 90000, "FY2024": 110000, "FY2025": 120000}),
            ("Debt", "fallback", {"FY2023": 550000, "FY2024": 525000, "FY2025": 500000}),
        ]
    for r, (name, source, values) in enumerate(rows, start=5):
        ws.cell(r, 2, name)
        ws.cell(r, 3, source)
        for c, period in enumerate(periods, start=4):
            _input(ws, r, c, float(values.get(period, 0)), styles)


def _build_entity_input(ws, entity: str, styles: dict) -> None:
    ws["B2"] = f"{entity} Data Input"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    rows = [
        ("Revenue", "000s", "='Revenue Drivers'!{col}14"),
        ("COGS", "000s", "='Product Build'!{col}14"),
        ("Payroll", "000s", "='Headcount'!{col}16"),
        ("Opex", "000s", "='Opex'!{col}15"),
        ("EBITDA", "000s", "={col}6+{col}7+{col}8+{col}9"),
        ("Capex", "000s", "='Capex D&A'!{col}13"),
        ("Receivables", "000s", "='Working Capital'!{col}9"),
        ("Inventory", "000s", "='Working Capital'!{col}10"),
        ("Payables", "000s", "='Working Capital'!{col}11"),
        ("Closing Debt", "000s", "='Financial Statements'!{col}24"),
    ]
    weight_row = 20 + ENTITIES.index(entity)
    _table_header(ws, 6, ["Metric", "Unit"], styles)
    for row, (label, unit, template) in enumerate(rows, start=7):
        ws.cell(row, 2, label)
        ws.cell(row, 3, unit)
        for col_idx in _period_cols():
            col = _col(col_idx)
            formula = template.format(col=col)
            if entity != "Group":
                if label in ["Revenue", "COGS", "EBITDA", "Receivables", "Inventory", "Payables"]:
                    formula = f"=({formula.lstrip('=')})*'Group Assumptions'!$C${weight_row}"
                elif label in ["Payroll", "Opex", "Capex"]:
                    formula = f"=({formula.lstrip('=')})*'Group Assumptions'!$D${weight_row}"
                elif label == "Closing Debt":
                    formula = f"=({formula.lstrip('=')})*'Group Assumptions'!$E${weight_row}"
            _formula(ws, row, col_idx, formula, styles, output=row in [11, 16])


def _build_entity_monthly_output(ws, entity: str, styles: dict) -> None:
    ws["B2"] = f"{entity} Monthly Output"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    rows = [
        ("Revenue", 7),
        ("Gross Profit", 8),
        ("EBITDA", 9),
        ("EBITDA Margin", 10),
        ("Capex", 11),
        ("Closing Debt", 12),
        ("Net Debt / EBITDA", 13),
    ]
    input_sheet = f"{entity} Data Input"
    for label, row in rows:
        ws.cell(row, 2, label)
    for col_idx in _period_cols():
        col = _col(col_idx)
        _formula(ws, 7, col_idx, f"='{input_sheet}'!{col}7", styles, output=True)
        _formula(ws, 8, col_idx, f"='{input_sheet}'!{col}7+'{input_sheet}'!{col}8", styles, output=True)
        _formula(ws, 9, col_idx, f"='{input_sheet}'!{col}11", styles, output=True)
        _formula(ws, 10, col_idx, f"=IFERROR({col}9/{col}7,0)", styles, output=True, fmt="0.0%")
        _formula(ws, 11, col_idx, f"='{input_sheet}'!{col}12", styles, output=True)
        _formula(ws, 12, col_idx, f"='{input_sheet}'!{col}16", styles, output=True)
        _formula(ws, 13, col_idx, f"=IFERROR({col}12/{col}9,0)", styles, output=True, fmt="0.0x")


def _build_entity_annual_output(ws, entity: str, styles: dict) -> None:
    ws["B2"] = f"{entity} Annual Output"
    ws["B2"].font = styles["section_font"]
    years = [2026, 2027, 2028, 2029, 2030]
    _table_header(ws, 4, ["Metric"] + [f"FY{year}" for year in years], styles)
    rows = ["Revenue", "Gross Profit", "EBITDA", "EBITDA Margin", "Capex", "Closing Debt", "Net Debt / EBITDA"]
    for row, label in enumerate(rows, start=5):
        ws.cell(row, 2, label)
    monthly = f"Output_{entity}_Monthly"
    for idx, year in enumerate(years, start=3):
        col = _col(idx)
        for row in range(5, 12):
            source_row = row + 2
            if row in [8, 11]:
                _formula(ws, row, idx, f"=IFERROR({col}{row-1}/{col}5,0)" if row == 8 else f"=IFERROR({col}10/{col}7,0)", styles, output=True, fmt="0.0%" if row == 8 else "0.0x")
            elif row in [10]:
                _formula(ws, row, idx, f"=XLOOKUP(DATE({year},12,31),'{monthly}'!$D$4:$BK$4,'{monthly}'!$D$12:$BK$12,0)", styles, output=True)
            else:
                _formula(ws, row, idx, f'=SUMIFS(\'{monthly}\'!$D${source_row}:$BK${source_row},\'{monthly}\'!$D$4:$BK$4,">="&DATE({year},1,1),\'{monthly}\'!$D$4:$BK$4,"<="&DATE({year},12,31))', styles, output=True)


def _build_revenue_drivers(ws, styles: dict, DataValidation) -> None:
    ws["B2"] = "Revenue Drivers"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    rows = [
        ("Product / Service 1", "Product"),
        ("Product / Service 2", "Service"),
        ("Product / Service 3", "Recurring"),
        ("Product / Service 4", "Project"),
        ("Product / Service 5", "Other"),
    ]
    _table_header(ws, 6, ["Revenue Stream", "Type", "Start Volume", "Start Price", "Monthly Volume Growth", "Monthly Price Growth"], styles)
    for idx, (name, typ) in enumerate(rows, start=7):
        ws.cell(idx, 2, name)
        _input(ws, idx, 3, typ, styles)
        _input(ws, idx, 4, 100 + (idx - 7) * 20, styles)
        _input(ws, idx, 5, 1000 - (idx - 7) * 100, styles)
        _input(ws, idx, 6, 0.01, styles)
        _input(ws, idx, 7, 0.002, styles)
        _add_list_validation(ws, f"C{idx}", "'Lists & Dates'!$H$2:$H$20", DataValidation)
    ws["B14"] = "Total Revenue"
    ws["B14"].font = styles["bold_font"]
    for c in _period_cols():
        letter = _col(c)
        formulas = []
        for r in range(7, 12):
            month_index = c - FIRST_PERIOD_COL
            formulas.append(f"($D{r}*(1+$F{r})^{month_index})*($E{r}*(1+$G{r})^{month_index})")
        _formula(ws, 14, c, "=" + "+".join(formulas), styles, output=True)


def _build_product_build(ws, styles: dict, DataValidation) -> None:
    ws["B2"] = "Product Build"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    _table_header(ws, 6, ["Stream", "COGS % Revenue", "Fulfilment Cost / Unit", "Direct FTE / Unit"], styles)
    for r in range(7, 12):
        ws.cell(r, 2, f"='Revenue Drivers'!B{r}")
        _input(ws, r, 3, 0.35, styles)
        _input(ws, r, 4, 100, styles)
        _input(ws, r, 5, 0.01, styles)
    ws["B14"] = "Total COGS"
    ws["B15"] = "Gross Profit"
    ws["B16"] = "Gross Margin"
    for c in _period_cols():
        letter = _col(c)
        revenue_col = letter
        formulas = [f"('Revenue Drivers'!{revenue_col}14*($C{r}/5))" for r in range(7, 12)]
        _formula(ws, 14, c, "=-(" + "+".join(formulas) + ")", styles, output=True)
        _formula(ws, 15, c, f"='Revenue Drivers'!{letter}14+{letter}14", styles, output=True)
        _formula(ws, 16, c, f"=IFERROR({letter}15/'Revenue Drivers'!{letter}14,0)", styles, output=True, fmt="0.0%")


def _build_headcount(ws, styles: dict) -> None:
    ws["B2"] = "Headcount"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    _table_header(ws, 6, ["Department", "Opening FTE", "Avg Salary / Month", "Hiring Every N Months", "New Hires"], styles)
    depts = ["Management", "Sales", "Operations", "Finance", "IT", "Admin"]
    for r, dept in enumerate(depts, start=7):
        ws.cell(r, 2, dept)
        _input(ws, r, 3, 2 if dept == "Management" else 4, styles)
        _input(ws, r, 4, 10000 if dept == "Management" else 4500, styles)
        _input(ws, r, 5, 6, styles)
        _input(ws, r, 6, 1, styles)
    ws["B15"] = "Total FTE"
    ws["B16"] = "Payroll Cost"
    for c in _period_cols():
        letter = _col(c)
        month_idx = c - FIRST_PERIOD_COL
        for r in range(7, 13):
            _formula(ws, r, c, f"=$C{r}+INT({month_idx}/$E{r})*$F{r}", styles)
        _formula(ws, 15, c, f"=SUM({letter}7:{letter}12)", styles, output=True)
        _formula(ws, 16, c, f"=-SUMPRODUCT({letter}7:{letter}12,$D$7:$D$12)", styles, output=True)


def _build_opex(ws, styles: dict, DataValidation) -> None:
    ws["B2"] = "Opex"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    _table_header(ws, 6, ["Cost Category", "Driver", "Monthly Fixed", "% Revenue", "Cost / FTE"], styles)
    rows = [
        ("Rent", "Fixed", 25000, 0, 0),
        ("Marketing", "% Revenue", 0, 0.03, 0),
        ("IT", "Per FTE", 5000, 0, 120),
        ("Professional Fees", "Fixed", 15000, 0, 0),
        ("Travel", "% Revenue", 0, 0.01, 0),
        ("Other SG&A", "Fixed", 10000, 0, 0),
    ]
    for r, row in enumerate(rows, start=7):
        ws.cell(r, 2, row[0])
        _input(ws, r, 3, row[1], styles)
        _input(ws, r, 4, row[2], styles)
        _input(ws, r, 5, row[3], styles)
        _input(ws, r, 6, row[4], styles)
        _add_list_validation(ws, f"C{r}", "'Lists & Dates'!$K$2:$K$8", DataValidation)
    ws["B15"] = "Total Opex excl Payroll"
    ws["B16"] = "Total Opex incl Payroll"
    for c in _period_cols():
        letter = _col(c)
        for r in range(7, 13):
            _formula(ws, r, c, f"=-($D{r}+('Revenue Drivers'!{letter}14*$E{r})+('Headcount'!{letter}15*$F{r}))", styles)
        _formula(ws, 15, c, f"=SUM({letter}7:{letter}12)", styles, output=True)
        _formula(ws, 16, c, f"={letter}15+'Headcount'!{letter}16", styles, output=True)


def _build_working_capital(ws, styles: dict) -> None:
    ws["B2"] = "Working Capital"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    rows = [("DSO", 60), ("DIO", 35), ("DPO", 55), ("Receivables", None), ("Inventory", None), ("Payables", None), ("Net Working Capital", None), ("Change in NWC", None)]
    for r, (label, value) in enumerate(rows, start=6):
        ws.cell(r, 2, label)
        if value is not None:
            _input(ws, r, 3, value, styles)
    for c in _period_cols():
        letter = _col(c)
        prev = _col(c - 1)
        _formula(ws, 9, c, f"='Revenue Drivers'!{letter}14/365*$C$6", styles)
        _formula(ws, 10, c, f"=ABS('Product Build'!{letter}14)/365*$C$7", styles)
        _formula(ws, 11, c, f"=ABS('Product Build'!{letter}14+'Opex'!{letter}16)/365*$C$8", styles)
        _formula(ws, 12, c, f"={letter}9+{letter}10-{letter}11", styles, output=True)
        _formula(ws, 13, c, f"={letter}12" if c == FIRST_PERIOD_COL else f"={letter}12-{prev}12", styles, output=True)


def _build_capex(ws, styles: dict) -> None:
    ws["B2"] = "Capex D&A"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    assumptions = [("Maintenance Capex % Revenue", 0.03), ("Growth Capex % Revenue", 0.02), ("Depreciation Life Months", 60)]
    for r, (label, value) in enumerate(assumptions, start=6):
        ws.cell(r, 2, label)
        _input(ws, r, 3, value, styles)
    rows = ["Maintenance Capex", "Growth Capex", "Total Capex", "Depreciation", "Net PPE"]
    for r, label in enumerate(rows, start=11):
        ws.cell(r, 2, label)
    for c in _period_cols():
        letter = _col(c)
        prev = _col(c - 1)
        _formula(ws, 11, c, f"=-'Revenue Drivers'!{letter}14*$C$6", styles)
        _formula(ws, 12, c, f"=-'Revenue Drivers'!{letter}14*$C$7", styles)
        _formula(ws, 13, c, f"={letter}11+{letter}12", styles, output=True)
        _formula(ws, 14, c, f"=-ABS({letter}13)/$C$8", styles)
        _formula(ws, 15, c, f"=ABS({letter}13)+{letter}14" if c == FIRST_PERIOD_COL else f"={prev}15+ABS({letter}13)+{letter}14", styles, output=True)


def _build_debt_config(ws, styles: dict, DataValidation) -> None:
    ws["B2"] = "Debt Config"
    ws["B2"].font = styles["section_font"]
    _table_header(
        ws,
        4,
        [
            "Tranche",
            "Debt Type",
            "Borrower",
            "Start Date",
            "Opening Balance",
            "Commitment",
            "Term Months",
            "Moratorium Months",
            "Interest Cap Months",
            "Margin",
            "Base Rate",
            "Amortization",
            "Bullet %",
            "Cash Sweep %",
            "PIK?",
            "Min Cash",
            "Maturity Date",
        ],
        styles,
    )
    defaults = [
        ("Senior Term Loan B", "Senior Term Loan B", "OpCo", "='Control Panel'!$C$8", 300000, 300000, 84, 12, 0, 0.035, 0.030, "Linear", 0.25, 0.15, "FALSE", 50000),
        ("Super Senior RCF", "Super Senior RCF", "OpCo", "='Control Panel'!$C$8", 0, 100000, 60, 0, 0, 0.020, 0.030, "Revolver", 1.00, 0.00, "FALSE", 50000),
        ("Mezzanine PIK", "Mezzanine PIK", "HoldCo", "='Control Panel'!$C$8", 200000, 200000, 96, 24, 24, 0.060, 0.060, "PIK", 1.00, 0.00, "TRUE", 50000),
        ("Seller Note", "Seller Note", "Seller", "='Control Panel'!$C$8", 50000, 50000, 36, 12, 0, 0.000, 0.060, "Bullet", 1.00, 0.00, "FALSE", 50000),
        ("DIP / Rescue", "DIP Financing", "OpCo", "='Control Panel'!$C$8", 0, 0, 24, 0, 0, 0.060, 0.060, "Cash Sweep", 0.00, 0.50, "FALSE", 50000),
    ]
    for r in range(5, 5 + MAX_DEBT_TRANCHES):
        values = defaults[r - 5] if r - 5 < len(defaults) else ("", "Senior Term Loan B", "", "='Control Panel'!$C$8", 0, 0, 60, 0, 0, 0.030, 0.030, "Bullet", 1.00, 0.00, "FALSE", 50000)
        for c, value in enumerate(values, start=2):
            _input(ws, r, c, value, styles)
        ws.cell(r, 18, f"=EDATE(E{r},H{r})")
        ws.cell(r, 5).number_format = "yyyy-mm-dd"
        ws.cell(r, 18).number_format = "yyyy-mm-dd"
        for col in [11, 12, 14, 15]:
            ws.cell(r, col).number_format = "0.0%"
        _add_list_validation(ws, f"C{r}", "'Lists & Dates'!$N$2:$N$80", DataValidation)
        _add_list_validation(ws, f"M{r}", "'Lists & Dates'!$Q$2:$Q$16", DataValidation)
        _add_list_validation(ws, f"P{r}", "'Lists & Dates'!$T$2:$T$3", DataValidation)


def _build_debt_schedule(ws, styles: dict) -> None:
    ws["B2"] = "Debt Schedule"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    start_row = 6
    for idx in range(MAX_DEBT_TRANCHES):
        block = start_row + idx * 16
        ws.cell(block, 2, f"='Debt Config'!B{5+idx}")
        ws.cell(block, 2).font = styles["section_font"]
        labels = [
            "Active?",
            "Month Since Start",
            "Opening Debt",
            "Drawdown",
            "Cash Interest",
            "PIK / Capitalised Interest",
            "Scheduled Amortization",
            "Bullet Repayment",
            "Cash Sweep",
            "Closing Debt",
            "Undrawn Commitment",
            "Total Cash Cost",
            "Maturity Flag",
        ]
        for offset, label in enumerate(labels, start=1):
            ws.cell(block + offset, 2, label)
        cfg = 5 + idx
        for c in _period_cols():
            letter = _col(c)
            prev = _col(c - 1)
            active = block + 1
            month = block + 2
            opening = block + 3
            draw = block + 4
            cash_interest = block + 5
            pik_interest = block + 6
            amort = block + 7
            bullet = block + 8
            sweep = block + 9
            closing = block + 10
            undrawn = block + 11
            cost = block + 12
            maturity = block + 13
            _formula(ws, active, c, f'=AND({letter}$4>=\'Debt Config\'!$E{cfg},{letter}$4<=\'Debt Config\'!$R{cfg},\'Debt Config\'!$B{cfg}<>"")', styles)
            _formula(ws, month, c, f'=IF({letter}{active},DATEDIF(\'Debt Config\'!$E{cfg},{letter}$4,"m")+1,0)', styles)
            _formula(ws, opening, c, f"=IF({letter}{active},'Debt Config'!$F{cfg},0)" if c == FIRST_PERIOD_COL else f"=IF({letter}{active},{prev}{closing},0)", styles)
            _formula(ws, draw, c, f'=IF(AND({letter}{active},{letter}{opening}=0,\'Debt Config\'!$M{cfg}="Revolver"),MIN(\'Debt Config\'!$G{cfg},{letter}{undrawn}),0)', styles)
            _formula(ws, cash_interest, c, f'=IF({letter}{active},IF(OR(\'Debt Config\'!$P{cfg}="TRUE",{letter}{month}<=\'Debt Config\'!$J{cfg}),0,{letter}{opening}*(\'Debt Config\'!$K{cfg}+\'Debt Config\'!$L{cfg})/12),0)', styles)
            _formula(ws, pik_interest, c, f'=IF({letter}{active},IF(OR(\'Debt Config\'!$P{cfg}="TRUE",{letter}{month}<=\'Debt Config\'!$J{cfg}),{letter}{opening}*(\'Debt Config\'!$K{cfg}+\'Debt Config\'!$L{cfg})/12,0),0)', styles)
            _formula(ws, amort, c, f'=IF({letter}{active},IF({letter}{month}<=\'Debt Config\'!$I{cfg},0,IF(OR(\'Debt Config\'!$M{cfg}="Bullet",\'Debt Config\'!$M{cfg}="PIK",\'Debt Config\'!$M{cfg}="Revolver"),0,IF(\'Debt Config\'!$M{cfg}="Annuity",PMT((\'Debt Config\'!$K{cfg}+\'Debt Config\'!$L{cfg})/12,MAX(1,\'Debt Config\'!$H{cfg}-{letter}{month}+1),-{letter}{opening}),{letter}{opening}*(1-\'Debt Config\'!$N{cfg})/MAX(1,\'Debt Config\'!$H{cfg}-{letter}{month}+1)))),0)', styles)
            _formula(ws, bullet, c, f'=IF({letter}{maturity},{letter}{opening}*\'Debt Config\'!$N{cfg},0)', styles)
            _formula(ws, sweep, c, f'=IF({letter}{active},MIN(MAX(0,{letter}{opening}+{letter}{draw}+{letter}{pik_interest}-{letter}{amort}-{letter}{bullet}),MAX(0,\'Financial Statements\'!{letter}20-\'Debt Config\'!$Q{cfg})*\'Debt Config\'!$O{cfg}),0)', styles)
            _formula(ws, closing, c, f"=MAX(0,{letter}{opening}+{letter}{draw}+{letter}{pik_interest}-{letter}{amort}-{letter}{bullet}-{letter}{sweep})", styles, output=True)
            _formula(ws, undrawn, c, f"=MAX(0,'Debt Config'!$G{cfg}-{letter}{closing})", styles)
            _formula(ws, cost, c, f"={letter}{cash_interest}", styles, output=True)
            _formula(ws, maturity, c, f'=AND({letter}{active},{letter}{month}>=\'Debt Config\'!$H{cfg})', styles)
    agg = start_row + MAX_DEBT_TRANCHES * 16 + 2
    summary = ["Total Opening Debt", "Total Drawdowns", "Total Cash Interest", "Total PIK Interest", "Total Amortization", "Total Bullet Repayment", "Total Cash Sweep", "Closing Debt", "Undrawn Commitments", "Total Cash Cost"]
    ws.cell(agg, 2, "Aggregate Debt Summary")
    ws.cell(agg, 2).font = styles["section_font"]
    for r, label in enumerate(summary, start=agg + 1):
        ws.cell(r, 2, label)
    for c in _period_cols():
        letter = _col(c)
        blocks = [start_row + idx * 16 for idx in range(MAX_DEBT_TRANCHES)]
        _formula(ws, agg + 1, c, "=" + "+".join(f"{letter}{b+3}" for b in blocks), styles, output=True)
        _formula(ws, agg + 2, c, "=" + "+".join(f"{letter}{b+4}" for b in blocks), styles, output=True)
        _formula(ws, agg + 3, c, "=" + "+".join(f"{letter}{b+5}" for b in blocks), styles, output=True)
        _formula(ws, agg + 4, c, "=" + "+".join(f"{letter}{b+6}" for b in blocks), styles, output=True)
        _formula(ws, agg + 5, c, "=" + "+".join(f"{letter}{b+7}" for b in blocks), styles, output=True)
        _formula(ws, agg + 6, c, "=" + "+".join(f"{letter}{b+8}" for b in blocks), styles, output=True)
        _formula(ws, agg + 7, c, "=" + "+".join(f"{letter}{b+9}" for b in blocks), styles, output=True)
        _formula(ws, agg + 8, c, "=" + "+".join(f"{letter}{b+10}" for b in blocks), styles, output=True)
        _formula(ws, agg + 9, c, "=" + "+".join(f"{letter}{b+11}" for b in blocks), styles, output=True)
        _formula(ws, agg + 10, c, "=" + "+".join(f"{letter}{b+12}" for b in blocks), styles, output=True)


def _build_financial_statements(ws, styles: dict) -> None:
    ws["B2"] = "Financial Statements"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    rows = ["Revenue", "COGS", "Gross Profit", "Payroll", "Opex", "EBITDA", "D&A", "EBIT", "Cash Interest", "PBT", "Tax", "Net Income", "Change in NWC", "Capex", "Cash Flow Before Debt", "Debt Amortization / Sweep", "Free Cash Flow", "Closing Cash", "Closing Debt", "Net Debt"]
    for r, label in enumerate(rows, start=6):
        ws.cell(r, 2, label)
    debt_agg = 6 + MAX_DEBT_TRANCHES * 16 + 2
    for c in _period_cols():
        letter = _col(c)
        prev = _col(c - 1)
        _formula(ws, 6, c, f"='Revenue Drivers'!{letter}14", styles)
        _formula(ws, 7, c, f"='Product Build'!{letter}14", styles)
        _formula(ws, 8, c, f"={letter}6+{letter}7", styles, output=True)
        _formula(ws, 9, c, f"='Headcount'!{letter}16", styles)
        _formula(ws, 10, c, f"='Opex'!{letter}15", styles)
        _formula(ws, 11, c, f"=SUM({letter}8:{letter}10)", styles, output=True)
        _formula(ws, 12, c, f"='Capex D&A'!{letter}14", styles)
        _formula(ws, 13, c, f"={letter}11+{letter}12", styles)
        _formula(ws, 14, c, f"='Debt Schedule'!{letter}{debt_agg+3}", styles)
        _formula(ws, 15, c, f"={letter}13-{letter}14", styles)
        _formula(ws, 16, c, f"=-MAX({letter}15,0)*'Control Panel'!$C$13", styles)
        _formula(ws, 17, c, f"={letter}15+{letter}16", styles, output=True)
        _formula(ws, 18, c, f"='Working Capital'!{letter}13", styles)
        _formula(ws, 19, c, f"='Capex D&A'!{letter}13", styles)
        _formula(ws, 20, c, f"={letter}11+{letter}16-{letter}18+{letter}19-{letter}14", styles, output=True)
        _formula(ws, 21, c, f"='Debt Schedule'!{letter}{debt_agg+5}+'Debt Schedule'!{letter}{debt_agg+6}+'Debt Schedule'!{letter}{debt_agg+7}", styles)
        _formula(ws, 22, c, f"={letter}20-{letter}21", styles, output=True)
        _formula(ws, 23, c, f"='Control Panel'!$C$11+{letter}22" if c == FIRST_PERIOD_COL else f"={prev}23+{letter}22", styles, output=True)
        _formula(ws, 24, c, f"='Debt Schedule'!{letter}{debt_agg+8}", styles, output=True)
        _formula(ws, 25, c, f"={letter}24-{letter}23", styles, output=True)


def _build_covenants(ws, styles: dict) -> None:
    ws["B2"] = "Covenants"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    assumptions = [("Max Net Debt / EBITDA", 3.5), ("Min ICR", 2.0), ("Min Liquidity", 50000)]
    for r, (label, value) in enumerate(assumptions, start=6):
        ws.cell(r, 2, label)
        _input(ws, r, 3, value, styles)
    rows = ["Net Debt / EBITDA", "Interest Cover", "Liquidity", "Leverage Pass?", "ICR Pass?", "Liquidity Pass?", "All Covenants Pass?"]
    for r, label in enumerate(rows, start=11):
        ws.cell(r, 2, label)
    for c in _period_cols():
        letter = _col(c)
        _formula(ws, 11, c, f"=IFERROR('Financial Statements'!{letter}25/'Financial Statements'!{letter}11,0)", styles, output=True, fmt="0.0x")
        _formula(ws, 12, c, f"=IFERROR('Financial Statements'!{letter}11/'Financial Statements'!{letter}14,99)", styles, output=True, fmt="0.0x")
        _formula(ws, 13, c, f"='Financial Statements'!{letter}23", styles, output=True)
        _formula(ws, 14, c, f"={letter}11<=$C$6", styles, output=True)
        _formula(ws, 15, c, f"={letter}12>=$C$7", styles, output=True)
        _formula(ws, 16, c, f"={letter}13>=$C$8", styles, output=True)
        _formula(ws, 17, c, f"=AND({letter}14,{letter}15,{letter}16)", styles, output=True)


def _build_outputs(ws, styles: dict) -> None:
    ws["B2"] = "Outputs"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    rows = ["Revenue", "Gross Margin", "EBITDA", "EBITDA Margin", "Free Cash Flow", "Closing Debt", "Net Debt / EBITDA", "Covenant Pass?"]
    for r, label in enumerate(rows, start=6):
        ws.cell(r, 2, label)
    for c in _period_cols():
        letter = _col(c)
        _formula(ws, 6, c, f"='Financial Statements'!{letter}6", styles, output=True)
        _formula(ws, 7, c, f"=IFERROR('Financial Statements'!{letter}8/'Financial Statements'!{letter}6,0)", styles, output=True, fmt="0.0%")
        _formula(ws, 8, c, f"='Financial Statements'!{letter}11", styles, output=True)
        _formula(ws, 9, c, f"=IFERROR({letter}8/{letter}6,0)", styles, output=True, fmt="0.0%")
        _formula(ws, 10, c, f"='Financial Statements'!{letter}22", styles, output=True)
        _formula(ws, 11, c, f"='Financial Statements'!{letter}24", styles, output=True)
        _formula(ws, 12, c, f"='Covenants'!{letter}11", styles, output=True, fmt="0.0x")
        _formula(ws, 13, c, f"='Covenants'!{letter}17", styles, output=True)


def _build_packaged_output(ws, project: dict, styles: dict) -> None:
    ws["B2"] = f"{project.get('company_name', 'Target Company')} - Packaged Output"
    ws["B2"].font = styles["title_font"]
    ws["B4"] = "Executive Summary"
    ws["B4"].font = styles["section_font"]
    _table_header(ws, 6, ["Metric", "Current Month", "FY2026", "FY2027", "FY2028", "FY2029", "FY2030"], styles)
    rows = [
        ("Revenue", "'Outputs'!D6", "'Output_Group_Annual'!C5", "'Output_Group_Annual'!D5", "'Output_Group_Annual'!E5", "'Output_Group_Annual'!F5", "'Output_Group_Annual'!G5"),
        ("EBITDA", "'Outputs'!D8", "'Output_Group_Annual'!C7", "'Output_Group_Annual'!D7", "'Output_Group_Annual'!E7", "'Output_Group_Annual'!F7", "'Output_Group_Annual'!G7"),
        ("EBITDA Margin", "'Outputs'!D9", "'Output_Group_Annual'!C8", "'Output_Group_Annual'!D8", "'Output_Group_Annual'!E8", "'Output_Group_Annual'!F8", "'Output_Group_Annual'!G8"),
        ("Free Cash Flow", "'Outputs'!D10", "'Financial Statements'!D22", "'Financial Statements'!P22", "'Financial Statements'!AB22", "'Financial Statements'!AN22", "'Financial Statements'!AZ22"),
        ("Closing Debt", "'Outputs'!D11", "'Output_Group_Annual'!C10", "'Output_Group_Annual'!D10", "'Output_Group_Annual'!E10", "'Output_Group_Annual'!F10", "'Output_Group_Annual'!G10"),
        ("Net Debt / EBITDA", "'Outputs'!D12", "'Output_Group_Annual'!C11", "'Output_Group_Annual'!D11", "'Output_Group_Annual'!E11", "'Output_Group_Annual'!F11", "'Output_Group_Annual'!G11"),
    ]
    for row, values in enumerate(rows, start=7):
        ws.cell(row, 2, values[0])
        for col_idx, ref in enumerate(values[1:], start=3):
            fmt = "0.0%" if values[0] == "EBITDA Margin" else "0.0x" if "/" in values[0] else None
            _formula(ws, row, col_idx, f"={ref}", styles, output=True, fmt=fmt)
    ws["B16"] = "Key Credit Messages"
    ws["B16"].font = styles["section_font"]
    messages = [
        "Revenue and EBITDA are driver-based and link to detailed product/service assumptions.",
        "Debt schedule supports start dates, maturities, moratorium, PIK, bullet repayment and cash sweep.",
        "Covenant outputs are linked to financial statements and debt schedule.",
        "All key outputs should be reviewed alongside the Checks sheet before external distribution.",
    ]
    for row, message in enumerate(messages, start=17):
        ws.cell(row, 2, f"{row-16}.")
        ws.cell(row, 3, message)


def _build_ic_summary(ws, project: dict, styles: dict) -> None:
    ws["B2"] = f"{project.get('company_name', 'Target Company')} - Investment Committee Summary"
    ws["B2"].font = styles["title_font"]
    ws["B4"] = "Executive Decision Page"
    ws["B4"].font = styles["section_font"]
    _table_header(ws, 6, ["Topic", "Current View", "Investment / Credit Implication", "Next Diligence Action"], styles)
    rows = [
        ("Trading trajectory", "='Packaged Output'!C7", "Revenue base and growth profile drive debt capacity.", "Validate revenue bridge by customer/product."),
        ("Profitability", "='Packaged Output'!C8", "EBITDA quality supports leverage and covenant sizing.", "Bridge EBITDA to QoE adjustments."),
        ("Cash conversion", "='Packaged Output'!C10", "FCF determines deleveraging and liquidity runway.", "Stress-test working capital and capex."),
        ("Leverage", "='Packaged Output'!C12", "Net debt / EBITDA frames sponsor and lender risk.", "Run downside covenant headroom."),
        ("Covenant status", "='Outputs'!D13", "Early breach risk drives amend-and-extend or restructuring path.", "Confirm covenant definitions in facility docs."),
        ("Model integrity", "='Checks'!D10", "External distribution depends on passing checks.", "Resolve all ERROR flags before release."),
    ]
    for row, values in enumerate(rows, start=7):
        ws.cell(row, 2, values[0])
        for col_idx, value in enumerate(values[1:], start=3):
            if isinstance(value, str) and value.startswith("="):
                _formula(ws, row, col_idx, value, styles, output=True, fmt="0.0x" if values[0] == "Leverage" else None)
            else:
                ws.cell(row, col_idx, value)

    ws["B16"] = "Banker Commentary Framework"
    ws["B16"].font = styles["section_font"]
    comments = [
        ("Base Case", "Business plan remains supportable if EBITDA growth converts to cash and covenant headroom remains above internal thresholds."),
        ("Downside Case", "Primary downside risk is lower cash conversion, delayed deleveraging and tighter liquidity headroom."),
        ("Restructuring Angle", "If debt service cannot be supported, options should be sequenced by liquidity runway, stakeholder consent and enterprise value preservation."),
    ]
    _table_header(ws, 18, ["Lens", "Draft institutional language"], styles)
    for row, values in enumerate(comments, start=19):
        ws.cell(row, 2, values[0])
        ws.cell(row, 3, values[1])


def _build_restructuring_options(ws, styles: dict) -> None:
    ws["B2"] = "Restructuring Options Paper"
    ws["B2"].font = styles["section_font"]
    _table_header(ws, 4, ["Option", "Trigger", "Liquidity Impact", "Leverage Impact", "Stakeholder Complexity", "Indicative Use Case"], styles)
    rows = [
        ("Status Quo / Self Help", "No covenant breach", "Low", "Low", "Low", "Management actions, cost take-out, WC release"),
        ("Amend & Extend", "Temporary covenant pressure", "Medium", "Medium", "Medium", "Fee + revised covenant package"),
        ("Payment Holiday / PIK Toggle", "Short-term liquidity gap", "High", "Medium", "Medium", "Preserve cash while business stabilises"),
        ("New Money Super Senior", "Liquidity runway insufficient", "High", "Medium", "High", "Bridge to disposal, M&A or operational turnaround"),
        ("Debt Buyback / Discounted Payoff", "Debt trades below par", "Medium", "High", "High", "Deleveraging with sponsor or asset-sale proceeds"),
        ("Debt-for-Equity Swap", "Unsustainable capital structure", "Medium", "Very High", "Very High", "Balance sheet reset and ownership transfer"),
        ("Accelerated M&A / Disposal", "Strategic value exceeds standalone recovery", "High", "High", "High", "Stakeholder-led exit or non-core disposals"),
    ]
    for row, values in enumerate(rows, start=5):
        for col_idx, value in enumerate(values, start=2):
            ws.cell(row, col_idx, value)

    ws["B15"] = "Option Scoring"
    ws["B15"].font = styles["section_font"]
    _table_header(ws, 17, ["Option", "Liquidity Relief", "Execution Risk", "Value Preservation", "Lender Acceptability", "Weighted Score"], styles)
    for row in range(18, 25):
        ws.cell(row, 2, f"=B{row-13}")
        _input(ws, row, 3, 3, styles)
        _input(ws, row, 4, 3, styles)
        _input(ws, row, 5, 3, styles)
        _input(ws, row, 6, 3, styles)
        _formula(ws, row, 7, f"=SUM(C{row},E{row},F{row})-D{row}", styles, output=True)


def _build_debt_capacity(ws, styles: dict) -> None:
    ws["B2"] = "Debt Capacity"
    ws["B2"].font = styles["section_font"]
    _table_header(ws, 4, ["Metric", "FY2026", "FY2027", "FY2028", "FY2029", "FY2030"], styles)
    metrics = [
        ("EBITDA", "'Output_Group_Annual'!C7", "'Output_Group_Annual'!D7", "'Output_Group_Annual'!E7", "'Output_Group_Annual'!F7", "'Output_Group_Annual'!G7"),
        ("Existing Net Debt", "'Output_Group_Annual'!C11", "'Output_Group_Annual'!D11", "'Output_Group_Annual'!E11", "'Output_Group_Annual'!F11", "'Output_Group_Annual'!G11"),
        ("Lender Leverage Threshold", "3.50", "3.25", "3.00", "2.75", "2.50"),
        ("Implied Debt Capacity", "=C5*C7", "=D5*D7", "=E5*E7", "=F5*F7", "=G5*G7"),
        ("Headroom / Shortfall", "=C8-C6", "=D8-D6", "=E8-E6", "=F8-F6", "=G8-G6"),
        ("Cash Sweep Capacity", "='Financial Statements'!D22", "='Financial Statements'!P22", "='Financial Statements'!AB22", "='Financial Statements'!AN22", "='Financial Statements'!AZ22"),
        ("Refinancing Risk Flag", '=IF(C9<0,"Shortfall","Headroom")', '=IF(D9<0,"Shortfall","Headroom")', '=IF(E9<0,"Shortfall","Headroom")', '=IF(F9<0,"Shortfall","Headroom")', '=IF(G9<0,"Shortfall","Headroom")'),
    ]
    for row, values in enumerate(metrics, start=5):
        ws.cell(row, 2, values[0])
        for col_idx, value in enumerate(values[1:], start=3):
            formula = value if str(value).startswith("=") else f"={value}" if "!" in str(value) else value
            if isinstance(formula, str) and formula.startswith("="):
                _formula(ws, row, col_idx, formula, styles, output=True, fmt="0.0x" if row == 7 else None)
            else:
                _input(ws, row, col_idx, formula, styles, fmt="0.0x" if row == 7 else None)


def _build_sensitivity_matrix(ws, styles: dict) -> None:
    ws["B2"] = "Sensitivity Matrix"
    ws["B2"].font = styles["section_font"]
    ws["B4"] = "Net Debt / EBITDA Sensitivity"
    ws["B4"].font = styles["section_font"]
    _table_header(ws, 6, ["EBITDA Case / Debt Case", "-15% Debt", "Base Debt", "+15% Debt"], styles)
    cases = [("-20% EBITDA", 0.80), ("Base EBITDA", 1.00), ("+20% EBITDA", 1.20)]
    debt_cases = [0.85, 1.00, 1.15]
    for row, (label, ebitda_factor) in enumerate(cases, start=7):
        ws.cell(row, 2, label)
        for col_idx, debt_factor in enumerate(debt_cases, start=3):
            _formula(
                ws,
                row,
                col_idx,
                f"=IFERROR(('Financial Statements'!D24*{debt_factor})/('Financial Statements'!D11*{ebitda_factor}),0)",
                styles,
                output=True,
                fmt="0.0x",
            )
    ws["B13"] = "Liquidity Runway Sensitivity"
    ws["B13"].font = styles["section_font"]
    _table_header(ws, 15, ["FCF Case / Opening Cash", "-25% Cash", "Base Cash", "+25% Cash"], styles)
    fcf_cases = [("-25% FCF", 0.75), ("Base FCF", 1.00), ("+25% FCF", 1.25)]
    cash_cases = [0.75, 1.00, 1.25]
    for row, (label, fcf_factor) in enumerate(fcf_cases, start=16):
        ws.cell(row, 2, label)
        for col_idx, cash_factor in enumerate(cash_cases, start=3):
            _formula(
                ws,
                row,
                col_idx,
                f"=('Control Panel'!$C$11*{cash_factor})+('Financial Statements'!D22*{fcf_factor})",
                styles,
                output=True,
            )


def _build_checks(ws, styles: dict) -> None:
    ws["B2"] = "Checks"
    ws["B2"].font = styles["section_font"]
    _write_period_headers(ws, 4, styles)
    rows = ["No Negative Cash", "Debt Roll Forward", "Revenue Positive", "Covenants Populated", "All Checks OK"]
    for r, label in enumerate(rows, start=6):
        ws.cell(r, 2, label)
    debt_agg = 6 + MAX_DEBT_TRANCHES * 16 + 2
    for c in _period_cols():
        letter = _col(c)
        _formula(ws, 6, c, f'=IF(\'Financial Statements\'!{letter}23>=0,"OK","ERROR")', styles, output=True)
        _formula(ws, 7, c, f'=IF(\'Debt Schedule\'!{letter}{debt_agg+8}>=0,"OK","ERROR")', styles, output=True)
        _formula(ws, 8, c, f'=IF(\'Financial Statements\'!{letter}6>0,"OK","ERROR")', styles, output=True)
        _formula(ws, 9, c, f'=IF(\'Covenants\'!{letter}17<>"","OK","ERROR")', styles, output=True)
        _formula(ws, 10, c, f'=IF(AND({letter}6="OK",{letter}7="OK",{letter}8="OK",{letter}9="OK"),"OK","ERROR")', styles, output=True)


def _build_lookup(ws, styles: dict) -> None:
    ws["B2"] = "Lookup"
    ws["B2"].font = styles["section_font"]
    _table_header(ws, 4, ["Level 0", "Level 1", "Level 2", "Model Line", "Target Sheet"], styles)
    rows = [
        ("Income Statement", "Revenue", "Trading", "Revenue", "Financial Statements"),
        ("Income Statement", "COGS", "Trading", "COGS", "Financial Statements"),
        ("Income Statement", "Opex", "Payroll", "Payroll", "Financial Statements"),
        ("Income Statement", "Opex", "SG&A", "Opex", "Financial Statements"),
        ("Balance Sheet", "Debt", "Borrowings", "Closing Debt", "Financial Statements"),
        ("Balance Sheet", "Cash", "Liquidity", "Closing Cash", "Financial Statements"),
        ("Cash Flow", "Working Capital", "NWC", "Change in NWC", "Financial Statements"),
        ("Cash Flow", "Investing", "Capex", "Capex", "Financial Statements"),
    ]
    for row, values in enumerate(rows, start=5):
        for col, value in enumerate(values, start=2):
            ws.cell(row, col, value)


def _build_mapping(ws, styles: dict) -> None:
    ws["B2"] = "Mapping"
    ws["B2"].font = styles["section_font"]
    _table_header(ws, 4, ["Output Line", "Source Sheet", "Source Row", "Formula Pattern", "Status"], styles)
    rows = [
        ("Revenue", "Revenue Drivers", "14", "Linked by monthly period", "Active"),
        ("COGS", "Product Build", "14", "Linked by monthly period", "Active"),
        ("Payroll", "Headcount", "16", "Linked by monthly period", "Active"),
        ("Opex", "Opex", "15", "Linked by monthly period", "Active"),
        ("Debt Interest", "Debt Schedule", "Aggregate cash interest", "Linked by monthly period", "Active"),
        ("Closing Debt", "Debt Schedule", "Aggregate closing debt", "Linked by monthly period", "Active"),
        ("Covenants", "Covenants", "11:17", "Linked to FS and Debt Schedule", "Active"),
    ]
    for row, values in enumerate(rows, start=5):
        for col, value in enumerate(values, start=2):
            ws.cell(row, col, value)


def _build_lists(ws, project: dict, styles: dict) -> None:
    ws["A1"] = "Lists & Fixed Dates"
    ws["A1"].font = styles["section_font"]
    currencies = ["EUR", "USD", "GBP", "AED", "CHF", "CAD", "AUD"]
    scenarios = ["Base", "Downside", "Upside", "Bank Case", "IC Case", "Restructuring Case"]
    products = ["Product", "Service", "Recurring", "Project", "Consulting", "License", "Maintenance", "Other"]
    drivers = ["Fixed", "% Revenue", "Per FTE"]
    amort = ["Bullet", "Linear", "Annuity", "Cash Sweep", "Revolver", "Borrowing Base", "PIK", "PIK Toggle", "Interest Only Then Linear", "Revenue Share", "Operational Run-Off", "Debt Sculpting", "Sculpted"]
    bools = ["TRUE", "FALSE"]
    lists = [
        (1, "Currencies", currencies),
        (4, "Scenarios", scenarios),
        (7, "Product / Service Types", products),
        (10, "Cost Drivers", drivers),
        (13, "Debt Types", debt_type_options()),
        (16, "Amortization Types", amort),
        (19, "Boolean", bools),
    ]
    for col, header, values in lists:
        ws.cell(1, col, header)
        ws.cell(1, col).fill = styles["header_fill"]
        ws.cell(1, col).font = styles["header_font"]
        for row, value in enumerate(values, start=2):
            ws.cell(row, col, value)
    ws["V1"] = "Monthly Dates"
    ws["V1"].fill = styles["header_fill"]
    ws["V1"].font = styles["header_font"]
    for idx in range(PERIODS):
        row = idx + 2
        ws.cell(row, 22, f"=EOMONTH('Control Panel'!$C$8,{idx})")
        ws.cell(row, 23, f"=YEAR(V{row})")
        ws.cell(row, 24, f'="Q"&ROUNDUP(MONTH(V{row})/3,0)')
        ws.cell(row, 22).number_format = "mmm-yy"


def _write_period_headers(ws, row: int, styles: dict, source_sheet: str = "'Lists & Dates'") -> None:
    ws.cell(row, 2, "Metric")
    ws.cell(row, 2).fill = styles["header_fill"]
    ws.cell(row, 2).font = styles["header_font"]
    for idx, c in enumerate(_period_cols(), start=2):
        cell = ws.cell(row, c)
        cell.value = f"={source_sheet}!V{idx}"
        cell.number_format = "mmm-yy"
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]


def _table_header(ws, row: int, labels: list[str], styles: dict) -> None:
    for col, label in enumerate(labels, start=2):
        cell = ws.cell(row, col, label)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]


def _input(ws, row: int, col: int, value, styles: dict, fmt: str | None = None) -> None:
    cell = ws.cell(row, col, value)
    cell.fill = styles["input_fill"]
    cell.font = styles["input_font"]
    if fmt:
        cell.number_format = fmt


def _formula(ws, row: int, col: int, formula: str, styles: dict, output: bool = False, fmt: str | None = None) -> None:
    cell = ws.cell(row, col, formula)
    cell.fill = styles["output_fill"] if output else styles["formula_fill"]
    if fmt:
        cell.number_format = fmt


def _label(ws, row: int, col: int, value: str, styles: dict) -> None:
    cell = ws.cell(row, col, value)
    cell.font = styles["bold_font"]
    cell.fill = styles["section_fill"]


def _add_list_validation(ws, cell_range: str, formula_range: str, DataValidation) -> None:
    validation = DataValidation(type="list", formula1=formula_range, allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(cell_range)


def _polish_sheet(ws, styles: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D5"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    for col in range(4, FIRST_PERIOD_COL + PERIODS):
        ws.column_dimensions[_col(col)].width = 12
    for row in ws.iter_rows():
        for cell in row:
            cell.border = styles["thin_border"]
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                if not cell.number_format or cell.number_format == "General":
                    cell.number_format = '#,##0;[Red](#,##0);-'


def _period_cols() -> range:
    return range(FIRST_PERIOD_COL, FIRST_PERIOD_COL + PERIODS)


def _styles(Font, PatternFill, Border, Side, Alignment) -> dict:
    side = Side(style="thin", color="D9E1F2")
    return {
        "title_font": Font(name="Arial", size=20, bold=True, color="1F4E78"),
        "subtitle_font": Font(name="Arial", size=14, bold=True, color="404040"),
        "section_font": Font(name="Arial", size=12, bold=True, color="1F4E78"),
        "bold_font": Font(name="Arial", bold=True, color="404040"),
        "input_font": Font(name="Arial", color="0000FF"),
        "header_font": Font(name="Arial", bold=True, color="FFFFFF"),
        "header_fill": PatternFill("solid", fgColor=HEADER_FILL),
        "section_fill": PatternFill("solid", fgColor=SECTION_FILL),
        "input_fill": PatternFill("solid", fgColor=INPUT_FILL),
        "output_fill": PatternFill("solid", fgColor=OUTPUT_FILL),
        "formula_fill": PatternFill("solid", fgColor=FORMULA_FILL),
        "check_fill": PatternFill("solid", fgColor=CHECK_FILL),
        "thin_border": Border(left=side, right=side, top=side, bottom=side),
        "center": Alignment(horizontal="center"),
    }


def _col(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result
